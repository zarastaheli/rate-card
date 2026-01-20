import os
import csv
import json
import uuid
import shutil
import subprocess
import re
import math
import zipfile
import urllib.request
import urllib.parse
import hashlib
import requests
from io import BytesIO
import tempfile
import threading
import time
from functools import lru_cache
from datetime import datetime, timedelta, timezone
from pathlib import Path
import logging
from flask import Flask, render_template, request, jsonify, send_file, session, after_this_request
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils.cell import range_boundaries, column_index_from_string
from werkzeug.utils import secure_filename

logging.basicConfig(level=logging.INFO)
app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'runs'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

dashboard_jobs = {}
dashboard_jobs_lock = threading.Lock()
summary_jobs = {}
summary_jobs_lock = threading.Lock()
carrier_details_jobs = {}
carrier_details_jobs_lock = threading.Lock()

# Template caching - load once at startup to avoid 25s load time per generation
_template_cache = {}
_template_cache_lock = threading.Lock()

# Preloaded parsed workbook for faster generation
_preloaded_wb_sheets = {}
_preloaded_wb_lock = threading.Lock()

# Rate tables cache - avoid re-parsing Excel on every dashboard call
_rate_tables_cache = {}
_rate_tables_cache_lock = threading.Lock()
_pricing_controls_cache = {}
_pricing_controls_cache_lock = threading.Lock()

def _get_cached_template():
    """Get the template workbook from cache, loading if needed."""
    global _template_cache
    template_path = Path('#New Template - Rate Card.xlsx')
    if not template_path.exists():
        template_path = Path('Rate Card Template.xlsx')
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found")
    
    cache_key = str(template_path)
    with _template_cache_lock:
        cached = _template_cache.get(cache_key)
        if cached and cached.get('mtime') == template_path.stat().st_mtime:
            # Return a copy of the cached BytesIO
            cached['buffer'].seek(0)
            return BytesIO(cached['buffer'].read())
        
        # Load and cache the template as bytes
        with open(template_path, 'rb') as f:
            template_bytes = f.read()
        _template_cache[cache_key] = {
            'buffer': BytesIO(template_bytes),
            'mtime': template_path.stat().st_mtime
        }
        return BytesIO(template_bytes)

def _load_workbook_with_retry(path, attempts=3, delay=0.2, **kwargs):
    """Load a workbook with retries to avoid transient read errors."""
    last_exc = None
    for attempt in range(attempts):
        try:
            return openpyxl.load_workbook(path, **kwargs)
        except (zipfile.BadZipFile, OSError, ValueError) as exc:
            last_exc = exc
            if attempt + 1 >= attempts:
                break
            time.sleep(delay * (attempt + 1))
    raise last_exc

# Thresholds for size/weight classification.
SMALL_MAX_VOLUME = 1728
MEDIUM_MAX_VOLUME = 5000
AMAZON_DAILY_MIN = 150
UNIUNI_WORKDAY_MIN = 300
DEFAULT_WORKING_DAYS_PER_YEAR = 261
WEIGHT_CLASS_BREAKS = [
    (1, '<1'),
    (5, '1-5'),
    (10, '5-10'),
    (float('inf'), '10+')
]

COUNTRY_NAME_TO_CODE = {
    'UNITED STATES': 'US',
    'UNITED STATES OF AMERICA': 'US',
    'USA': 'US',
    'U.S.': 'US',
    'U.S.A.': 'US',
    'US': 'US',
    'CANADA': 'CA',
    'CA': 'CA',
    'MEXICO': 'MX',
    'MEX': 'MX',
    'MX': 'MX',
    'UNITED KINGDOM': 'GB',
    'GREAT BRITAIN': 'GB',
    'ENGLAND': 'GB',
    'UK': 'GB',
    'GB': 'GB',
    'GERMANY': 'DE',
    'DE': 'DE',
    'FRANCE': 'FR',
    'FR': 'FR'
}

CODE_TO_COUNTRY_NAME = {
    'US': 'United States',
    'CA': 'Canada',
    'MX': 'Mexico',
    'GB': 'United Kingdom',
    'DE': 'Germany',
    'FR': 'France'
}

# Ensure runs directory exists
Path(app.config['UPLOAD_FOLDER']).mkdir(exist_ok=True)

# Standard field mappings
STANDARD_FIELDS = {
    'required': [
        'Order Number',
        'Order Date',
        'Zip',
        'Weight',
        'Shipping Carrier',
        'Shipping Service',
        'Package Height',
        'Package Width',
        'Package Length',
        'Zone'
    ],
    'optional': ['Label Cost']
}

def required_fields_for_structure(structure):
    required = ['Shipping Carrier', 'Shipping Service']
    if structure == 'zip':
        required.append('Zip')
    elif structure == 'zone':
        required.append('Zone')
    return required

# Service level normalization
SERVICE_LEVELS = [
    'UPSÂ® Ground',
    'DHL Parcel International Direct - DDU',
    'DHL SM Parcel Expedited',
    'USPS Ground Advantage',
    'UPS 2nd Day AirÂ®',
    'DHL SM Parcel Expedited Max'
]

REDO_CARRIERS = [
    "UniUni",
    "USPS Market",
    "UPS Ground",
    "UPS Ground Saver",
    "FedEx",
    "Amazon",
    "DHL"
]

REDO_FORCED_ON = [
    "USPS Market",
    "UPS Ground",
    "UPS Ground Saver"
]
MERCHANT_CARRIERS = ['USPS', 'UPS', 'Amazon', 'FedEx', 'UniUni']
DASHBOARD_CARRIERS = ['UniUni', 'USPS Market', 'UPS Ground', 'UPS Ground Saver', 'FedEx', 'Amazon']
FAST_DASHBOARD_METRICS = False
WEIGHT_BUCKETS = [i / 16 for i in range(1, 16)] + list(range(1, 21))

USPS_ZONE_CACHE = {}
USPS_ZONE_CACHE_LOCK = threading.Lock()
USPS_ZONE_CACHE_PATH = os.environ.get('USPS_ZONE_CACHE_PATH') or str(
    Path(__file__).resolve().parent / 'runs' / 'usps_zone_cache.json'
)
USPS_ZONE_CSV_PATH = os.environ.get('USPS_ZONE_CSV_PATH') or str(
    Path(__file__).resolve().parent / 'zip_code_zones_new.csv'
)

PROGRESS_PHASE_SEQUENCE = ['normalize', 'qualification', 'write_template', 'saving']
DEFAULT_PHASE_ESTIMATES = {
    'normalize': 3.0,
    'qualification': 2.0,
    'write_template': 25.0,
    'saving': 6.0
}
_PROGRESS_STATS_FILE = Path(app.config['UPLOAD_FOLDER']) / '.progress_stats.json'
_PROGRESS_STATS_LOCK = threading.Lock()
USPS_ZONE_CSV_LOADED = False

MANUAL_ZONE_TABLE_604 = """
005 5 399 4 550---551 4 780---785 6
006---009 8 400---402 3 553---558 4 786---787 5
010---045 5 403---404 4 559 3 788 6
046---047 6 405---406 3 560---566 4 789---796 5
048---089 5 407---409 4 567 5 797---799 6
090---099 5+ 410 3 570---573 4 800---810 5
100---139 5 411---418 4 574---577 5 811 6
140---173 4 420---422 4 580---581 4 812 5
174---176 5 423---424 3 582---588 5 813---816 6
177 4 425---426 4 590---591 6 820 5
178---200 5 427 3 592---593 5 821 6
201 4 430---436 3 594---599 6 822---823 5
202---207 5 437---447 4 600---602 2* 824---825 6
208 4 448---455 3 603---606 1* 826---828 5
209---212 5 456---457 4 607 2* 829---834 6
214 5 458---459 3 608 1* 835---838 7
215 4 460 2 609 1 840---847 6
216 5 461 3 610---611 2* 850 7
217 4 462 2 612 2 851---852 6
218---224 5 463 2* 613 1 853 7
225---230 4 464 1* 614---619 2 855 6
231---238 5 465---467 2 620 3 856---857 7
239---268 4 468 3 622---624 3 859---860 6
270---272 4 469 2 625---627 2 863---864 7
273 5 470---477 3 628---631 3 865 6
274 4 478---479 2 633---637 3 870---871 6
275---279 5 480---483 3 638---641 4 873---876 6
280---282 4 484 4 644---649 4 877 5
283---285 5 485---486 3 650---652 3 878---880 6
286---289 4 487 4 653---658 4 881 5
290---292 5 488---489 3 660---662 4 882---883 6
293 4 490---491 2 664---669 4 884 5
294---295 5 492---496 3 670---671 5 885 6
296---297 4 497---500 4 672---674 4 889---891 7
298---299 5 501 3 675---679 5 893---895 7
300---303 4 502 4 680---681 4 897---898 7
304 5 503---504 3 683---689 4 900---908 7
305---307 4 505 4 690---693 5 910---928 7
308---310 5 506---507 3 700---701 5 930---938 7
311 4 508 4 703---708 5 939---941 8
312---328 5 509 3 710---714 5 942 7
329---334 6 510---516 4 716---717 4 943---951 8
335---337 5 520---526 3 718 5 952---953 7
338---342 6 527---528 2 719---729 4 954---955 8
344 5 530 3* 730---731 5 956---961 7
346---347 5 531---532 2* 733---739 5 962---966 8+
349 6 534 2* 740---741 4 96700 8
350---352 4 535 3* 743---744 4 968 8
354---359 4 537 2* 745 5 969 9+
360---361 5 538---539 3* 746 4 970---982 7
362 4 540 4 747---748 5 983 8
363---369 5 541---544 3 749 4 984---986 7
370---390 4 545 4 750---768 5 988---994 7
391---396 5 546 3 769 6 995---999 8
397 4 547---548 4 770 5
398 5 549 3* 772---779 5
"""

def strip_after_dash(value):
    if value is None:
        return ""
    text = str(value)
    for sep in (' - ', ' – ', ' — ', '-', '–', '—'):
        if sep in text:
            return text.split(sep, 1)[0].strip()
    return text

BASE_DIR = Path(__file__).resolve().parent
AMAZON_ZIP_PATH = BASE_DIR / 'Amazon Zip list  - Zip Code List.csv'
AMAZON_ZIPS = None
UNIUNI_ZIP_PATH = BASE_DIR / 'UniUni Qualified Zips.txt'
UNIUNI_ZIPS = None

def _load_amazon_zips():
    global AMAZON_ZIPS
    if AMAZON_ZIPS is not None:
        if isinstance(AMAZON_ZIPS, dict):
            return AMAZON_ZIPS
        if isinstance(AMAZON_ZIPS, (set, list)):
            zip5 = set(str(z)[:5].zfill(5) for z in AMAZON_ZIPS if str(z).strip())
            zip3 = {z[:3] for z in zip5 if len(z) >= 3}
            AMAZON_ZIPS = {'zip5': zip5, 'zip3': zip3}
            return AMAZON_ZIPS
    zip5 = set()
    zip3 = set()
    if AMAZON_ZIP_PATH.exists():
        try:
            with AMAZON_ZIP_PATH.open(newline='', encoding='utf-8', errors='ignore') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    raw = row.get('Zip Code') or ''
                    digits = re.sub(r'\D', '', str(raw))
                    if len(digits) < 5:
                        continue
                    digits = digits[:5]
                    zip5.add(digits)
                    zip3.add(digits[:3])
        except Exception:
            zip5 = set()
            zip3 = set()
    AMAZON_ZIPS = {'zip5': zip5, 'zip3': zip3}
    return AMAZON_ZIPS

def _load_uniuni_zips():
    global UNIUNI_ZIPS
    if UNIUNI_ZIPS is not None:
        if isinstance(UNIUNI_ZIPS, dict):
            return UNIUNI_ZIPS
        if isinstance(UNIUNI_ZIPS, (set, list)):
            zip5 = set(str(z)[:5].zfill(5) for z in UNIUNI_ZIPS if str(z).strip())
            zip3 = {z[:3] for z in zip5 if len(z) >= 3}
            UNIUNI_ZIPS = {'zip3': zip3, 'zip5': zip5}
            return UNIUNI_ZIPS
    zip3 = set()
    zip5 = set()
    if UNIUNI_ZIP_PATH.exists():
        try:
            with UNIUNI_ZIP_PATH.open('r', encoding='utf-8-sig', errors='ignore') as f:
                for line in f:
                    digits = re.sub(r'\D', '', line.strip())
                    if len(digits) == 3:
                        zip3.add(digits)
                    elif len(digits) == 5:
                        zip5.add(digits)
        except Exception:
            zip3 = set()
            zip5 = set()
    UNIUNI_ZIPS = {'zip3': zip3, 'zip5': zip5}
    return UNIUNI_ZIPS

def get_working_days_per_year():
    raw = os.getenv('WORKING_DAYS_PER_YEAR', str(DEFAULT_WORKING_DAYS_PER_YEAR))
    try:
        value = int(raw)
    except Exception:
        value = DEFAULT_WORKING_DAYS_PER_YEAR
    return value if value > 0 else DEFAULT_WORKING_DAYS_PER_YEAR

def is_amazon_eligible(origin_zip):
    digits = re.sub(r'\D', '', str(origin_zip or ''))
    if len(digits) < 3:
        return False
    amazon_zips = _load_amazon_zips()
    if len(digits) >= 5:
        digits = digits[:5]
        if digits in amazon_zips.get('zip5', set()):
            return True
        return digits[:3] in amazon_zips.get('zip3', set())
    return digits[:3] in amazon_zips.get('zip3', set())

def is_uniuni_zip_eligible(origin_zip):
    digits = re.sub(r'\D', '', str(origin_zip or ''))
    if not digits:
        return False
    uniuni_zips = _load_uniuni_zips()
    zip3 = uniuni_zips.get('zip3', set())
    zip5 = uniuni_zips.get('zip5', set())
    if len(digits) >= 5:
        first5 = digits[:5]
        if first5 in zip5:
            return True
        return first5[:3] in zip3
    if len(digits) >= 3:
        return digits[:3] in zip3
    return False

def _parse_annual_orders(value):
    return _parse_numeric_value(value)

def compute_eligibility(origin_zip, annual_orders, working_days_per_year=None, mapping_config=None):
    """Compute eligibility for Amazon and UniUni carriers.
    
    Eligibility requires BOTH ZIP code whitelist AND volume thresholds:
    - Amazon: ZIP in Amazon whitelist AND >= 150 orders per day (annual_orders / 365)
    - UniUni: ZIP in UniUni whitelist AND >= 300 orders per workday (annual_orders / working_days_per_year)
    
    Explicit overrides in mapping_config can bypass these requirements.
    """
    zip_eligible_amazon = is_amazon_eligible(origin_zip)
    zip_eligible_uniuni = is_uniuni_zip_eligible(origin_zip)

    # Check for explicit eligibility overrides in mapping config
    amazon_override = None
    uniuni_override = None
    if mapping_config:
        amazon_override = mapping_config.get('amazon_eligible')
        uniuni_override = (
            mapping_config.get('uniuni_eligible')
            or mapping_config.get('uniuni_qualified')
            or mapping_config.get('uniuni')
        )

    annual_orders_value = _parse_annual_orders(annual_orders)
    if annual_orders_value is None:
        amazon_volume_avg = 0
        uniuni_volume_avg = 0
    else:
        amazon_volume_avg = annual_orders_value / 365
        days = working_days_per_year or get_working_days_per_year()
        uniuni_volume_avg = annual_orders_value / days

    # Volume eligibility checks
    amazon_volume_eligible = amazon_volume_avg >= AMAZON_DAILY_MIN
    uniuni_volume_eligible = uniuni_volume_avg >= UNIUNI_WORKDAY_MIN

    # Final eligibility: BOTH ZIP whitelist AND volume threshold must pass
    # Explicit overrides take precedence if set
    if amazon_override is not None:
        if isinstance(amazon_override, bool):
            amazon_eligible_final = amazon_override
        else:
            amazon_eligible_final = str(amazon_override).strip().lower() in ('1', 'true', 'yes', 'y')
    else:
        amazon_eligible_final = zip_eligible_amazon and amazon_volume_eligible

    if uniuni_override is not None:
        if isinstance(uniuni_override, bool):
            uniuni_eligible_final = uniuni_override
        else:
            uniuni_eligible_final = str(uniuni_override).strip().lower() in ('1', 'true', 'yes', 'y')
    else:
        uniuni_eligible_final = zip_eligible_uniuni and uniuni_volume_eligible

    return {
        'zip_eligible_amazon': zip_eligible_amazon,
        'zip_eligible_uniuni': zip_eligible_uniuni,
        'amazon_volume_avg': amazon_volume_avg,
        'amazon_volume_eligible': amazon_volume_eligible,
        'uniuni_volume_avg': uniuni_volume_avg,
        'uniuni_volume_eligible': uniuni_volume_eligible,
        'amazon_eligible_final': amazon_eligible_final,
        'uniuni_eligible_final': uniuni_eligible_final,
        'working_days_per_year': working_days_per_year or get_working_days_per_year()
    }

def _parse_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {'true', '1', 'yes', 'y'}

def _parse_numeric_value(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)) and not (isinstance(value, float) and pd.isna(value)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(',', '')
    negative = False
    if text.startswith('(') and text.endswith(')'):
        negative = True
        text = text[1:-1].strip()
    lowered = text.lower()
    multiplier = 1
    if lowered.endswith('%'):
        lowered = lowered[:-1]
    if lowered.endswith('k'):
        multiplier = 1000
        lowered = lowered[:-1]
    elif lowered.endswith('m'):
        multiplier = 1_000_000
        lowered = lowered[:-1]
    elif lowered.endswith('b'):
        multiplier = 1_000_000_000
        lowered = lowered[:-1]
    match = re.search(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', lowered)
    if not match:
        return None
    try:
        number = float(match.group()) * multiplier
    except Exception:
        return None
    if negative:
        number = -number
    return number

def _parse_number(value):
    parsed = _parse_numeric_value(value)
    return parsed if parsed is not None else 0.0

def _coerce_numeric_series(series):
    if series is None:
        return None
    if isinstance(series, pd.Series) and series.empty:
        return series
    if not isinstance(series, pd.Series):
        series = pd.Series(series)
    cleaned = series
    if cleaned.dtype == object:
        text = cleaned.astype(str)
        text = text.str.replace(',', '', regex=False)
        text = text.str.replace(r'^\((.*)\)$', r'-\1', regex=True)
        text = text.str.replace(r'[^0-9.+-Ee]', '', regex=True)
        cleaned = text
    return pd.to_numeric(cleaned, errors='coerce')

def _saas_tier_name(annual_orders):
    try:
        orders = float(annual_orders)
    except Exception:
        orders = 0
    if orders <= 0:
        return ''
    if orders < 5000:
        return 'Starter Tier'
    if orders < 15000:
        return 'Growth Tier'
    if orders < 35000:
        return 'Pro Tier'
    if orders < 50000:
        return 'Scale Tier'
    return 'Enterprise Tier'

def _effective_orders(annual_orders, comment_sold, ebay, live_selling):
    adjusted = annual_orders
    if comment_sold:
        adjusted *= 0.6
    if ebay:
        adjusted *= 0.6 if live_selling else 0.95
    return adjusted

def default_included_services(services):
    if not services:
        return []
    exclude_tokens = {'PRIORITY', 'NEXT', '2ND', 'SECOND', '2_DAY', '2DAY'}
    international_tokens = {'INTERNATIONAL', 'INTL'}
    included = []
    for service in services:
        normalized = normalize_service_name(service)
        normalized_compact = normalized.replace(' ', '')
        if any(token in normalized for token in international_tokens):
            continue
        is_two_day = any(token in normalized for token in exclude_tokens) or any(
            token in normalized_compact for token in ('2DAY', '2NDDAY', 'SECONDDAY')
        )
        if is_two_day:
            if 'DHL' in normalized and 'EXPEDITED' in normalized:
                included.append(service)
                continue
            continue
        included.append(service)
    return included

def normalize_service_name(service):
    """Normalize service name for matching"""
    if not service:
        return ""
    cleaned = strip_after_dash(str(service)).replace('Â', '').replace('®', '')
    # Remove punctuation and symbols, collapse whitespace, uppercase
    normalized = re.sub(r'[^\w\s]', '', cleaned)
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized.upper().strip()

def clean_shipping_service(service):
    """Normalize shipping service for cleaned column output."""
    if not service:
        return ""
    cleaned = strip_after_dash(str(service)).replace('Â', '').replace('®', '')
    cleaned = re.sub(r'[^\w\s]', ' ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.upper().strip()

def extract_zip5(value):
    if value is None:
        return None
    zip_match = re.search(r'\d{5}', str(value))
    if zip_match:
        return int(zip_match.group())
    return None

def extract_origin_zip(value):
    if value is None:
        return None
    digits = re.sub(r'\D', '', str(value))
    if not digits:
        return None
    if len(digits) >= 5:
        return int(digits[:5])
    return int(digits)

def _zip3_from_zip(value):
    digits = re.sub(r'\D', '', str(value or ''))
    if len(digits) < 3:
        return None
    return digits[:3].zfill(3)

def _zip3s_from_range(value):
    if value is None:
        return []
    text = str(value)
    digits = re.findall(r'\d{3,5}', text)
    if not digits:
        return []
    if len(digits) >= 2 and re.search(r'(-|–|—|to|thru|through|~|---)', text, re.IGNORECASE):
        start = int(digits[0][:3])
        end = int(digits[1][:3])
        if end < start:
            start, end = end, start
        return [f"{zip3:03d}" for zip3 in range(start, end + 1)]
    return [digits[0][:3].zfill(3)]

@lru_cache(maxsize=1)
def _build_manual_zone_map_604():
    mapping = {}
    for raw_line in MANUAL_ZONE_TABLE_604.strip().splitlines():
        line = raw_line.strip()
        if not line:
            continue
        tokens = line.split()
        for idx in range(0, len(tokens) - 1, 2):
            zip_token = tokens[idx]
            zone_token = tokens[idx + 1]
            zone_match = re.search(r'\d+', zone_token)
            if not zone_match:
                continue
            zone_value = zone_match.group()
            if '---' in zip_token:
                start_raw, end_raw = zip_token.split('---', 1)
                start_digits = re.sub(r'\D', '', start_raw)
                end_digits = re.sub(r'\D', '', end_raw)
                if not start_digits or not end_digits:
                    continue
                start_zip = int(start_digits[:3])
                end_zip = int(end_digits[:3])
            else:
                digits = re.sub(r'\D', '', zip_token)
                if not digits:
                    continue
                start_zip = end_zip = int(digits[:3])
            for zip3 in range(start_zip, end_zip + 1):
                mapping[f"{zip3:03d}"] = zone_value
    return mapping

def _load_usps_zone_cache():
    if not USPS_ZONE_CACHE_PATH:
        return
    cache_path = Path(USPS_ZONE_CACHE_PATH)
    if not cache_path.exists():
        return
    try:
        with open(cache_path, 'r') as f:
            data = json.load(f)
        if isinstance(data, dict):
            USPS_ZONE_CACHE.update(data)
    except Exception:
        pass

def _save_usps_zone_cache():
    if not USPS_ZONE_CACHE_PATH:
        return
    cache_path = Path(USPS_ZONE_CACHE_PATH)
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        with open(cache_path, 'w') as f:
            json.dump(USPS_ZONE_CACHE, f)
    except Exception:
        pass

def _choose_origin_zip_column(fieldnames):
    for name in fieldnames:
        if name.lower() in {'zipcodeprefix', 'zip_code_prefix', 'zip code prefix'}:
            return name
    for name in fieldnames:
        if 'origin' in name.lower() and 'zip' in name.lower():
            return name
    return None

def _choose_dest_zip_column(fieldnames):
    candidates = {
        'destinationzip', 'destzip', 'destination zip', 'dest zip',
        'destinationzipcode', 'destzipcode', 'destination zip code',
        'dest zip code', 'zip', 'zipcode', 'zip code', 'destination'
    }
    for name in fieldnames:
        if name.lower() in candidates:
            return name
    for name in fieldnames:
        lowered = name.lower()
        if 'zip' in lowered and 'prefix' not in lowered and 'origin' not in lowered:
            return name
    return None

def _choose_zone_column(fieldnames):
    for name in fieldnames:
        if 'zone' in name.lower():
            return name
    return None

def _load_usps_zone_csv_cache():
    global USPS_ZONE_CSV_LOADED
    if USPS_ZONE_CSV_LOADED:
        return
    USPS_ZONE_CSV_LOADED = True
    path = Path(USPS_ZONE_CSV_PATH)
    if not path.exists():
        return
    try:
        with open(path, newline='', encoding='utf-8') as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                return
            origin_col = _choose_origin_zip_column(reader.fieldnames) or 'ZipCodePrefix'
            dest_col = _choose_dest_zip_column(reader.fieldnames)
            zone_col = _choose_zone_column(reader.fieldnames)
            if dest_col is None or zone_col is None:
                return
            for row in reader:
                origin_raw = row.get(origin_col)
                origin_zip3 = _zip3_from_zip(origin_raw)
                if not origin_zip3:
                    continue
                dest_raw = row.get(dest_col)
                zone_raw = row.get(zone_col)
                if not dest_raw or not zone_raw:
                    continue
                zone_match = re.search(r'\d+', str(zone_raw).strip())
                if not zone_match:
                    continue
                zone_value = zone_match.group()
                for dest_zip3 in _zip3s_from_range(dest_raw):
                    USPS_ZONE_CACHE.setdefault(origin_zip3, {})[dest_zip3] = zone_value
    except Exception:
        return

def _zone_mapping_from_usps_json(data):
    mapping = {}
    if not isinstance(data, dict):
        return mapping
    for key, rows in data.items():
        if not str(key).lower().startswith('column'):
            continue
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            dest_value = None
            for candidate in (
                'DestinationZip', 'DestZip', 'DestZipCode', 'Destination ZIP',
                'Destination Zip', 'ZIP', 'Zip', 'ZipCode', 'ZIP Code', 'Destination'
            ):
                if candidate in row and row.get(candidate):
                    dest_value = row.get(candidate)
                    break
            if dest_value is None:
                for value in row.values():
                    if value and re.search(r'\d{3,5}', str(value)):
                        dest_value = value
                        break
            zone_value = None
            for candidate in ('Zone', 'ZONE', 'zone'):
                if candidate in row and row.get(candidate):
                    zone_value = row.get(candidate)
                    break
            if zone_value is None:
                for key_name, value in row.items():
                    if 'zone' in str(key_name).lower():
                        zone_value = value
                        break
            if not zone_value:
                continue
            for dest_zip3 in _zip3s_from_range(dest_value):
                zone_match = re.search(r'\d+', str(zone_value).strip())
                if zone_match:
                    mapping[dest_zip3] = zone_match.group()
    return mapping

def _fetch_usps_zone_chart_json(origin_zip3):
    if not origin_zip3:
        return {}
    shipping_date = datetime.now().strftime('%m/%d/%Y')
    params = urllib.parse.urlencode({
        'zipCode3Digit': origin_zip3,
        'shippingDate': shipping_date
    })
    url = f"https://postcalc.usps.com/DomesticZoneChart/GetZoneChart?{params}"
    headers = {
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Referer': 'https://postcalc.usps.com/domesticzonechart',
        'User-Agent': 'Mozilla/5.0',
        'X-Requested-With': 'XMLHttpRequest'
    }
    try:
        request_obj = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(request_obj, timeout=20) as response:
            content = response.read()
    except Exception:
        return {}
    try:
        data = json.loads(content.decode('utf-8', errors='ignore'))
    except Exception:
        return {}
    return _zone_mapping_from_usps_json(data)

def _fetch_usps_zone_chart(origin_zip3):
    if not origin_zip3:
        return {}
    with USPS_ZONE_CACHE_LOCK:
        _load_usps_zone_csv_cache()
        cached = USPS_ZONE_CACHE.get(origin_zip3)
        if isinstance(cached, dict) and cached:
            return cached
    if origin_zip3 == '604':
        return _build_manual_zone_map_604()
    return {}

def to_float(value):
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        return float(value)
    except Exception:
        return None

def normalize_country_code(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    raw = str(value).strip()
    if len(raw) == 2 and raw.isalpha():
        return raw.upper()
    lookup = COUNTRY_NAME_TO_CODE.get(raw.upper())
    return lookup or ""

def normalize_country_name(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    raw = str(value).strip()
    if len(raw) == 2 and raw.isalpha():
        return CODE_TO_COUNTRY_NAME.get(raw.upper(), "")
    return raw

def calculate_shipping_priority(cleaned_service):
    if not cleaned_service:
        return ""
    service = cleaned_service.upper()
    if 'GROUND' in service:
        return 'GROUND'
    if '2ND DAY' in service or '2 DAY' in service or '2DAY' in service:
        return 'AIR'
    if 'EXPEDITED' in service:
        return 'EXPEDITED'
    return 'OTHER'

def classify_package_size(volume):
    if volume is None:
        return ""
    if volume < SMALL_MAX_VOLUME:
        return 'SMALL'
    if volume < MEDIUM_MAX_VOLUME:
        return 'MEDIUM'
    return 'LARGE'

def classify_weight(weight_lbs):
    if weight_lbs is None:
        return ""
    for threshold, label in WEIGHT_CLASS_BREAKS:
        if weight_lbs < threshold:
            return label
    return ""

def infer_redo_carrier(carrier_value, service_value):
    """Infer Redo carrier bucket from carrier/service text."""
    combined = f"{carrier_value or ''} {service_value or ''}"
    text = normalize_service_name(combined)
    if not text:
        return None

    if 'USPS' in text or 'POSTAL' in text or 'GROUND ADVANTAGE' in text or 'USPS MARKET' in text:
        return 'USPS Market'
    if 'UPS GROUND SAVER' in text or 'UPS SAVER' in text:
        return 'UPS Ground Saver'
    if 'UPS' in text and 'GROUND' in text:
        return 'UPS Ground'
    if 'UNIUNI' in text or 'UNI UNI' in text:
        return 'UniUni'
    if 'FEDEX' in text:
        return 'FedEx'
    if 'AMAZON' in text:
        return 'Amazon'
    if 'DHL' in text:
        return 'DHL'
    return None

def extract_invoice_services(raw_df, mapping_config):
    mapping_value = mapping_config.get('mapping', {}).get('Shipping Service')
    normalized_cols = {
        re.sub(r'\W+', '', str(c).strip().lower()): c for c in raw_df.columns
    }
    candidates = []
    for norm, original in normalized_cols.items():
        score = 0
        if norm == 'shippingservice':
            score += 100
        if 'shipping' in norm and 'service' in norm:
            score += 60
        if mapping_value:
            mapping_norm = re.sub(r'\W+', '', str(mapping_value).strip().lower())
            if norm == mapping_norm:
                score += 80
        if score:
            candidates.append((score, original))
    if mapping_value and mapping_value in raw_df.columns:
        candidates.append((90, mapping_value))
    if not candidates:
        return []
    candidates.sort(key=lambda x: x[0], reverse=True)
    for _, service_col in candidates:
        if service_col not in raw_df.columns:
            continue
        services = raw_df[service_col].dropna().astype(str).tolist()
        if any(s.strip() for s in services):
            return services
    return []

def detect_redo_carriers(raw_df, mapping_config):
    carrier_col = mapping_config.get('mapping', {}).get('Shipping Carrier')
    service_col = mapping_config.get('mapping', {}).get('Shipping Service')
    detected = set()
    for _, row in raw_df.iterrows():
        carrier_val = row.get(carrier_col) if carrier_col in raw_df.columns else None
        service_val = row.get(service_col) if service_col in raw_df.columns else None
        inferred = infer_redo_carrier(carrier_val, service_val)
        if inferred:
            detected.add(inferred)
    ordered = [c for c in REDO_CARRIERS if c in detected]
    return ordered

def available_merchant_carriers(raw_df, mapping_config):
    carrier_col = mapping_config.get('mapping', {}).get('Shipping Carrier')
    detected = set()
    for _, row in raw_df.iterrows():
        carrier_val = row.get(carrier_col) if carrier_col in raw_df.columns else None
        normalized = normalize_merchant_carrier(carrier_val)
        if normalized:
            detected.add(normalized)
    display_map = {
        'USPS': 'USPS',
        'UPS': 'UPS',
        'FEDEX': 'FedEx',
        'AMAZON': 'Amazon',
        'UNIUNI': 'UniUni'
    }
    ordered = []
    for carrier in MERCHANT_CARRIERS:
        key = normalize_merchant_carrier(carrier)
        if key in detected:
            ordered.append(display_map.get(key, carrier))
    return ordered

def available_merchant_services(raw_df, mapping_config):
    services = extract_invoice_services(raw_df, mapping_config)
    if not services:
        return []
    canonical_map = {normalize_service_name(s): s for s in SERVICE_LEVELS}
    seen = set()
    ordered = []
    for service in services:
        if not service:
            continue
        cleaned = str(service).replace('Â', '').replace('®', '').strip()
        cleaned = re.sub(r'\s+', ' ', cleaned)
        norm = normalize_service_name(cleaned)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        ordered.append(canonical_map.get(norm, cleaned))
    return ordered

def normalize_redo_label(label):
    if not label:
        return ""
    text = re.sub(r'\([^)]*\)', '', strip_after_dash(str(label)))
    text = text.replace('Â', '').replace('®', '')
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.upper().strip()

def normalize_merchant_carrier(value):
    text = normalize_redo_label(value)
    if not text:
        return ""
    if 'USPS' in text or 'POSTAL' in text:
        return 'USPS'
    if 'UPS' in text:
        return 'UPS'
    if 'FEDEX' in text:
        return 'FEDEX'
    if 'AMAZON' in text:
        return 'AMAZON'
    if 'UNIUNI' in text or 'UNI UNI' in text:
        return 'UNIUNI'
    return ""

def _dashboard_selected_from_redo(selected_redo):
    selected = set()
    if not selected_redo:
        return selected
    if 'UniUni' in selected_redo:
        selected.add('UniUni')
    if 'USPS Market' in selected_redo:
        selected.add('USPS Market')
    if 'UPS Ground' in selected_redo:
        selected.add('UPS Ground')
    if 'UPS Ground Saver' in selected_redo:
        selected.add('UPS Ground Saver')
    if 'Amazon' in selected_redo:
        selected.add('Amazon')
    if 'FedEx' in selected_redo:
        selected.add('FedEx')
    return selected

def _redo_selection_from_dashboard(selected_dashboard):
    selected = set()
    if not selected_dashboard:
        return selected
    if 'UniUni' in selected_dashboard:
        selected.add('UNIUNI')
    if 'USPS Market' in selected_dashboard:
        selected.add('USPS MARKET')
    if 'UPS Ground' in selected_dashboard:
        selected.add('UPS GROUND')
    if 'UPS Ground Saver' in selected_dashboard:
        selected.add('UPS GROUND SAVER')
    if 'Amazon' in selected_dashboard:
        selected.add('AMAZON')
    if 'FedEx' in selected_dashboard:
        selected.add('FEDEX')
    return selected

def _read_summary_metrics(ws):
    return {
        'Est. Merchant Annual Savings': ws['C5'].value,
        'Est. Redo Deal Size': ws['C6'].value,
        'Spread Available': ws['C7'].value,
        '% Orders We Could Win': ws['C11'].value,
        '% Orders Won W/ Spread': ws['C12'].value
    }

def _normalize_cell_ref(cell_ref):
    text = str(cell_ref).replace('$', '')
    if '!' in text:
        sheet, ref = text.split('!', 1)
        return f"{sheet}!{ref.upper()}"
    return text.upper()

class FormulaEvaluator:
    def __init__(self, ws, overrides=None, data_only_wb=None):
        self.ws = ws
        self.overrides = { _normalize_cell_ref(k): v for k, v in (overrides or {}).items() }
        self.cache = {}
        self.workbook = ws.parent
        self.data_only_wb = data_only_wb
        self._stack = []
        self._stack_set = set()
        self._max_depth = 200

    def get(self, cell_ref):
        return self._eval_cell(_normalize_cell_ref(cell_ref))

    def _eval_cell(self, cell_ref):
        sheet_name = None
        ref = cell_ref
        if '!' in cell_ref:
            sheet_name, ref = cell_ref.split('!', 1)
        cache_key = cell_ref if sheet_name else ref
        if cache_key in self.cache:
            return self.cache[cache_key]
        if cache_key in self._stack_set:
            return ''
        if len(self._stack) >= self._max_depth:
            return ''
        self._stack.append(cache_key)
        self._stack_set.add(cache_key)
        if ref in self.overrides and (sheet_name is None or sheet_name == self.ws.title):
            value = self.overrides[ref]
            self.cache[cache_key] = value
            self._stack.pop()
            self._stack_set.discard(cache_key)
            return value
        target_ws = self.ws
        if sheet_name:
            target_ws = self.workbook[sheet_name]
        cell = target_ws[ref]
        value = cell.value
        if isinstance(value, str) and value.startswith('='):
            formula = value[1:]
            if self.data_only_wb and self._should_use_cached(formula):
                cached = self._get_cached_value(sheet_name or target_ws.title, ref)
                if cached is not None:
                    value = cached
                else:
                    value = ''
                self.cache[cache_key] = value
                self._stack.pop()
                self._stack_set.discard(cache_key)
                return value
            try:
                value = self._eval_formula(formula)
            except Exception:
                cached = self._get_cached_value(sheet_name or target_ws.title, ref)
                value = cached if cached is not None else ''
        if value is None:
            value = ''
        self.cache[cache_key] = value
        self._stack.pop()
        self._stack_set.discard(cache_key)
        return value

    def _eval_formula(self, formula):
        tokens = self._tokenize(formula)
        parser = _FormulaParser(tokens, self)
        return parser.parse_expression()

    def _get_cached_value(self, sheet_name, ref):
        if not self.data_only_wb:
            return None
        try:
            ws = self.data_only_wb[sheet_name]
        except Exception:
            return None
        try:
            return ws[ref].value
        except Exception:
            return None

    def _should_use_cached(self, formula):
        upper = formula.upper()
        if any(name in upper for name in ('COUNTIF', 'COUNTIFS', 'SUMIF', 'SUMIFS')):
            return True
        if re.search(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}", upper):
            return True
        return False

    def _split_sheet_ref(self, cell_ref):
        if '!' in cell_ref:
            sheet_name, ref = cell_ref.split('!', 1)
            return sheet_name, _normalize_cell_ref(ref)
        return None, _normalize_cell_ref(cell_ref)

    def _tokenize(self, formula):
        tokens = []
        i = 0
        length = len(formula)
        while i < length:
            ch = formula[i]
            if ch.isspace():
                i += 1
                continue
            if ch == '"':
                i += 1
                start = i
                while i < length and formula[i] != '"':
                    i += 1
                tokens.append(('STRING', formula[start:i]))
                i += 1
                continue
            if ch == "'":
                i += 1
                start = i
                while i < length and formula[i] != "'":
                    i += 1
                sheet_name = formula[start:i]
                i += 1
                if i < length and formula[i] == '!':
                    i += 1
                    cell_start = i
                    while i < length and (formula[i].isalnum() or formula[i] in '$'):
                        i += 1
                    cell_ref = formula[cell_start:i]
                    tokens.append(('CELL', _normalize_cell_ref(f"{sheet_name}!{cell_ref}")))
                    continue
                tokens.append(('IDENT', sheet_name.upper()))
                continue
            if ch in '(),:':
                tokens.append((ch, ch))
                i += 1
                continue
            if ch in '+-*/^':
                tokens.append(('OP', ch))
                i += 1
                continue
            if ch in '<>=':
                op = ch
                if i + 1 < length and formula[i + 1] in '=>':
                    op += formula[i + 1]
                    i += 1
                if op == '<>':
                    tokens.append(('OP', op))
                else:
                    tokens.append(('OP', op))
                i += 1
                continue
            if ch.isdigit() or ch == '.':
                start = i
                i += 1
                while i < length and (formula[i].isdigit() or formula[i] == '.'):
                    i += 1
                tokens.append(('NUMBER', formula[start:i]))
                continue
            if ch.isalpha() or ch == '$' or ch == '_':
                start = i
                i += 1
                while i < length and (formula[i].isalnum() or formula[i] in '$_'):
                    i += 1
                token = formula[start:i]
                if i < length and formula[i] == '!':
                    sheet_name = token.replace('$', '')
                    i += 1
                    cell_start = i
                    while i < length and (formula[i].isalnum() or formula[i] in '$'):
                        i += 1
                    cell_ref = formula[cell_start:i]
                    tokens.append(('CELL', _normalize_cell_ref(f"{sheet_name}!{cell_ref}")))
                elif re.fullmatch(r'\$?[A-Z]{1,3}\$?\d+', token.upper()):
                    tokens.append(('CELL', _normalize_cell_ref(token)))
                else:
                    tokens.append(('IDENT', token.upper()))
                continue
            raise ValueError(f"Unsupported token in formula: {formula}")
        return tokens

class _FormulaParser:
    def __init__(self, tokens, evaluator):
        self.tokens = tokens
        self.pos = 0
        self.evaluator = evaluator

    def _peek(self):
        if self.pos >= len(self.tokens):
            return None
        return self.tokens[self.pos]

    def _consume(self):
        tok = self._peek()
        if tok is not None:
            self.pos += 1
        return tok

    def _match(self, value):
        tok = self._peek()
        if tok and tok[0] == value:
            self.pos += 1
            return True
        return False

    def parse_expression(self):
        return self._parse_comparison()

    def _parse_comparison(self):
        left = self._parse_additive()
        tok = self._peek()
        if tok and tok[0] == 'OP' and tok[1] in ('=', '<>', '<', '>', '<=', '>='):
            op = self._consume()[1]
            right = self._parse_additive()
            return self._compare(op, left, right)
        return left

    def _parse_additive(self):
        value = self._parse_term()
        while True:
            tok = self._peek()
            if tok and tok[0] == 'OP' and tok[1] in ('+', '-'):
                op = self._consume()[1]
                rhs = self._parse_term()
                value = self._apply_op(op, value, rhs)
            else:
                break
        return value

    def _parse_term(self):
        value = self._parse_factor()
        while True:
            tok = self._peek()
            if tok and tok[0] == 'OP' and tok[1] in ('*', '/'):
                op = self._consume()[1]
                rhs = self._parse_factor()
                value = self._apply_op(op, value, rhs)
            else:
                break
        return value

    def _parse_factor(self):
        tok = self._peek()
        if tok is None:
            return ''
        if tok[0] == 'OP' and tok[1] in ('+', '-'):
            op = self._consume()[1]
            value = self._parse_factor()
            return self._apply_op(op, 0, value) if op == '-' else value
        if tok[0] == 'NUMBER':
            self._consume()
            return float(tok[1])
        if tok[0] == 'STRING':
            self._consume()
            return tok[1]
        if tok[0] == 'CELL':
            self._consume()
            if self._match(':'):
                end_tok = self._consume()
                if not end_tok:
                    raise ValueError('Invalid range')
                end_ref = None
                if end_tok[0] == 'CELL':
                    end_ref = end_tok[1]
                else:
                    end_ref = self._coerce_range_end(tok[1], end_tok)
                if not end_ref:
                    raise ValueError('Invalid range')
                return self._eval_range(tok[1], end_ref)
            return self.evaluator._eval_cell(tok[1])
        if tok[0] == 'IDENT':
            name = self._consume()[1]
            if not self._match('('):
                return name
            args = []
            if not self._match(')'):
                while True:
                    args.append(self.parse_expression())
                    if self._match(')'):
                        break
                    if not self._match(','):
                        raise ValueError('Expected comma')
            return self._eval_function(name, args)
        if tok[0] == '(':
            self._consume()
            value = self.parse_expression()
            if not self._match(')'):
                raise ValueError('Expected closing parenthesis')
            return value
        raise ValueError('Unsupported expression')

    def _coerce_range_end(self, start_cell, end_tok):
        if end_tok[0] == 'IDENT':
            token = end_tok[1]
            if re.fullmatch(r'\$?[A-Z]{1,3}', token):
                col = token.replace('$', '')
                return _normalize_cell_ref(f"{col}1048576")
        if end_tok[0] == 'NUMBER':
            try:
                row = int(float(end_tok[1]))
            except Exception:
                return None
            _, start_ref = self.evaluator._split_sheet_ref(start_cell)
            match = re.match(r'([A-Z]{1,3})', start_ref)
            if not match:
                return None
            col = match.group(1)
            return _normalize_cell_ref(f"{col}{row}")
        return None

    def _eval_range(self, start, end):
        start_sheet, start_ref = self.evaluator._split_sheet_ref(start)
        end_sheet, end_ref = self.evaluator._split_sheet_ref(end)
        sheet_name = start_sheet or end_sheet
        target_ws = self.evaluator.ws
        if sheet_name:
            target_ws = self.evaluator.workbook[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(f"{start_ref}:{end_ref}")
        values = []
        for row in target_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                coord = _normalize_cell_ref(cell.coordinate)
                if sheet_name:
                    coord = f"{sheet_name}!{coord}"
                values.append(self.evaluator._eval_cell(coord))
        return values

    def _apply_op(self, op, left, right):
        if op in ('+', '-', '*', '/'):
            left_num = self._to_number(left)
            right_num = self._to_number(right)
            if op == '+':
                return left_num + right_num
            if op == '-':
                return left_num - right_num
            if op == '*':
                return left_num * right_num
            if op == '/':
                return left_num / right_num if right_num != 0 else 0
        return 0

    def _compare(self, op, left, right):
        left_val, right_val, numeric = self._coerce_compare(left, right)
        if op == '=':
            return left_val == right_val
        if op == '<>':
            return left_val != right_val
        if op == '<':
            return left_val < right_val
        if op == '>':
            return left_val > right_val
        if op == '<=':
            return left_val <= right_val
        if op == '>=':
            return left_val >= right_val
        return False

    def _eval_function(self, name, args):
        if name == 'IF':
            condition = args[0] if args else ''
            true_val = args[1] if len(args) > 1 else ''
            false_val = args[2] if len(args) > 2 else ''
            return true_val if self._truthy(condition) else false_val
        if name == 'OR':
            return any(self._truthy(arg) for arg in args)
        if name == 'AND':
            return all(self._truthy(arg) for arg in args)
        if name == 'IFERROR':
            if not args:
                return ''
            try:
                return args[0]
            except Exception:
                return args[1] if len(args) > 1 else ''
        if name == 'SUM':
            total = 0.0
            for arg in args:
                if isinstance(arg, list):
                    for item in arg:
                        total += self._to_number(item)
                else:
                    total += self._to_number(arg)
            return total
        return ''

    def _truthy(self, value):
        if isinstance(value, bool):
            return value
        if value in ('', None):
            return False
        if isinstance(value, (int, float)):
            return value != 0
        return True

    def _to_number(self, value):
        if value is None or value == '':
            return 0.0
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(',', ''))
        except Exception:
            return 0.0

    def _coerce_compare(self, left, right):
        if self._is_number_like(left) and self._is_number_like(right):
            return self._to_number(left), self._to_number(right), True
        return str(left), str(right), False

    def _is_number_like(self, value):
        if isinstance(value, (int, float)):
            return True
        try:
            float(str(value).replace(',', ''))
            return True
        except Exception:
            return False

def _apply_redo_selection(ws, selected_dashboard):
    header_row_idx, label_col, use_col = _find_pricing_section(ws, 'Redo Carriers')
    if header_row_idx is None:
        return
    selection = _redo_selection_from_dashboard(selected_dashboard)
    stop_titles = {'MERCHANT CARRIERS', 'MERCHANT CARRIER', 'MERCHANT SERVICE LEVELS'}
    for row_idx, label_val in _iter_section_rows(ws, header_row_idx + 1, label_col, stop_titles):
        normalized = normalize_redo_label(label_val)
        target_cell = ws.cell(row_idx, use_col)
        if 'FIRST MILE' in normalized:
            target_cell.value = 'No'
            continue
        target_cell.value = 'Yes' if normalized in selection else 'No'

def _build_redo_overrides(ws, selected_dashboard):
    header_row_idx, label_col, use_col = _find_pricing_section(ws, 'Redo Carriers')
    if header_row_idx is None:
        return {}
    selection = _redo_selection_from_dashboard(selected_dashboard)
    stop_titles = {'MERCHANT CARRIERS', 'MERCHANT CARRIER', 'MERCHANT SERVICE LEVELS'}
    overrides = {}
    for row_idx, label_val in _iter_section_rows(ws, header_row_idx + 1, label_col, stop_titles):
        normalized = normalize_redo_label(label_val)
        coord = ws.cell(row_idx, use_col).coordinate
        if 'FIRST MILE' in normalized:
            overrides[coord] = 'No'
            continue
        overrides[coord] = 'Yes' if normalized in selection else 'No'
    return overrides

def _calculate_metrics_from_formulas(rate_card_path, selected_dashboard):
    wb = _load_workbook_with_retry(rate_card_path, data_only=False, read_only=True)
    data_wb = _load_workbook_with_retry(rate_card_path, data_only=True, read_only=True)
    if 'Pricing & Summary' not in wb.sheetnames:
        wb.close()
        data_wb.close()
        raise ValueError('Pricing & Summary sheet not found')
    ws = wb['Pricing & Summary']
    metrics = _calculate_metrics_from_formulas_ws(ws, selected_dashboard, data_wb)
    wb.close()
    data_wb.close()
    return metrics

def _calculate_metrics_from_formulas_ws(ws, selected_dashboard, data_only_wb=None):
    overrides = _build_redo_overrides(ws, selected_dashboard)
    evaluator = FormulaEvaluator(ws, overrides, data_only_wb=data_only_wb)
    return {
        'Est. Merchant Annual Savings': evaluator.get('C5'),
        'Est. Redo Deal Size': evaluator.get('C6'),
        'Spread Available': evaluator.get('C7'),
        '% Orders We Could Win': evaluator.get('C11'),
        '% Orders Won W/ Spread': evaluator.get('C12')
    }

RATE_TABLE_COLUMNS = {
    'UPS Ground': ('N', 'U'),
    'FedEx': ('X', 'AE'),
    'UniUni': ('AH', 'AO'),
    'Amazon': ('BC', 'BJ'),
    'USPS Market': ('BO', 'BV'),
    'UPS Ground Saver': ('CI', 'CP')
}
CARRIER_PRIORITY = [
    'UPS Ground',
    'FedEx',
    'UniUni',
    'Amazon',
    'USPS Market',
    'UPS Ground Saver'
]

@lru_cache(maxsize=4)
def _get_pricing_controls(template_path_str):
    global _pricing_controls_cache
    path = Path(template_path_str)
    mtime = path.stat().st_mtime
    
    with _pricing_controls_cache_lock:
        cached = _pricing_controls_cache.get(str(path))
        if cached and cached.get('mtime') == mtime:
            return cached['controls']
    
    wb = _load_workbook_with_retry(path, data_only=True)
    ws = wb['Pricing & Summary']
    controls = {
        'k2': ws['K2'].value,
        'g2': ws['G2'].value,
        'c2': ws['C2'].value,
        'c19': ws['C19'].value or 0,
        'c20': ws['C20'].value or 0,
        'c22': ws['C22'].value or 0,
        'c23': ws['C23'].value or 0,
        'c25': ws['C25'].value or 0,
        'c26': ws['C26'].value or 0
    }
    wb.close()
    
    with _pricing_controls_cache_lock:
        _pricing_controls_cache[str(path)] = {'controls': controls, 'mtime': mtime}
    
    return controls

def _load_rate_tables(template_path_str):
    global _rate_tables_cache
    path = Path(template_path_str)
    mtime = path.stat().st_mtime
    
    with _rate_tables_cache_lock:
        cached = _rate_tables_cache.get(str(path))
        if cached and cached.get('mtime') == mtime:
            return cached['tables']
    
    wb = _load_workbook_with_retry(path, data_only=True)
    ws = wb['Redo Rate Cards']
    tables = {}
    for carrier, (start_col, end_col) in RATE_TABLE_COLUMNS.items():
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        rates = {}
        for row in range(145, 210):
            zone_rates = {}
            for zone, col_idx in enumerate(range(start_idx, end_idx + 1), start=1):
                value = ws.cell(row, col_idx).value
                if value is None:
                    zone_rates[zone] = None
                else:
                    try:
                        zone_rates[zone] = float(value)
                    except Exception:
                        zone_rates[zone] = None
            rates[row] = zone_rates
        tables[carrier] = rates
    wb.close()
    
    with _rate_tables_cache_lock:
        _rate_tables_cache[str(path)] = {'tables': tables, 'mtime': mtime}
    
    return tables

def _rate_row_for_bucket(weight_bucket):
    if weight_bucket is None:
        return None
    if weight_bucket < 1:
        oz = int(round(weight_bucket * 16))
        if oz <= 0:
            return None
        return 144 + oz
    lbs = int(round(weight_bucket))
    if lbs <= 0:
        return None
    return 159 + lbs

def _compute_first_mile_weight(weight_oz, weight_lbs):
    weight_oz = pd.to_numeric(weight_oz, errors='coerce')
    weight_lbs = pd.to_numeric(weight_lbs, errors='coerce')
    output = pd.Series(np.nan, index=weight_oz.index)
    oz_mask = weight_oz.notna()
    if oz_mask.any():
        oz = weight_oz[oz_mask]
        output.loc[oz_mask] = np.where(
            oz < 16,
            np.ceil(oz).astype(int) / 16,
            np.ceil(oz / 16).astype(int)
        )
    lbs_mask = ~oz_mask & weight_lbs.notna()
    if lbs_mask.any():
        lbs = weight_lbs[lbs_mask]
        output.loc[lbs_mask] = np.where(
            lbs < 1,
            np.ceil(lbs * 16).astype(int) / 16,
            np.ceil(lbs).astype(int)
        )
    return output.round(4)

def _coerce_float(value):
    if value is None or value == '':
        return None
    try:
        return float(value)
    except Exception:
        return None

def _normalize_percent(value):
    number = _coerce_float(value)
    if number is None:
        return None
    if 1 < number <= 100:
        return number / 100
    return number

def _read_carrier_details_from_ws(ws):
    details = {}
    for row in range(35, 41):
        carrier = ws.cell(row, 2).value
        if not carrier:
            continue
        spread = _coerce_float(ws.cell(row, 4).value)
        orders_won_pct = _normalize_percent(ws.cell(row, 3).value)
        key = normalize_redo_label(carrier)
        details[key] = {
            'carrier': str(carrier).strip(),
            'spread': spread,
            'orders_won_pct': orders_won_pct
        }
    return details

@lru_cache(maxsize=16)
def _load_carrier_details_cached(rate_card_path_str, source_mtime):
    path = Path(rate_card_path_str)
    if not path.exists():
        return {}
    wb = _load_workbook_with_retry(path, data_only=True, read_only=True)
    if 'Pricing & Summary' not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb['Pricing & Summary']
    details = _read_carrier_details_from_ws(ws)
    wb.close()
    return details

def _load_carrier_details(rate_card_path, source_mtime=None):
    if not rate_card_path:
        return {}
    try:
        mtime = source_mtime or int(rate_card_path.stat().st_mtime)
    except Exception:
        mtime = 0
    return _load_carrier_details_cached(str(rate_card_path), mtime)

def _load_carrier_details_for_selection(job_dir, selected_dashboard, mapping_config, profile_dir=None):
    return _calculate_carrier_details_fast(job_dir, selected_dashboard, mapping_config)

def _carrier_details_cache_path(job_dir):
    return Path(job_dir) / 'carrier_details.json'

def _carrier_details_job_key(job_dir, source_mtime, selection_key):
    return f"{Path(job_dir).name}:{source_mtime}:{selection_key}"

def _read_carrier_details_cache(job_dir, source_mtime, selection_key):
    cache_path = _carrier_details_cache_path(job_dir)
    if not cache_path.exists():
        return None
    try:
        with open(cache_path, 'r') as f:
            cache = json.load(f)
        if cache.get('source_mtime') != source_mtime:
            return None
        entries = cache.get('entries', {})
        return entries.get(selection_key)
    except Exception:
        return None

def _write_carrier_details_cache(job_dir, source_mtime, selection_key, details):
    cache_path = _carrier_details_cache_path(job_dir)
    if not isinstance(details, dict):
        details = {}
    payload = {'source_mtime': source_mtime, 'updated_at': datetime.now(timezone.utc).isoformat(), 'entries': {}}
    if cache_path.exists():
        try:
            with open(cache_path, 'r') as f:
                existing = json.load(f)
            if existing.get('source_mtime') == source_mtime:
                payload = existing
        except Exception:
            payload = {'source_mtime': source_mtime, 'updated_at': datetime.now(timezone.utc).isoformat(), 'entries': {}}
    payload['source_mtime'] = source_mtime
    payload['updated_at'] = datetime.now(timezone.utc).isoformat()
    payload.setdefault('entries', {})
    payload['entries'][selection_key] = details
    with open(cache_path, 'w') as f:
        json.dump(payload, f)

def _build_carrier_details_cache(job_dir, source_mtime, selection_key, selected_dashboard, mapping_config, job_key):
    try:
        details = _calculate_carrier_details_fast(job_dir, selected_dashboard, mapping_config)
        if not isinstance(details, dict):
            details = {}
        _write_carrier_details_cache(job_dir, source_mtime, selection_key, details)
    finally:
        with carrier_details_jobs_lock:
            carrier_details_jobs.pop(job_key, None)

def _start_carrier_details_cache(job_dir, source_mtime, selection_key, selected_dashboard, mapping_config):
    cached = _read_carrier_details_cache(job_dir, source_mtime, selection_key)
    if cached is not None:
        return cached, False
    details = _calculate_carrier_details_fast(job_dir, selected_dashboard, mapping_config)
    if not isinstance(details, dict):
        details = {}
    _write_carrier_details_cache(job_dir, source_mtime, selection_key, details)
    return details, False

def _mode_or_min(series):
    series = series.dropna()
    if series.empty:
        return None
    counts = series.value_counts()
    if counts.empty:
        return None
    max_count = counts.max()
    if max_count <= 1:
        return float(series.min())
    candidates = counts[counts == max_count].index
    return float(min(candidates))

def _calculate_all_carriers_batch(job_dir, all_carriers, mapping_config):
    """Calculate metrics for all carriers in a single pass - much faster than per-carrier calls."""
    normalized_csv = Path(job_dir) / 'normalized.csv'
    if not normalized_csv.exists():
        return {}, {}
    normalized_df = pd.read_csv(normalized_csv)
    if normalized_df.empty:
        return {}, {}

    template_path = Path('#New Template - Rate Card.xlsx')
    if not template_path.exists():
        template_path = Path('Rate Card Template.xlsx')
    rate_tables = _load_rate_tables(str(template_path))
    controls = dict(_get_pricing_controls(str(template_path)))
    pct_off, dollar_off = _usps_market_discount_values(mapping_config)
    controls['c19'] = pct_off
    controls['c20'] = dollar_off

    merchant_pricing = {'excluded_carriers': [], 'included_services': []}
    pricing_file = Path(job_dir) / 'merchant_pricing.json'
    if pricing_file.exists():
        with open(pricing_file, 'r') as f:
            merchant_pricing = json.load(f)
    excluded_carriers = merchant_pricing.get('excluded_carriers', [])
    included_services = merchant_pricing.get('included_services', [])
    if not included_services:
        available_services = _unique_cleaned_services(normalized_df)
        included_services = default_included_services(available_services)

    normalized_selected = {normalize_service_name(s) for s in included_services}
    normalized_excluded = {normalize_merchant_carrier(c) for c in excluded_carriers}

    service_series = normalized_df.get('CLEANED_SHIPPING_SERVICE')
    if service_series is None:
        service_series = normalized_df.get('Shipping Service')
    carrier_series = normalized_df.get('Shipping Carrier')
    if service_series is None:
        service_series = pd.Series([""] * len(normalized_df))
    if carrier_series is None:
        carrier_series = pd.Series([""] * len(normalized_df))

    service_norm = service_series.fillna("").astype(str).apply(normalize_service_name)
    carrier_norm = carrier_series.fillna("").astype(str).apply(normalize_merchant_carrier)
    carrier_allowed = ~carrier_norm.isin(normalized_excluded)
    qualified = service_norm.isin(normalized_selected) & carrier_allowed

    weight_oz = normalized_df.get('WEIGHT_IN_OZ')
    if weight_oz is None:
        weight_oz = normalized_df.get('Weight')
    if weight_oz is None:
        weight_oz = pd.Series([np.nan] * len(normalized_df))
    weight_lbs = normalized_df.get('WEIGHT_IN_LBS')
    if weight_lbs is None:
        weight_lbs = pd.Series([np.nan] * len(normalized_df))

    weight_bucket = _compute_first_mile_weight(weight_oz, weight_lbs)
    zone_series = normalized_df.get('Zone')
    if zone_series is None:
        zone_series = normalized_df.get('ZONE')
    if zone_series is None:
        zone_series = pd.Series([np.nan] * len(normalized_df))
    zone = pd.to_numeric(zone_series, errors='coerce').astype('Int64')

    label_cost = normalized_df.get('Label Cost')
    if label_cost is None:
        label_cost = normalized_df.get('LABEL_COST')
    if label_cost is None:
        label_cost = pd.Series([np.nan] * len(normalized_df))
    label_cost = _coerce_numeric_series(label_cost)

    work_df = pd.DataFrame({
        'zone': zone,
        'weight_bucket': weight_bucket,
        'label_cost': label_cost,
        'qualified': qualified
    })
    work_df = work_df[work_df['zone'].between(1, 8)]
    work_df = work_df[work_df['weight_bucket'].isin(WEIGHT_BUCKETS)]
    if work_df.empty:
        return {}, {}

    count_all = work_df.groupby(['zone', 'weight_bucket']).size()
    qualified_df = work_df[work_df['qualified']]
    count_qualified = qualified_df.groupby(['zone', 'weight_bucket']).size()

    merchant_rate = None
    if controls['k2'] == 'USPS Market Rates':
        merchant_rate = {}
        usps_rates = rate_tables.get('USPS Market', {})
        for (zone_val, weight_val), count_val in count_all.items():
            row_idx = _rate_row_for_bucket(weight_val)
            rate = None
            if row_idx and row_idx in usps_rates:
                rate = usps_rates[row_idx].get(int(zone_val))
            if rate is not None:
                merchant_rate[(zone_val, weight_val)] = rate
    else:
        if controls['g2'] == 'Minimum Rates':
            nonzero = qualified_df[qualified_df['label_cost'] > 0]
            merchant_rate = nonzero.groupby(['zone', 'weight_bucket'])['label_cost'].min()
        else:
            merchant_rate = qualified_df.groupby(['zone', 'weight_bucket'])['label_cost'].apply(_mode_or_min)

    total_qualified = 0
    for key, count_val in count_qualified.items():
        if count_val > 0:
            total_qualified += count_val
    if total_qualified <= 0:
        return {}, {}

    annual_orders = None
    try:
        annual_orders = int(float(mapping_config.get('annual_orders'))) if mapping_config.get('annual_orders') else None
    except Exception:
        annual_orders = None
    scale_factor = 1.0
    orders_in_analysis = total_qualified
    if annual_orders and orders_in_analysis:
        scale_factor = orders_in_analysis / annual_orders

    c19 = float(controls['c19'] or 0)
    c20 = float(controls['c20'] or 0)
    usps_rates = rate_tables.get('USPS Market', {})

    avg_qualified_label_cost = (
        float(qualified_df['label_cost'].mean())
        if not qualified_df.empty and qualified_df['label_cost'].notna().any()
        else 0.0
    )
    annual_orders_value = annual_orders or orders_in_analysis

    carrier_metrics = {}
    for carrier in all_carriers:
        if carrier not in rate_tables:
            continue
        selected_carriers = [carrier]
        savings_all = 0.0
        savings_won = 0.0
        spread_all = 0.0
        spread_won = 0.0
        winable_count = 0.0
        won_count = 0.0
        usps_won_count = 0.0
        ups_won_count = 0.0

        for (zone_val, weight_val), count_val in count_all.items():
            count_q = count_qualified.get((zone_val, weight_val), 0)
            if count_q <= 0:
                continue
            merchant = merchant_rate.get((zone_val, weight_val))
            if merchant is None or (isinstance(merchant, float) and math.isnan(merchant)):
                continue
            row_idx = _rate_row_for_bucket(weight_val)
            if not row_idx:
                continue
            redo_rates = {}
            for c in selected_carriers:
                rate = rate_tables.get(c, {}).get(row_idx, {}).get(int(zone_val))
                if rate is not None:
                    redo_rates[c] = rate
            if not redo_rates:
                continue
            min_rate = min(redo_rates.values())
            winning_carrier = carrier

            redo_rate = min_rate
            usps_market_rate = usps_rates.get(row_idx, {}).get(int(zone_val)) if row_idx else None
            if winning_carrier in {'USPS Market', 'UPS Ground', 'UPS Ground Saver'}:
                rate_offered = redo_rate
            else:
                base_rate = merchant
                if winning_carrier in {'UniUni', 'Amazon'} and usps_market_rate is not None:
                    base_rate = usps_market_rate
                if base_rate is None or (isinstance(base_rate, float) and math.isnan(base_rate)):
                    continue
                if c19 > 0:
                    rate_offered = max(redo_rate, base_rate * (1 - c19))
                else:
                    rate_offered = max(redo_rate, base_rate - c20)

            savings = merchant - rate_offered
            base_savings = merchant - redo_rate
            spread = rate_offered - redo_rate

            savings_all += savings * count_val
            spread_all += spread * count_val
            if savings >= 0:
                savings_won += savings * count_val
                spread_won += spread * count_val
                won_count += count_val
                if winning_carrier == 'USPS Market':
                    usps_won_count += count_val
                if winning_carrier in {'UPS Ground', 'UPS Ground Saver'}:
                    ups_won_count += count_val
            if base_savings >= 0:
                winable_count += count_val

        if controls['c2'] == 'All Orders':
            est_savings = savings_all
            est_redo_deal = spread_all
        else:
            est_savings = savings_won / scale_factor if scale_factor else savings_won
            est_redo_deal = spread_won / scale_factor if scale_factor else spread_won

        usps_won_pct = usps_won_count / total_qualified if total_qualified else 0
        ups_won_pct = ups_won_count / total_qualified if total_qualified else 0
        selected_set = {carrier}
        if selected_set.issubset({'USPS Market'}):
            est_redo_deal = 0.20 * annual_orders_value * usps_won_pct
        elif selected_set.issubset({'UPS Ground', 'UPS Ground Saver'}):
            est_redo_deal = avg_qualified_label_cost * 0.11 * annual_orders_value * ups_won_pct

        spread_available = est_savings + est_redo_deal
        orders_winable = winable_count / total_qualified if total_qualified else 0
        orders_won = won_count / total_qualified if total_qualified else 0

        carrier_metrics[carrier] = {
            'Est. Merchant Annual Savings': est_savings,
            'Est. Redo Deal Size': est_redo_deal,
            'Spread Available': spread_available,
            '% Orders We Could Win': orders_winable,
            '% Orders Won W/ Spread': orders_won,
            'Orders Analyzed': total_qualified,
            'Average Label Cost': avg_qualified_label_cost
        }

    context = {
        'rate_tables': rate_tables,
        'controls': controls,
        'count_all': count_all,
        'count_qualified': count_qualified,
        'merchant_rate': merchant_rate,
        'total_qualified': total_qualified,
        'c19': c19,
        'c20': c20,
        'usps_rates': usps_rates,
        'avg_qualified_label_cost': avg_qualified_label_cost,
        'annual_orders_value': annual_orders_value,
        'scale_factor': scale_factor
    }
    return carrier_metrics, context

def _calculate_summary_from_context(selected_dashboard, context):
    """Calculate summary metrics for a carrier selection using pre-loaded context."""
    if not context:
        return {}
    rate_tables = context['rate_tables']
    controls = context['controls']
    count_all = context['count_all']
    count_qualified = context['count_qualified']
    merchant_rate = context['merchant_rate']
    total_qualified = context['total_qualified']
    c19 = context['c19']
    c20 = context['c20']
    usps_rates = context['usps_rates']
    avg_qualified_label_cost = context['avg_qualified_label_cost']
    annual_orders_value = context['annual_orders_value']
    scale_factor = context['scale_factor']

    selected_carriers = [c for c in selected_dashboard if c in rate_tables]
    if not selected_carriers:
        return {}

    savings_all = 0.0
    savings_won = 0.0
    spread_all = 0.0
    spread_won = 0.0
    winable_count = 0.0
    won_count = 0.0
    usps_won_count = 0.0
    ups_won_count = 0.0

    for (zone_val, weight_val), count_val in count_all.items():
        count_q = count_qualified.get((zone_val, weight_val), 0)
        if count_q <= 0:
            continue
        merchant = merchant_rate.get((zone_val, weight_val))
        if merchant is None or (isinstance(merchant, float) and math.isnan(merchant)):
            continue
        row_idx = _rate_row_for_bucket(weight_val)
        if not row_idx:
            continue
        redo_rates = {}
        for carrier in selected_carriers:
            rate = rate_tables.get(carrier, {}).get(row_idx, {}).get(int(zone_val))
            if rate is not None:
                redo_rates[carrier] = rate
        if not redo_rates:
            continue
        min_rate = min(redo_rates.values())
        winning_carrier = None
        for carrier in CARRIER_PRIORITY:
            rate = redo_rates.get(carrier)
            if rate is not None and abs(rate - min_rate) < 1e-9:
                winning_carrier = carrier
                break
        if winning_carrier is None:
            winning_carrier = min(redo_rates, key=redo_rates.get)

        redo_rate = min_rate
        usps_market_rate = usps_rates.get(row_idx, {}).get(int(zone_val)) if row_idx else None
        if winning_carrier in {'USPS Market', 'UPS Ground', 'UPS Ground Saver'}:
            rate_offered = redo_rate
        else:
            base_rate = merchant
            if winning_carrier in {'UniUni', 'Amazon'} and usps_market_rate is not None:
                base_rate = usps_market_rate
            if base_rate is None or (isinstance(base_rate, float) and math.isnan(base_rate)):
                continue
            if c19 > 0:
                rate_offered = max(redo_rate, base_rate * (1 - c19))
            else:
                rate_offered = max(redo_rate, base_rate - c20)

        savings = merchant - rate_offered
        base_savings = merchant - redo_rate
        spread = rate_offered - redo_rate

        savings_all += savings * count_val
        spread_all += spread * count_val
        if savings >= 0:
            savings_won += savings * count_val
            spread_won += spread * count_val
            won_count += count_val
            if winning_carrier == 'USPS Market':
                usps_won_count += count_val
            if winning_carrier in {'UPS Ground', 'UPS Ground Saver'}:
                ups_won_count += count_val
        if base_savings >= 0:
            winable_count += count_val

    if controls['c2'] == 'All Orders':
        est_savings = savings_all
        est_redo_deal = spread_all
    else:
        est_savings = savings_won / scale_factor if scale_factor else savings_won
        est_redo_deal = spread_won / scale_factor if scale_factor else spread_won

    usps_won_pct = usps_won_count / total_qualified if total_qualified else 0
    ups_won_pct = ups_won_count / total_qualified if total_qualified else 0
    selected_set = set(selected_dashboard or [])
    if selected_set and selected_set.issubset({'USPS Market'}):
        est_redo_deal = 0.20 * annual_orders_value * usps_won_pct
    elif selected_set and selected_set.issubset({'UPS Ground', 'UPS Ground Saver'}):
        est_redo_deal = avg_qualified_label_cost * 0.11 * annual_orders_value * ups_won_pct

    spread_available = est_savings + est_redo_deal
    orders_winable = winable_count / total_qualified if total_qualified else 0
    orders_won = won_count / total_qualified if total_qualified else 0

    return {
        'Est. Merchant Annual Savings': est_savings,
        'Est. Redo Deal Size': est_redo_deal,
        'Spread Available': spread_available,
        '% Orders We Could Win': orders_winable,
        '% Orders Won W/ Spread': orders_won,
        'Orders Analyzed': total_qualified,
        'Average Label Cost': avg_qualified_label_cost
    }

def _calculate_metrics_fast(job_dir, selected_dashboard, mapping_config):
    normalized_csv = job_dir / 'normalized.csv'
    if not normalized_csv.exists():
        return {}
    normalized_df = pd.read_csv(normalized_csv)
    if normalized_df.empty:
        return {}

    template_path = Path('#New Template - Rate Card.xlsx')
    if not template_path.exists():
        template_path = Path('Rate Card Template.xlsx')
    rate_tables = _load_rate_tables(str(template_path))
    controls = dict(_get_pricing_controls(str(template_path)))
    pct_off, dollar_off = _usps_market_discount_values(mapping_config)
    controls['c19'] = pct_off
    controls['c20'] = dollar_off

    merchant_pricing = {'excluded_carriers': [], 'included_services': []}
    pricing_file = job_dir / 'merchant_pricing.json'
    if pricing_file.exists():
        with open(pricing_file, 'r') as f:
            merchant_pricing = json.load(f)
    excluded_carriers = merchant_pricing.get('excluded_carriers', [])
    included_services = merchant_pricing.get('included_services', [])
    if not included_services:
        available_services = _unique_cleaned_services(normalized_df)
        included_services = default_included_services(available_services)

    normalized_selected = {normalize_service_name(s) for s in included_services}
    normalized_excluded = {normalize_merchant_carrier(c) for c in excluded_carriers}

    service_series = normalized_df.get('CLEANED_SHIPPING_SERVICE')
    if service_series is None:
        service_series = normalized_df.get('Shipping Service')
    carrier_series = normalized_df.get('Shipping Carrier')
    if service_series is None:
        service_series = pd.Series([""] * len(normalized_df))
    if carrier_series is None:
        carrier_series = pd.Series([""] * len(normalized_df))

    service_norm = service_series.fillna("").astype(str).apply(normalize_service_name)
    carrier_norm = carrier_series.fillna("").astype(str).apply(normalize_merchant_carrier)
    carrier_allowed = ~carrier_norm.isin(normalized_excluded)
    qualified = service_norm.isin(normalized_selected) & carrier_allowed

    weight_oz = normalized_df.get('WEIGHT_IN_OZ')
    if weight_oz is None:
        weight_oz = normalized_df.get('Weight')
    if weight_oz is None:
        weight_oz = pd.Series([np.nan] * len(normalized_df))
    weight_lbs = normalized_df.get('WEIGHT_IN_LBS')
    if weight_lbs is None:
        weight_lbs = pd.Series([np.nan] * len(normalized_df))

    weight_bucket = _compute_first_mile_weight(weight_oz, weight_lbs)
    zone_series = normalized_df.get('Zone')
    if zone_series is None:
        zone_series = normalized_df.get('ZONE')
    if zone_series is None:
        zone_series = pd.Series([np.nan] * len(normalized_df))
    zone = pd.to_numeric(zone_series, errors='coerce').astype('Int64')

    label_cost = normalized_df.get('Label Cost')
    if label_cost is None:
        label_cost = normalized_df.get('LABEL_COST')
    if label_cost is None:
        label_cost = pd.Series([np.nan] * len(normalized_df))
    label_cost = pd.to_numeric(label_cost, errors='coerce')

    work_df = pd.DataFrame({
        'zone': zone,
        'weight_bucket': weight_bucket,
        'label_cost': label_cost,
        'qualified': qualified
    })
    work_df = work_df[work_df['zone'].between(1, 8)]
    work_df = work_df[work_df['weight_bucket'].isin(WEIGHT_BUCKETS)]
    if work_df.empty:
        return {}

    count_all = work_df.groupby(['zone', 'weight_bucket']).size()
    qualified_df = work_df[work_df['qualified']]
    count_qualified = qualified_df.groupby(['zone', 'weight_bucket']).size()

    merchant_rate = None
    if controls['k2'] == 'USPS Market Rates':
        merchant_rate = {}
        usps_rates = rate_tables.get('USPS Market', {})
        for (zone_val, weight_val), count_val in count_all.items():
            row_idx = _rate_row_for_bucket(weight_val)
            rate = None
            if row_idx and row_idx in usps_rates:
                rate = usps_rates[row_idx].get(int(zone_val))
            if rate is not None:
                merchant_rate[(zone_val, weight_val)] = rate
    else:
        if controls['g2'] == 'Minimum Rates':
            nonzero = qualified_df[qualified_df['label_cost'] > 0]
            merchant_rate = nonzero.groupby(['zone', 'weight_bucket'])['label_cost'].min()
        else:
            merchant_rate = qualified_df.groupby(['zone', 'weight_bucket'])['label_cost'].apply(_mode_or_min)

    total_qualified = 0
    for key, count_val in count_qualified.items():
        if count_val > 0:
            total_qualified += count_val
    if total_qualified <= 0:
        return {}

    selected_carriers = [c for c in selected_dashboard if c in rate_tables]
    if not selected_carriers:
        return {}

    savings_all = 0.0
    savings_won = 0.0
    spread_all = 0.0
    spread_won = 0.0
    winable_count = 0.0
    won_count = 0.0
    usps_won_count = 0.0
    ups_won_count = 0.0

    c19 = float(controls['c19'] or 0)
    c20 = float(controls['c20'] or 0)
    usps_rates = rate_tables.get('USPS Market', {})

    for (zone_val, weight_val), count_val in count_all.items():
        count_q = count_qualified.get((zone_val, weight_val), 0)
        if count_q <= 0:
            continue
        merchant = merchant_rate.get((zone_val, weight_val))
        if merchant is None or (isinstance(merchant, float) and math.isnan(merchant)):
            continue
        row_idx = _rate_row_for_bucket(weight_val)
        if not row_idx:
            continue
        redo_rates = {}
        for carrier in selected_carriers:
            rate = rate_tables.get(carrier, {}).get(row_idx, {}).get(int(zone_val))
            if rate is not None:
                redo_rates[carrier] = rate
        if not redo_rates:
            continue
        min_rate = min(redo_rates.values())
        winning_carrier = None
        for carrier in CARRIER_PRIORITY:
            rate = redo_rates.get(carrier)
            if rate is not None and abs(rate - min_rate) < 1e-9:
                winning_carrier = carrier
                break
        if winning_carrier is None:
            winning_carrier = min(redo_rates, key=redo_rates.get)

        redo_rate = min_rate
        usps_market_rate = None
        if row_idx and row_idx in usps_rates:
            usps_market_rate = usps_rates[row_idx].get(int(zone_val))
        if winning_carrier in {'USPS Market', 'UPS Ground', 'UPS Ground Saver'}:
            rate_offered = redo_rate
        else:
            base_rate = merchant
            if winning_carrier in {'UniUni', 'Amazon'} and usps_market_rate is not None:
                base_rate = usps_market_rate
            if base_rate is None or (isinstance(base_rate, float) and math.isnan(base_rate)):
                continue
            if c19 > 0:
                rate_offered = max(redo_rate, base_rate * (1 - c19))
            else:
                rate_offered = max(redo_rate, base_rate - c20)

        savings = merchant - rate_offered
        base_savings = merchant - redo_rate
        spread = rate_offered - redo_rate

        savings_all += savings * count_val
        spread_all += spread * count_val
        if savings >= 0:
            savings_won += savings * count_val
            spread_won += spread * count_val
            won_count += count_val
            if winning_carrier == 'USPS Market':
                usps_won_count += count_val
            if winning_carrier in {'UPS Ground', 'UPS Ground Saver'}:
                ups_won_count += count_val
        if base_savings >= 0:
            winable_count += count_val

    orders_in_analysis = total_qualified
    annual_orders = None
    try:
        annual_orders = int(float(mapping_config.get('annual_orders'))) if mapping_config.get('annual_orders') else None
    except Exception:
        annual_orders = None
    scale_factor = 1.0
    if annual_orders and orders_in_analysis:
        scale_factor = orders_in_analysis / annual_orders

    if controls['c2'] == 'All Orders':
        est_savings = savings_all
        est_redo_deal = spread_all
    else:
        est_savings = savings_won / scale_factor if scale_factor else savings_won
        est_redo_deal = spread_won / scale_factor if scale_factor else spread_won

    annual_orders_value = annual_orders or orders_in_analysis
    usps_won_pct = usps_won_count / total_qualified if total_qualified else 0
    ups_won_pct = ups_won_count / total_qualified if total_qualified else 0
    avg_qualified_label_cost = (
        float(qualified_df['label_cost'].mean())
        if not qualified_df.empty and qualified_df['label_cost'].notna().any()
        else 0.0
    )
    selected_set = set(selected_dashboard or [])
    if selected_set and selected_set.issubset({'USPS Market'}):
        est_redo_deal = 0.20 * annual_orders_value * usps_won_pct
    elif selected_set and selected_set.issubset({'UPS Ground', 'UPS Ground Saver'}):
        est_redo_deal = avg_qualified_label_cost * 0.11 * annual_orders_value * ups_won_pct

    spread_available = est_savings + est_redo_deal
    orders_winable = winable_count / total_qualified if total_qualified else 0
    orders_won = won_count / total_qualified if total_qualified else 0

    return {
        'Est. Merchant Annual Savings': est_savings,
        'Est. Redo Deal Size': est_redo_deal,
        'Spread Available': spread_available,
        '% Orders We Could Win': orders_winable,
        '% Orders Won W/ Spread': orders_won
    }

def _calculate_carrier_details_fast(job_dir, selected_dashboard, mapping_config):
    normalized_csv = Path(job_dir) / 'normalized.csv'
    if not normalized_csv.exists():
        return {}
    normalized_df = pd.read_csv(normalized_csv)
    if normalized_df.empty:
        return {}

    template_path = Path('#New Template - Rate Card.xlsx')
    if not template_path.exists():
        template_path = Path('Rate Card Template.xlsx')
    rate_tables = _load_rate_tables(str(template_path))
    controls = dict(_get_pricing_controls(str(template_path)))
    pct_off, dollar_off = _usps_market_discount_values(mapping_config)
    controls['c19'] = pct_off
    controls['c20'] = dollar_off

    merchant_pricing = {'excluded_carriers': [], 'included_services': []}
    pricing_file = Path(job_dir) / 'merchant_pricing.json'
    if pricing_file.exists():
        with open(pricing_file, 'r') as f:
            merchant_pricing = json.load(f)
    excluded_carriers = merchant_pricing.get('excluded_carriers', [])
    included_services = merchant_pricing.get('included_services', [])
    if not included_services:
        available_services = _unique_cleaned_services(normalized_df)
        included_services = default_included_services(available_services)

    normalized_selected = {normalize_service_name(s) for s in included_services}
    normalized_excluded = {normalize_merchant_carrier(c) for c in excluded_carriers}

    service_series = normalized_df.get('CLEANED_SHIPPING_SERVICE')
    if service_series is None:
        service_series = normalized_df.get('Shipping Service')
    carrier_series = normalized_df.get('Shipping Carrier')
    if service_series is None:
        service_series = pd.Series([""] * len(normalized_df))
    if carrier_series is None:
        carrier_series = pd.Series([""] * len(normalized_df))

    service_norm = service_series.fillna("").astype(str).apply(normalize_service_name)
    carrier_norm = carrier_series.fillna("").astype(str).apply(normalize_merchant_carrier)
    carrier_allowed = ~carrier_norm.isin(normalized_excluded)
    qualified = service_norm.isin(normalized_selected) & carrier_allowed

    weight_oz = normalized_df.get('WEIGHT_IN_OZ')
    if weight_oz is None:
        weight_oz = normalized_df.get('Weight')
    if weight_oz is None:
        weight_oz = pd.Series([np.nan] * len(normalized_df))
    weight_lbs = normalized_df.get('WEIGHT_IN_LBS')
    if weight_lbs is None:
        weight_lbs = pd.Series([np.nan] * len(normalized_df))

    weight_bucket = _compute_first_mile_weight(weight_oz, weight_lbs)
    zone_series = normalized_df.get('Zone')
    if zone_series is None:
        zone_series = normalized_df.get('ZONE')
    if zone_series is None:
        zone_series = pd.Series([np.nan] * len(normalized_df))
    zone = pd.to_numeric(zone_series, errors='coerce').astype('Int64')

    label_cost = normalized_df.get('Label Cost')
    if label_cost is None:
        label_cost = normalized_df.get('LABEL_COST')
    if label_cost is None:
        label_cost = pd.Series([np.nan] * len(normalized_df))
    label_cost = pd.to_numeric(label_cost, errors='coerce')

    work_df = pd.DataFrame({
        'zone': zone,
        'weight_bucket': weight_bucket,
        'label_cost': label_cost,
        'qualified': qualified
    })
    work_df = work_df[work_df['zone'].between(1, 8)]
    work_df = work_df[work_df['weight_bucket'].isin(WEIGHT_BUCKETS)]
    if work_df.empty:
        return {}

    count_all = work_df.groupby(['zone', 'weight_bucket']).size()
    qualified_df = work_df[work_df['qualified']]
    count_qualified = qualified_df.groupby(['zone', 'weight_bucket']).size()

    merchant_rate = None
    if controls['k2'] == 'USPS Market Rates':
        merchant_rate = {}
        usps_rates = rate_tables.get('USPS Market', {})
        for (zone_val, weight_val), count_val in count_all.items():
            row_idx = _rate_row_for_bucket(weight_val)
            rate = None
            if row_idx and row_idx in usps_rates:
                rate = usps_rates[row_idx].get(int(zone_val))
            if rate is not None:
                merchant_rate[(zone_val, weight_val)] = rate
    else:
        if controls['g2'] == 'Minimum Rates':
            nonzero = qualified_df[qualified_df['label_cost'] > 0]
            merchant_rate = nonzero.groupby(['zone', 'weight_bucket'])['label_cost'].min()
        else:
            merchant_rate = qualified_df.groupby(['zone', 'weight_bucket'])['label_cost'].apply(_mode_or_min)

    total_qualified = 0
    for key, count_val in count_qualified.items():
        if count_val > 0:
            total_qualified += count_val
    if total_qualified <= 0:
        return {}

    selected_carriers = [c for c in (selected_dashboard or []) if c in rate_tables]
    if not selected_carriers:
        return {}

    won_counts = {carrier: 0.0 for carrier in DASHBOARD_CARRIERS}
    spread_sums = {carrier: 0.0 for carrier in DASHBOARD_CARRIERS}

    c19 = float(controls['c19'] or 0)
    c20 = float(controls['c20'] or 0)
    usps_rates = rate_tables.get('USPS Market', {})

    for (zone_val, weight_val), count_val in count_all.items():
        count_q = count_qualified.get((zone_val, weight_val), 0)
        if count_q <= 0:
            continue
        merchant = merchant_rate.get((zone_val, weight_val))
        if merchant is None or (isinstance(merchant, float) and math.isnan(merchant)):
            continue
        row_idx = _rate_row_for_bucket(weight_val)
        if not row_idx:
            continue
        redo_rates = {}
        for carrier in selected_carriers:
            rate = rate_tables.get(carrier, {}).get(row_idx, {}).get(int(zone_val))
            if rate is not None:
                redo_rates[carrier] = rate
        if not redo_rates:
            continue
        min_rate = min(redo_rates.values())
        winning_carrier = None
        for carrier in CARRIER_PRIORITY:
            rate = redo_rates.get(carrier)
            if rate is not None and abs(rate - min_rate) < 1e-9:
                winning_carrier = carrier
                break
        if winning_carrier is None:
            winning_carrier = min(redo_rates, key=redo_rates.get)

        redo_rate = min_rate
        usps_market_rate = None
        if row_idx and row_idx in usps_rates:
            usps_market_rate = usps_rates[row_idx].get(int(zone_val))
        if winning_carrier in {'USPS Market', 'UPS Ground', 'UPS Ground Saver'}:
            rate_offered = redo_rate
        else:
            base_rate = merchant
            if winning_carrier in {'UniUni', 'Amazon'} and usps_market_rate is not None:
                base_rate = usps_market_rate
            if base_rate is None or (isinstance(base_rate, float) and math.isnan(base_rate)):
                continue
            if c19 > 0:
                rate_offered = max(redo_rate, base_rate * (1 - c19))
            else:
                rate_offered = max(redo_rate, base_rate - c20)

        spread = rate_offered - redo_rate
        won_counts[winning_carrier] += count_q
        spread_sums[winning_carrier] += spread * count_q

    details = {}
    for carrier in DASHBOARD_CARRIERS:
        won_count = won_counts.get(carrier, 0.0)
        orders_pct = won_count / total_qualified if total_qualified else 0.0
        spread_total = spread_sums.get(carrier, 0.0)
        spread_avg = spread_total / won_count if won_count else 0.0
        details[normalize_redo_label(carrier)] = {
            'carrier': carrier,
            'spread': spread_avg,
            'spread_total': spread_total,
            'orders_won_pct': orders_pct
        }
    return details

def _calculate_metrics(job_dir, selected_dashboard, profile_dir=None):
    mapping_file = job_dir / 'mapping.json'
    mapping_config = {}
    if mapping_file.exists():
        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)
    rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
    if not rate_card_files:
        raise FileNotFoundError('Rate card not found')
    return _calculate_metrics_from_formulas(rate_card_files[0], selected_dashboard)

def _calculate_metrics_batch(job_dir, selections, profile_dir=None):
    rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
    if not rate_card_files:
        return {}
    source_path = rate_card_files[0]
    try:
        wb = _load_workbook_with_retry(source_path, data_only=False, read_only=True)
        data_wb = _load_workbook_with_retry(source_path, data_only=True, read_only=True)
    except Exception:
        return {}
    if 'Pricing & Summary' not in wb.sheetnames:
        wb.close()
        data_wb.close()
        return {}
    ws = wb['Pricing & Summary']
    results = {}
    for key, selected_dashboard in selections.items():
        results[key] = _calculate_metrics_from_formulas_ws(ws, selected_dashboard, data_only_wb=data_wb)
    wb.close()
    data_wb.close()
    return results

def _cache_path_for_job(job_dir):
    return job_dir / 'dashboard_breakdown.json'

def _summary_cache_path(job_dir):
    return job_dir / 'dashboard_summary.json'

def _selection_cache_key(selected_dashboard):
    return '|'.join(sorted(selected_dashboard))

def _compute_source_hash(file_path):
    """Compute SHA256 hash of a file for cache invalidation."""
    if not file_path or not Path(file_path).exists():
        return None
    sha = hashlib.sha256()
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            sha.update(chunk)
    return sha.hexdigest()[:16]

def _compute_config_hash(mapping_config, redo_config):
    """Compute hash of config for cache invalidation."""
    sha = hashlib.sha256()
    sha.update(json.dumps(mapping_config or {}, sort_keys=True).encode())
    sha.update(json.dumps(redo_config or {}, sort_keys=True).encode())
    return sha.hexdigest()[:16]

def _compute_full_cache_hash(job_dir, mapping_config=None, redo_config=None):
    """Compute combined hash of rate card file and configs."""
    rate_card_files = list(Path(job_dir).glob('* - Rate Card.xlsx'))
    if not rate_card_files:
        return None
    file_hash = _compute_source_hash(rate_card_files[0])
    config_hash = _compute_config_hash(mapping_config, redo_config)
    return f"{file_hash}:{config_hash}"

_dashboard_cache_lock = threading.Lock()

def _read_dashboard_cache(job_dir):
    """Read pre-computed dashboard cache from JSON files."""
    breakdown_path = _cache_path_for_job(Path(job_dir))
    summary_path = _summary_cache_path(Path(job_dir))
    result = {'breakdown': {}, 'summary': {}, 'source_hash': None, 'ready': False}
    with _dashboard_cache_lock:
        if breakdown_path.exists():
            try:
                with open(breakdown_path, 'r') as f:
                    data = json.load(f)
                    result['breakdown'] = data.get('carriers', {})
                    result['source_hash'] = data.get('source_hash')
                    result['ready'] = True
            except json.JSONDecodeError as e:
                app.logger.error(f"Failed to parse dashboard cache: {e}")
            except Exception as e:
                app.logger.error(f"Failed to read dashboard cache: {e}")
        if summary_path.exists():
            try:
                with open(summary_path, 'r') as f:
                    result['summary'] = json.load(f)
            except json.JSONDecodeError as e:
                app.logger.error(f"Failed to parse summary cache: {e}")
            except Exception as e:
                app.logger.error(f"Failed to read summary cache: {e}")
    return result

def _write_dashboard_cache(job_dir, breakdown, summary, source_hash):
    """Write pre-computed dashboard cache to JSON files with atomic writes."""
    import tempfile
    import shutil
    
    breakdown_path = _cache_path_for_job(Path(job_dir))
    summary_path = _summary_cache_path(Path(job_dir))
    
    with _dashboard_cache_lock:
        try:
            fd, tmp_breakdown = tempfile.mkstemp(suffix='.json', dir=job_dir)
            with os.fdopen(fd, 'w') as f:
                json.dump({'carriers': breakdown, 'source_hash': source_hash}, f)
            shutil.move(tmp_breakdown, breakdown_path)
            
            fd, tmp_summary = tempfile.mkstemp(suffix='.json', dir=job_dir)
            with os.fdopen(fd, 'w') as f:
                json.dump(summary, f)
            shutil.move(tmp_summary, summary_path)
        except Exception as e:
            app.logger.error(f"Failed to write dashboard cache: {e}")

def _is_cache_valid(job_dir, current_hash):
    """Check if cache is valid based on source hash."""
    cache = _read_dashboard_cache(job_dir)
    return cache.get('source_hash') == current_hash and cache.get('ready')

def _aggregate_metrics_from_carriers(carrier_metrics, selected_carriers):
    """Aggregate overall metrics from per-carrier metrics (fast, no Excel)."""
    if not carrier_metrics or not selected_carriers:
        return {}
    
    total_savings = 0.0
    total_spread = 0.0
    total_deal_size = 0.0
    win_pcts = []
    spread_pcts = []
    
    for carrier in selected_carriers:
        cm = carrier_metrics.get(carrier, {})
        if not cm:
            continue
        savings = cm.get('Est. Merchant Annual Savings')
        spread = cm.get('Spread Available')
        deal = cm.get('Est. Redo Deal Size')
        win_pct = cm.get('% Orders We Could Win')
        spread_pct = cm.get('% Orders Won W/ Spread')
        
        if savings is not None:
            total_savings += float(savings)
        if spread is not None:
            total_spread += float(spread)
        if deal is not None:
            total_deal_size += float(deal)
        if win_pct is not None:
            win_pcts.append(float(win_pct))
        if spread_pct is not None:
            spread_pcts.append(float(spread_pct))
    
    return {
        'Est. Merchant Annual Savings': total_savings if total_savings else None,
        'Est. Redo Deal Size': total_deal_size if total_deal_size else None,
        'Spread Available': total_spread if total_spread else None,
        '% Orders We Could Win': max(win_pcts) if win_pcts else None,
        '% Orders Won W/ Spread': max(spread_pcts) if spread_pcts else None
    }

background_cache_jobs = {}
background_cache_jobs_lock = threading.Lock()
CACHE_COMPUTATION_TIMEOUT = 900  # 15 minutes for LibreOffice recalculation per carrier

POWER_AUTOMATE_URL = os.environ.get('POWER_AUTOMATE_URL', '')

def _call_power_automate(toggles=None, outputs=None, timeout=30):
    """Call Power Automate to recalculate Excel and read cell values.
    
    Args:
        toggles: List of dicts with 'address' and 'value' keys for cells to modify
        outputs: List of cell addresses to read (e.g. ["'Pricing & Summary'!C5"])
        timeout: Request timeout in seconds
    
    Returns:
        Dict mapping cell addresses to their values, or None on error
    """
    if not POWER_AUTOMATE_URL:
        app.logger.warning("POWER_AUTOMATE_URL not configured")
        return None
    
    payload = {
        'toggles': toggles or [],
        'outputs': outputs or []
    }
    
    try:
        app.logger.info(f"Calling Power Automate with {len(outputs or [])} outputs")
        response = requests.post(
            POWER_AUTOMATE_URL,
            json=payload,
            headers={'Content-Type': 'application/json'},
            timeout=timeout
        )
        
        if not response.ok:
            app.logger.error(f"Power Automate returned {response.status_code}: {response.text[:500]}")
            return None
        
        result = response.json()
        app.logger.info(f"Power Automate returned: {result}")
        return result
    except requests.Timeout:
        app.logger.error("Power Automate request timed out")
        return None
    except Exception as e:
        app.logger.error(f"Power Automate error: {e}")
        return None

def _get_dashboard_metrics_via_power_automate(selected_carriers, all_carriers):
    """Get dashboard metrics by calling Power Automate with carrier toggles.
    
    Args:
        selected_carriers: List of carrier names to enable
        all_carriers: List of all available carrier names
    
    Returns:
        Dict with metric names as keys and values, or None on error
    """
    if not POWER_AUTOMATE_URL:
        return None
    
    selected_set = _redo_selection_from_dashboard(selected_carriers)
    
    toggles = []
    carrier_row_map = {
        'UNIUNI': 5,
        'USPS MARKET': 6, 
        'UPS GROUND': 7,
        'UPS GROUND SAVER': 8,
        'FEDEX': 9,
        'AMAZON': 10
    }
    
    for carrier_name, row in carrier_row_map.items():
        value = 'Yes' if carrier_name in selected_set else 'No'
        toggles.append({
            'address': f"'Pricing & Summary'!F{row}",
            'value': value
        })
    
    outputs = [
        "'Pricing & Summary'!C5",
        "'Pricing & Summary'!C6",
        "'Pricing & Summary'!C7",
        "'Pricing & Summary'!C11",
        "'Pricing & Summary'!C12"
    ]
    
    result = _call_power_automate(toggles=toggles, outputs=outputs, timeout=60)
    if not result:
        return None
    
    return {
        'Est. Merchant Annual Savings': result.get("'Pricing & Summary'!C5"),
        'Est. Redo Deal Size': result.get("'Pricing & Summary'!C6"),
        'Spread Available': result.get("'Pricing & Summary'!C7"),
        '% Orders We Could Win': result.get("'Pricing & Summary'!C11"),
        '% Orders Won W/ Spread': result.get("'Pricing & Summary'!C12")
    }

def _start_background_cache_job(job_dir, mapping_config, redo_config):
    """Start background job to compute dashboard cache with timeout."""
    job_key = str(job_dir)
    with background_cache_jobs_lock:
        if job_key in background_cache_jobs:
            return
        background_cache_jobs[job_key] = {'status': 'running', 'started_at': time.time()}
    
    def compute():
        try:
            _precompute_dashboard_metrics(job_dir, mapping_config, redo_config, timeout=CACHE_COMPUTATION_TIMEOUT)
            with background_cache_jobs_lock:
                if job_key in background_cache_jobs:
                    background_cache_jobs[job_key]['status'] = 'done'
        except Exception as e:
            with background_cache_jobs_lock:
                if job_key in background_cache_jobs:
                    background_cache_jobs[job_key]['status'] = 'failed'
                    background_cache_jobs[job_key]['error'] = str(e)
        finally:
            with background_cache_jobs_lock:
                background_cache_jobs.pop(job_key, None)
    
    thread = threading.Thread(target=compute, daemon=True)
    thread.start()

def _get_background_cache_status(job_dir):
    """Check status of background cache computation."""
    job_key = str(job_dir)
    with background_cache_jobs_lock:
        job = background_cache_jobs.get(job_key)
        if not job:
            return None
        elapsed = time.time() - job.get('started_at', time.time())
        if elapsed > CACHE_COMPUTATION_TIMEOUT:
            background_cache_jobs.pop(job_key, None)
            return {'status': 'timeout', 'error': 'Cache computation timed out'}
        return job
    return None

def _recalculate_excel_with_libreoffice(excel_path, timeout=60):
    """Use LibreOffice to recalculate Excel formulas and save."""
    import subprocess
    import tempfile
    import shutil
    
    excel_path = Path(excel_path)
    temp_dir = tempfile.mkdtemp()
    try:
        result = subprocess.run(
            ['soffice', '--headless', '--calc', '--convert-to', 'xlsx', 
             '--outdir', temp_dir, str(excel_path)],
            capture_output=True, text=True, timeout=timeout
        )
        if result.returncode != 0:
            app.logger.error(f"LibreOffice recalc failed: {result.stderr}")
            return False
        output_file = Path(temp_dir) / excel_path.name
        if output_file.exists():
            shutil.copy2(output_file, excel_path)
            return True
        return False
    except subprocess.TimeoutExpired:
        app.logger.error("LibreOffice recalc timed out")
        return False
    except Exception as e:
        app.logger.error(f"LibreOffice recalc error: {e}")
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

def _toggle_carriers_and_read_metrics(rate_card_path, selected_carriers, all_carriers, recalc_timeout=180):
    """Toggle carriers in Excel, recalculate, and read metrics from C5/C7/C11/C12."""
    import tempfile
    import shutil
    
    rate_card_path = Path(rate_card_path)
    temp_file = Path(tempfile.mktemp(suffix='.xlsx'))
    
    try:
        shutil.copy2(rate_card_path, temp_file)
        wb = _load_workbook_with_retry(temp_file, data_only=False)
        if 'Pricing & Summary' not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb['Pricing & Summary']
        
        header_row_idx, label_col, use_col = _find_pricing_section(ws, 'Redo Carriers')
        if header_row_idx is None:
            wb.close()
            return {}
        
        selected_set = _redo_selection_from_dashboard(selected_carriers)
        stop_titles = {'MERCHANT CARRIERS', 'MERCHANT CARRIER', 'MERCHANT SERVICE LEVELS'}
        for row_idx, label_val in _iter_section_rows(ws, header_row_idx + 1, label_col, stop_titles):
            normalized = normalize_redo_label(label_val)
            target_cell = ws.cell(row_idx, use_col)
            if 'FIRST MILE' in normalized:
                target_cell.value = 'No'
                continue
            target_cell.value = 'Yes' if normalized in selected_set else 'No'
        
        wb.save(temp_file)
        wb.close()
        
        if not _recalculate_excel_with_libreoffice(temp_file, timeout=recalc_timeout):
            return {}
        
        wb = _load_workbook_with_retry(temp_file, data_only=True)
        ws = wb['Pricing & Summary']
        metrics = {
            'Est. Merchant Annual Savings': ws['C5'].value,
            'Spread Available': ws['C7'].value,
            '% Orders We Could Win': ws['C11'].value,
            '% Orders Won W/ Spread': ws['C12'].value
        }
        wb.close()
        return metrics
    except Exception as e:
        app.logger.error(f"Toggle carriers error: {e}")
        return {}
    finally:
        if temp_file.exists():
            temp_file.unlink()

def _read_metrics_from_excel_cells(rate_card_path):
    """Read metrics directly from Excel cells after LibreOffice recalculation."""
    try:
        wb = openpyxl.load_workbook(rate_card_path, data_only=True, read_only=True)
        if 'Pricing & Summary' not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb['Pricing & Summary']
        metrics = {
            'Est. Merchant Annual Savings': ws['C5'].value,
            'Est. Redo Deal Size': ws['C6'].value,
            'Spread Available': ws['C7'].value,
            '% Orders We Could Win': ws['C11'].value,
            '% Orders Won W/ Spread': ws['C12'].value
        }
        wb.close()
        return {k: v for k, v in metrics.items() if v is not None}
    except Exception as e:
        app.logger.error(f"Error reading metrics from Excel: {e}")
        return {}

def _precompute_dashboard_metrics(job_dir, mapping_config, redo_config, timeout=None):
    """Pre-compute dashboard metrics using pure Python calculations."""
    job_dir = Path(job_dir)
    
    # Check that normalized CSV exists (required for calculations)
    normalized_csv = job_dir / 'normalized.csv'
    if not normalized_csv.exists():
        app.logger.error("No normalized.csv found for precompute")
        return False
    
    # Use rate card file for hash if it exists
    rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
    full_hash = _compute_full_cache_hash(job_dir, mapping_config, redo_config)
    selected_redo = redo_config.get('selected_carriers', [])
    selected_dashboard = _dashboard_selected_from_redo(selected_redo)
    available_carriers = list(DASHBOARD_CARRIERS)
    
    # Try Power Automate first if configured (fast + accurate)
    if POWER_AUTOMATE_URL:
        app.logger.info("Using Power Automate for dashboard metrics")
        carrier_metrics = {}
        for carrier in available_carriers:
            metrics = _get_dashboard_metrics_via_power_automate([carrier], available_carriers)
            if metrics:
                carrier_metrics[carrier] = metrics
        summary_metrics = _get_dashboard_metrics_via_power_automate(list(selected_dashboard), available_carriers)
        if carrier_metrics and summary_metrics:
            summary_by_selection = {_selection_cache_key(list(selected_dashboard)): summary_metrics}
            _write_dashboard_cache(job_dir, carrier_metrics, summary_by_selection, full_hash)
            app.logger.info("Dashboard metrics cached (Power Automate)")
            return True
    
    # Fast Python calculations - load data once, calculate all carriers
    app.logger.info("Using fast batch Python calculations for dashboard metrics")
    
    carrier_metrics, context = _calculate_all_carriers_batch(job_dir, available_carriers, mapping_config)
    for carrier, metrics in carrier_metrics.items():
        app.logger.info(f"  {carrier}: Spread={metrics.get('Spread Available')}")
    
    # Summary metrics using pre-loaded context (instant)
    summary_metrics = _calculate_summary_from_context(list(selected_dashboard), context)
    
    summary_by_selection = {}
    default_key = _selection_cache_key(list(selected_dashboard))
    summary_by_selection[default_key] = summary_metrics if summary_metrics else {}
    
    _write_dashboard_cache(job_dir, carrier_metrics, summary_by_selection, full_hash)
    app.logger.info("Dashboard metrics cached (batch Python calculations)")
    
    return True

def _summary_job_key(job_dir, source_mtime, selection_key):
    return f"{job_dir.name}:{source_mtime}:{selection_key}"

def _read_summary_cache(job_dir, source_mtime, selection_key):
    cache_path = _summary_cache_path(job_dir)
    if not cache_path.exists():
        return None
    try:
        with open(cache_path, 'r') as f:
            cache = json.load(f)
        if cache.get('source_mtime') != source_mtime:
            return None
        entries = cache.get('entries', {})
        return entries.get(selection_key)
    except Exception:
        return None

def _write_summary_cache(job_dir, source_mtime, selection_key, metrics):
    cache_path = _summary_cache_path(job_dir)
    payload = {'source_mtime': source_mtime, 'updated_at': datetime.now(timezone.utc).isoformat(), 'entries': {}}
    if cache_path.exists():
        try:
            with open(cache_path, 'r') as f:
                existing = json.load(f)
            if existing.get('source_mtime') == source_mtime:
                payload = existing
        except Exception:
            payload = {'source_mtime': source_mtime, 'updated_at': datetime.now(timezone.utc).isoformat(), 'entries': {}}
    payload['source_mtime'] = source_mtime
    payload['updated_at'] = datetime.now(timezone.utc).isoformat()
    payload.setdefault('entries', {})
    payload['entries'][selection_key] = metrics
    with open(cache_path, 'w') as f:
        json.dump(payload, f)

def _read_breakdown_cache(job_dir, source_mtime):
    cache_path = _cache_path_for_job(job_dir)
    if not cache_path.exists():
        return None, False
    try:
        with open(cache_path, 'r') as f:
            cache = json.load(f)
        if cache.get('source_mtime') != source_mtime:
            return None, False
        return cache.get('per_carrier', []), bool(cache.get('complete', False))
    except Exception:
        return None, False

def _write_breakdown_cache(job_dir, source_mtime, per_carrier, complete=True):
    cache_path = _cache_path_for_job(job_dir)
    payload = {
        'source_mtime': source_mtime,
        'updated_at': datetime.now(timezone.utc).isoformat(),
        'complete': complete,
        'per_carrier': per_carrier
    }
    with open(cache_path, 'w') as f:
        json.dump(payload, f)

def _build_breakdown_cache(job_dir, source_mtime, job_key, selected_dashboard=None, selection_key=None, available_carriers=None):
    try:
        carriers = available_carriers or list(DASHBOARD_CARRIERS)
        selections = {carrier: [carrier] for carrier in carriers}
        if selected_dashboard and selection_key:
            selections['__overall__'] = list(selected_dashboard)
        try:
            metrics_map = _calculate_metrics_batch(job_dir, selections)
        except Exception:
            metrics_map = {}

        if selected_dashboard and selection_key:
            overall_metrics = metrics_map.get('__overall__', {})
            _write_summary_cache(job_dir, source_mtime, selection_key, overall_metrics)

        per_carrier = []
        for carrier in carriers:
            metrics = metrics_map.get(carrier, {})
            per_carrier.append({'carrier': carrier, 'metrics': metrics})
            _write_breakdown_cache(job_dir, source_mtime, per_carrier, complete=False)
        _write_breakdown_cache(job_dir, source_mtime, per_carrier, complete=True)
    finally:
        with dashboard_jobs_lock:
            dashboard_jobs.pop(job_key, None)

def _build_summary_cache(job_dir, source_mtime, selection_key, selected_dashboard, job_key):
    try:
        metrics = _calculate_metrics(job_dir, selected_dashboard)
        _write_summary_cache(job_dir, source_mtime, selection_key, metrics)
    finally:
        with summary_jobs_lock:
            summary_jobs.pop(job_key, None)

def _start_breakdown_cache(job_dir, source_mtime, selected_dashboard=None, selection_key=None, available_carriers=None):
    cached, complete = _read_breakdown_cache(job_dir, source_mtime)
    if cached is not None:
        return cached, not complete
    job_key = f"{job_dir.name}:{source_mtime}"
    with dashboard_jobs_lock:
        if job_key not in dashboard_jobs:
            thread = threading.Thread(
                target=_build_breakdown_cache,
                args=(job_dir, source_mtime, job_key, selected_dashboard, selection_key, available_carriers),
                daemon=True
            )
            dashboard_jobs[job_key] = thread
            thread.start()
    return None, True

def _start_summary_cache(job_dir, source_mtime, selected_dashboard):
    selection_key = _selection_cache_key(selected_dashboard)
    cached = _read_summary_cache(job_dir, source_mtime, selection_key)
    if cached is not None:
        return cached, False
    if selected_dashboard and len(selected_dashboard) == 1:
        breakdown_cached, _ = _read_breakdown_cache(job_dir, source_mtime)
        if breakdown_cached:
            carrier = selected_dashboard[0]
            for entry in breakdown_cached:
                if entry.get('carrier') == carrier:
                    return entry.get('metrics', {}), False
    job_key = _summary_job_key(job_dir, source_mtime, selection_key)
    with summary_jobs_lock:
        if job_key not in summary_jobs:
            thread = threading.Thread(
                target=_build_summary_cache,
                args=(job_dir, source_mtime, selection_key, selected_dashboard, job_key),
                daemon=True
            )
            summary_jobs[job_key] = thread
            thread.start()
    return None, True

def _find_pricing_section(ws, section_title):
    title_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == section_title:
                title_cell = cell
                break
        if title_cell:
            break
    if not title_cell:
        return None, None, None

    header_row_idx = title_cell.row
    label_col = title_cell.column
    use_col = None

    for cell in ws[header_row_idx]:
        if cell.value and str(cell.value).strip() == 'Use in Pricing':
            use_col = cell.column
            break

    if use_col is None:
        for cell in ws[header_row_idx + 1]:
            if cell.value and str(cell.value).strip() == 'Use in Pricing':
                use_col = cell.column
                header_row_idx = header_row_idx + 1
                break

    if use_col is None:
        return None, None, None

    return header_row_idx, label_col, use_col

def _iter_section_rows(ws, start_row, label_col, stop_titles):
    row_idx = start_row
    max_row = ws.max_row
    while row_idx <= max_row:
        label_val = ws.cell(row_idx, label_col).value
        if label_val is None:
            row_idx += 1
            continue
        normalized = normalize_redo_label(label_val)
        if not normalized:
            row_idx += 1
            continue
        if normalized in stop_titles:
            break
        yield row_idx, label_val
        row_idx += 1

def _scan_section_rows(ws, section_title, stop_titles):
    header_row_idx, label_col, use_col = _find_pricing_section(ws, section_title)
    if header_row_idx is None:
        return None, None, None, []
    rows = []
    row_idx = header_row_idx + 1
    while True:
        label_val = ws.cell(row_idx, label_col).value
        normalized = normalize_redo_label(label_val)
        if not normalized or normalized in stop_titles:
            break
        rows.append((row_idx, label_val))
        row_idx += 1
    return header_row_idx, label_col, use_col, rows

def update_pricing_summary_redo_carriers(ws, selected_redo_carriers):
    """Update Use in Pricing for Redo Carriers section."""
    selected = set(selected_redo_carriers or [])
    canonical_map = {normalize_redo_label(c): c for c in REDO_CARRIERS}

    header_row_idx, label_col, use_col = _find_pricing_section(ws, 'Redo Carriers')
    if header_row_idx is None:
        return

    stop_titles = {'MERCHANT CARRIERS', 'MERCHANT CARRIER', 'MERCHANT SERVICE LEVELS'}
    seen = set()
    last_row = header_row_idx
    for row_idx, label_val in _iter_section_rows(ws, header_row_idx + 1, label_col, stop_titles):
        normalized = normalize_redo_label(label_val)
        canonical = canonical_map.get(normalized)
        if canonical:
            seen.add(canonical)

        target_cell = ws.cell(row_idx, use_col)
        target_cell.value = 'Yes' if canonical in selected else 'No'
        last_row = row_idx

    missing_carriers = [c for c in REDO_CARRIERS if c not in seen]
    for carrier in missing_carriers:
        insert_row = last_row + 1
        ws.cell(insert_row, label_col, carrier)
        ws.cell(insert_row, use_col, 'Yes' if carrier in selected else 'No')
        last_row = insert_row

    if 'First Mile' not in seen:
        insert_row = last_row + 1
        ws.cell(insert_row, label_col, 'First Mile')
        ws.cell(insert_row, use_col, 'Yes' if 'First Mile' in selected else 'No')

def update_pricing_summary_merchant_carriers(ws, excluded_carriers):
    """Update Use in Pricing for Merchant Carriers section."""
    excluded = {normalize_merchant_carrier(c) for c in (excluded_carriers or [])}

    header_row_idx, label_col, use_col = _find_pricing_section(ws, 'Merchant Carriers')
    if header_row_idx is None:
        return

    stop_titles = {'MERCHANT SERVICE LEVELS', 'REDO CARRIERS'}
    for row_idx, label_val in _iter_section_rows(ws, header_row_idx + 1, label_col, stop_titles):
        label_norm = normalize_redo_label(label_val)
        target_cell = ws.cell(row_idx, use_col)
        if 'FIRST MILE' in label_norm:
            target_cell.value = 'No'
            continue
        carrier = normalize_merchant_carrier(label_val)
        if not carrier:
            continue
        target_cell.value = 'No' if carrier in excluded else 'Yes'

def _unique_cleaned_services(normalized_df):
    if normalized_df is None or 'CLEANED_SHIPPING_SERVICE' not in normalized_df.columns:
        return []
    seen = set()
    ordered = []
    for value in normalized_df['CLEANED_SHIPPING_SERVICE']:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        text = str(value).strip()
        if not text:
            continue
        norm = normalize_service_name(text)
        if norm not in seen:
            seen.add(norm)
            ordered.append(text)
    return ordered

def update_pricing_summary_merchant_service_levels(ws, selected_services, normalized_df=None):
    """Update Use in Pricing for Merchant Service Levels section."""
    selected_normalized = {normalize_service_name(s) for s in (selected_services or [])}

    stop_titles = {'REDO CARRIERS', 'MERCHANT CARRIERS'}
    header_row_idx, label_col, use_col, rows = _scan_section_rows(
        ws, 'Merchant Service Levels', stop_titles
    )
    if header_row_idx is None:
        return
    services = SERVICE_LEVELS
    row_idx = header_row_idx + 1
    for service in services:
        ws.cell(row_idx, label_col, service)
        normalized = normalize_service_name(service)
        ws.cell(row_idx, use_col, 'Yes' if normalized in selected_normalized else 'No')
        row_idx += 1

    while True:
        cell = ws.cell(row_idx, label_col)
        normalized = normalize_redo_label(cell.value)
        if not normalized or normalized in stop_titles:
            break
        cell.value = None
        ws.cell(row_idx, use_col, None)
        row_idx += 1

def detect_structure(csv_path):
    """Detect if invoice is zone-based or zip-based"""
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        headers = [h.lower() for h in reader.fieldnames or []]
        # Check for zone column (case-insensitive, allow variants)
        zone_keywords = ['zone', 'shipment - zone', 'shipment-zone']
        has_zone = any(any(kw in h for h in headers) for kw in zone_keywords)
        return 'zone' if has_zone else 'zip'

def _detect_weight_unit_from_text(text):
    if not text:
        return None
    cleaned = re.sub(r'[^a-z0-9\s]', ' ', str(text).lower())
    patterns = {
        'oz': re.compile(r'\b(oz|ounce|ounces)\b'),
        'lb': re.compile(r'\b(lb|lbs|pound|pounds)\b'),
        'kg': re.compile(r'\b(kg|kilogram|kilograms)\b')
    }
    for unit, pattern in patterns.items():
        if pattern.search(cleaned):
            return unit
    return None

def detect_weight_unit_from_values(series, sample_size=50):
    if series is None:
        return None
    sample = series.dropna().astype(str).head(sample_size)
    counts = {'oz': 0, 'lb': 0, 'kg': 0}
    for value in sample:
        unit = _detect_weight_unit_from_text(value)
        if unit:
            counts[unit] += 1
    if not any(counts.values()):
        return None
    return max(counts, key=counts.get)

def detect_weight_unit_fallback(df):
    unit_columns = [
        col for col in df.columns
        if 'unit' in col.lower() or 'uom' in col.lower()
    ]
    for col in unit_columns:
        unit = detect_weight_unit_from_values(df[col])
        if unit:
            return unit
    return None

def _annual_orders_missing(mapping_config):
    if not mapping_config:
        return True
    raw_value = mapping_config.get('annual_orders')
    if raw_value is None or str(raw_value).strip() == '':
        return True
    try:
        return float(raw_value) <= 0
    except Exception:
        return True

def _usps_market_discount_values(mapping_config):
    pct = None
    dollar = None
    if mapping_config:
        pct = mapping_config.get('usps_market_pct_off')
        dollar = mapping_config.get('usps_market_dollar_off')
    try:
        pct = float(pct) if pct is not None else 0.05
    except Exception:
        pct = 0.05
    if pct > 1:
        pct = pct / 100
    try:
        dollar = float(dollar) if dollar is not None else 0.0
    except Exception:
        dollar = 0.0
    return pct, dollar

def _avg_label_cost_from_job(job_dir):
    normalized_csv = Path(job_dir) / 'normalized.csv'
    if not normalized_csv.exists():
        return None
    try:
        df = pd.read_csv(normalized_csv)
    except Exception:
        return None
    if df.empty:
        return None
    series = None
    if 'Label Cost' in df.columns:
        series = df['Label Cost']
    elif 'LABEL_COST' in df.columns:
        series = df['LABEL_COST']
    if series is None:
        return None
    numeric = _coerce_numeric_series(series)
    if numeric.notna().any():
        return float(numeric.mean())
    return None

def _carrier_distribution(job_dir, mapping_config, available_carriers):
    normalized_csv = Path(job_dir) / 'normalized.csv'
    if not normalized_csv.exists():
        return {carrier: 0 for carrier in available_carriers}
    try:
        df = pd.read_csv(normalized_csv)
    except Exception:
        return {carrier: 0 for carrier in available_carriers}
    if df.empty:
        return {carrier: 0 for carrier in available_carriers}
    carrier_series = df['Shipping Carrier'] if 'Shipping Carrier' in df.columns else pd.Series([""] * len(df))
    service_series = df['Shipping Service'] if 'Shipping Service' in df.columns else pd.Series([""] * len(df))
    carrier_series = carrier_series.fillna("").astype(str)
    service_series = service_series.fillna("").astype(str)
    counts = {carrier: 0 for carrier in available_carriers}
    total = len(df)
    for carrier_val, service_val in zip(carrier_series, service_series):
        inferred = infer_redo_carrier(carrier_val, service_val)
        if inferred in counts:
            counts[inferred] += 1
    if total <= 0:
        return {carrier: 0 for carrier in available_carriers}
    return {carrier: counts.get(carrier, 0) / total for carrier in available_carriers}

def suggest_mapping(invoice_columns, standard_field):
    """Suggest best matching column for a standard field"""
    invoice_lower = [str(c).lower() for c in invoice_columns]
    field_lower = standard_field.lower()

    def _compact(text):
        return re.sub(r'[^a-z0-9]+', '', text.lower())

    def _tokenize(text):
        cleaned = re.sub(r'[^a-z0-9]+', ' ', text.lower()).strip()
        return cleaned.split()
    
    def _is_bad_label_cost(col_text):
        bad_tokens = ('insurance', 'labelcreatedate', 'create date', 'createdate', 'shipdate', 'date')
        return any(token in col_text for token in bad_tokens)

    compact_field = _compact(standard_field)
    for i, col in enumerate(invoice_lower):
        if _compact(col) == compact_field:
            if standard_field == 'Label Cost' and _is_bad_label_cost(col):
                continue
            return invoice_columns[i]

    rules = {
        'Order Number': {
            'positive': ['order number', 'order #', 'order no', 'order id', 'ordernumber', 'orderid', 'order'],
            'negative': ['date', 'ship', 'time']
        },
        'Order Date': {
            'positive': ['order date', 'ship date', 'shipped date', 'orderdate', 'shipdate', 'date', 'shipped'],
            'negative': ['number', 'id', '#', 'qty', 'count']
        },
        'Zip': {
            'positive': ['zip', 'postal', 'postal code', 'zipcode', 'postalcode'],
            'negative': []
        },
        'Weight': {
            'positive': ['weight', 'oz', 'ounce', 'lb', 'lbs', 'pound', 'kg', 'kilogram'],
            'negative': ['unit']
        },
        'Shipping Carrier': {
            'positive': ['carrier', 'shipper', 'courier'],
            'negative': ['service', 'method', 'level']
        },
        'Shipping Service': {
            'positive': ['service', 'method', 'level'],
            'negative': ['carrier']
        },
        'Package Height': {
            'positive': ['height', 'ht'],
            'negative': ['weight', 'lb', 'lbs', 'oz', 'ounce', 'unit']
        },
        'Package Width': {
            'positive': ['width', 'wd'],
            'negative': ['weight', 'lb', 'lbs', 'oz', 'ounce', 'unit']
        },
        'Package Length': {
            'positive': ['length', 'len'],
            'negative': ['weight', 'lb', 'lbs', 'oz', 'ounce', 'unit']
        },
        'Zone': {
            'positive': ['zone'],
            'negative': []
        },
        'Label Cost': {
            'positive': ['label cost', 'shipping rate', 'rate', 'postage', 'cost', 'carrier fee', 'fee'],
            'negative': ['insurance', 'labelcreatedate', 'createdate', 'shipdate', 'date']
        }
    }

    rule = rules.get(standard_field)
    if not rule:
        return None

    best = (0, None)
    for idx, col in enumerate(invoice_lower):
        if standard_field == 'Label Cost' and _is_bad_label_cost(col):
            continue
        tokens = _tokenize(col)
        col_text = f" {' '.join(tokens)} "
        score = 0
        for phrase in rule['positive']:
            if phrase in col:
                score += 10
            if phrase in col_text:
                score += 10
        for token in rule['positive']:
            if token in tokens:
                score += 5
        for token in rule['negative']:
            if token in tokens or token in col:
                score -= 10
        if standard_field == 'Order Date' and not any(t in tokens for t in ['date', 'ship', 'shipped']):
            score = 0
        if standard_field == 'Order Number' and any(t in tokens for t in ['date', 'ship', 'shipped']):
            score -= 10
        if score > best[0]:
            best = (score, idx)

    if best[1] is not None and best[0] > 0:
        return invoice_columns[best[1]]
    return None

def clean_old_runs():
    """Remove runs older than 24 hours"""
    runs_dir = Path(app.config['UPLOAD_FOLDER'])
    if not runs_dir.exists():
        return
    
    cutoff = datetime.now() - timedelta(hours=24)
    for run_dir in runs_dir.iterdir():
        if run_dir.is_dir():
            try:
                mtime = datetime.fromtimestamp(run_dir.stat().st_mtime)
                if mtime < cutoff:
                    shutil.rmtree(run_dir)
            except Exception:
                pass

ADMIN_LOG_PATH = BASE_DIR / 'admin_log.xlsx'
DEAL_SIZING_HEADERS = [
    'Merchant Name',
    'SE/AE On Deal',
    'Annual Orders',
    '% USPS',
    '% FedEx',
    '% UPS',
    '% Amazon',
    '% UniUni',
    'UPS Average Label Cost',
    'FedEx Actual Spread',
    'Amazon Actual Spread',
    'UniUni Actual Spread',
    'Carrier Spread',
    'Adjusted Annual Orders',
    'Attach Rate',
    'SaaS Fee',
    'SaaS Tier',
    'Per Label Fee',
    '% Of Orders With Fee',
    'CommentSold',
    'Ebay',
    'Ebay Live Selling',
    'Printing Today',
    'Total Deal Size'
]
RATE_CARD_HEADERS = [
    'Timestamp',
    'Job ID',
    'Flow Type',
    'Merchant Name',
    'Merchant ID',
    'Existing Customer',
    'Origin ZIP',
    'Annual Orders',
    'Structure',
    'Zone Column',
    'Mapping JSON',
    'Merchant Pricing JSON',
    'Redo Carriers JSON',
    'USPS Market % Off',
    'USPS Market $ Off'
]

def _ensure_admin_log():
    if not ADMIN_LOG_PATH.exists():
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        ws = wb.create_sheet('Deal sizing')
        ws.append(DEAL_SIZING_HEADERS)
        ws = wb.create_sheet('Rate card + deal sizing')
        ws.append(RATE_CARD_HEADERS)
        wb.save(ADMIN_LOG_PATH)
        return

    wb = openpyxl.load_workbook(ADMIN_LOG_PATH)
    if 'Deal sizing' not in wb.sheetnames:
        ws = wb.create_sheet('Deal sizing')
        ws.append(DEAL_SIZING_HEADERS)
    else:
        ws = wb['Deal sizing']
        for idx, header in enumerate(DEAL_SIZING_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
    if 'Rate card + deal sizing' not in wb.sheetnames:
        ws = wb.create_sheet('Rate card + deal sizing')
        ws.append(RATE_CARD_HEADERS)
    else:
        ws = wb['Rate card + deal sizing']
        for idx, header in enumerate(RATE_CARD_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
    wb.save(ADMIN_LOG_PATH)

def _upsert_admin_row(ws, job_id, row_values):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if str(row[1].value) == str(job_id):
            for idx, value in enumerate(row_values, start=1):
                row[idx - 1].value = value
            return
    ws.append(row_values)

def log_admin_entry(job_id, mapping_config, merchant_pricing, redo_config):
    _ensure_admin_log()
    flow_type = mapping_config.get('flow_type', 'rate_card_plus_deal_sizing') if mapping_config else ''
    if flow_type == 'deal_sizing':
        return
    sheet_name = 'Deal sizing' if flow_type == 'deal_sizing' else 'Rate card + deal sizing'
    pct_off, dollar_off = _usps_market_discount_values(mapping_config)
    row = [
        datetime.now(timezone.utc).isoformat(),
        job_id,
        flow_type,
        mapping_config.get('merchant_name', '') if mapping_config else '',
        mapping_config.get('merchant_id', '') if mapping_config else '',
        bool(mapping_config.get('existing_customer')) if mapping_config else False,
        mapping_config.get('origin_zip', '') if mapping_config else '',
        mapping_config.get('annual_orders', '') if mapping_config else '',
        mapping_config.get('structure', '') if mapping_config else '',
        mapping_config.get('zone_column', '') if mapping_config else '',
        json.dumps(mapping_config.get('mapping', {})) if mapping_config else '{}',
        json.dumps(merchant_pricing or {}),
        json.dumps(redo_config or {}),
        pct_off,
        dollar_off
    ]
    wb = openpyxl.load_workbook(ADMIN_LOG_PATH)
    ws = wb[sheet_name]
    _upsert_admin_row(ws, job_id, row)
    wb.save(ADMIN_LOG_PATH)

@app.route('/')
def index():
    clean_old_runs()
    return render_template('entry.html')

@app.route('/upload')
def upload_page():
    clean_old_runs()
    return render_template('screen1.html')

@app.route('/deal-sizing')
def deal_sizing_page():
    return render_template('deal_sizing.html')

def _build_admin_view_data():
    _ensure_admin_log()
    wb = openpyxl.load_workbook(ADMIN_LOG_PATH, data_only=True)
    deal_ws = wb['Deal sizing'] if 'Deal sizing' in wb.sheetnames else None
    rate_ws = wb['Rate card + deal sizing'] if 'Rate card + deal sizing' in wb.sheetnames else None

    def _sheet_data(ws):
        if ws is None:
            return {'headers': [], 'rows': [], 'row_ids': []}
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        row_ids = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            rows.append(list(row))
            row_ids.append(idx)
        if 'Timestamp' in headers:
            ts_idx = headers.index('Timestamp')
            def _parse_ts(value):
                if not value:
                    return datetime.min
                if isinstance(value, datetime):
                    return value
                try:
                    return datetime.fromisoformat(str(value))
                except Exception:
                    return datetime.min
            combined = list(zip(rows, row_ids))
            combined.sort(key=lambda item: _parse_ts(item[0][ts_idx] if ts_idx < len(item[0]) else None), reverse=True)
            rows = [item[0] for item in combined]
            row_ids = [item[1] for item in combined]
        return {'headers': headers, 'rows': rows, 'row_ids': row_ids}

    deal_data = _sheet_data(deal_ws)
    rate_data = _sheet_data(rate_ws)

    def _format_currency(value):
        try:
            number = float(value)
        except Exception:
            return ''
        return f"${number:,.0f}"

    def _format_percent(value):
        try:
            number = float(value)
        except Exception:
            return ''
        if number <= 1:
            number *= 100
        return f"{number:.0f}%"

    def _format_number(value):
        try:
            number = float(value)
        except Exception:
            return ''
        return f"{number:,.0f}"

    def _format_bool(value):
        return 'Yes' if value else 'No'

    def _parse_bool(value):
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        return str(value).strip().lower() in {'true', '1', 'yes', 'y'}

    def _parse_number(value):
        try:
            return float(value)
        except Exception:
            return 0.0

    def _saas_tier_name(annual_orders):
        try:
            orders = float(annual_orders)
        except Exception:
            orders = 0
        if orders <= 0:
            return ''
        if orders < 5000:
            return 'Starter Tier'
        if orders < 15000:
            return 'Growth Tier'
        if orders < 35000:
            return 'Pro Tier'
        if orders < 50000:
            return 'Scale Tier'
        return 'Enterprise Tier'

    def _summarize_list(values):
        items = [str(value).strip() for value in (values or []) if value]
        return ', '.join(items)

    def _effective_orders(annual_orders, comment_sold, ebay, live_selling):
        adjusted = annual_orders
        if comment_sold:
            adjusted *= 0.6
        if ebay:
            adjusted *= 0.6 if live_selling else 0.95
        return adjusted

    def _deal_spread_from_details(details, effective_orders):
        total = 0.0
        for detail in (details or []):
            spread = _parse_number(detail.get('spread'))
            pct = detail.get('orders_won_pct')
            pct = _parse_number(pct)
            if pct > 1:
                pct /= 100
            total += spread * pct * effective_orders
        return total

    def _reorder_with_total(data, download_label='Download Rate Card'):
        headers = list(data.get('headers') or [])
        rows = list(data.get('rows') or [])
        if not headers:
            return data
        total_label = 'Total Deal Size'
        total_index = None
        if 'Total Size' in headers:
            total_index = headers.index('Total Size')
        elif total_label in headers:
            total_index = headers.index(total_label)
        if total_index is not None:
            headers = headers[:total_index] + headers[total_index + 1:] + [total_label]
            updated = []
            for row in rows:
                total_value = row[total_index] if total_index < len(row) else ''
                new_row = []
                for header in headers:
                    if header == total_label:
                        new_row.append(total_value)
                        continue
                    try:
                        idx = data['headers'].index(header)
                    except ValueError:
                        new_row.append('')
                        continue
                    new_row.append(row[idx] if idx < len(row) else '')
                updated.append(new_row)
            data['headers'] = headers
            data['rows'] = updated
        else:
            if total_label not in headers:
                headers.append(total_label)
                rows = [list(row) + [''] for row in rows]
                data['headers'] = headers
                data['rows'] = rows
        if download_label in headers and headers[-1] != total_label:
            download_idx = headers.index(download_label)
            headers.pop(download_idx)
            headers.insert(len(headers) - 1, download_label)
            updated = []
            for row in data.get('rows', []):
                values = list(row)
                value = values.pop(download_idx) if download_idx < len(values) else ''
                values.insert(len(values) - 1, value)
                updated.append(values)
            data['headers'] = headers
            data['rows'] = updated
        return data

    def _build_admin_groups(headers, sheet_type):
        def group_label(header):
            header = str(header or '')
            if sheet_type == 'deal':
                if header in ('Merchant Name', 'SE/AE On Deal', 'Annual Orders'):
                    return 'Merchant Info'
                if header.startswith('% '):
                    return 'Carrier Mix'
                if header in (
                    'UPS Average Label Cost',
                    'FedEx Actual Spread',
                    'Amazon Actual Spread',
                    'UniUni Actual Spread'
                ):
                    return 'Carrier Inputs'
                if header in (
                    'Carrier Spread',
                    'Adjusted Annual Orders',
                    'Attach Rate',
                    'SaaS Fee',
                    'SaaS Tier',
                    'Per Label Fee',
                    '% Of Orders With Fee',
                    'CommentSold',
                    'Ebay',
                    'Ebay Live Selling',
                    'Printing Today'
                ):
                    return 'Deal Sizing Inputs'
                if header == 'Total Deal Size':
                    return 'Deal Sizes'
                return 'Other'
            if sheet_type == 'rate':
                if header in (
                    'Timestamp',
                    'Merchant Name',
                    'Existing Customer',
                    'Merchant ID',
                    'SE/AE On Deal',
                    'Origin ZIP',
                    'Annual Orders',
                    'Structure',
                    'Raw Invoice File',
                    'Amazon Eligible',
                    'UniUni Eligible'
                ):
                    return 'Merchant Info'
                if header.endswith('Mapped To') or header == 'Units':
                    return 'Mappings'
                if header in ('Merchant Carriers', 'Merchant Service Levels'):
                    return 'Merchant Pricing'
                if header == 'Redo Carriers':
                    return 'Redo Carriers'
                if header in (
                    'Est. Merchant Annual Savings',
                    'Total Spread Available',
                    '% Orders We Could Win',
                    '% Orders Won W/ Spread'
                ):
                    return 'Dashboard Figures'
                if header.endswith('Breakdown'):
                    return 'Carrier Breakdown'
                if header in ('% Off USPS Market Rate', '$ Off USPS Market Rate'):
                    return 'Pricing Inputs'
                if header in (
                    'Est. Redo Deal Size (Popup)',
                    'Carrier Spread',
                    'Adjusted Annual Orders',
                    'Attach Rate',
                    'File Attached',
                    'SaaS Fee',
                    'Per Label Fee',
                    '% Of Orders With Fee',
                    'Annual Orders (Deal Sizing)',
                    'Average Label Cost',
                    'CommentSold',
                    'Ebay',
                    'Live Selling',
                    'Printing Today'
                ):
                    return 'Deal Sizing Inputs'
                if header in ('Download Rate Card', 'View File In Sharepoint'):
                    return 'Files'
                return 'Other'
            return 'Other'

        groups = []
        if not headers:
            return groups
        current = group_label(headers[0])
        start = 0
        for idx, header in enumerate(headers):
            label = group_label(header)
            if label != current:
                groups.append({'label': current, 'span': idx - start})
                current = label
                start = idx
        groups.append({'label': current, 'span': len(headers) - start})
        return groups

    def _format_breakdown(metrics):
        if not metrics:
            return ''
        savings = _format_currency(metrics.get('Est. Merchant Annual Savings'))
        deal_size = _format_currency(metrics.get('Est. Redo Deal Size'))
        spread = _format_currency(metrics.get('Spread Available'))
        could_win = _format_percent(metrics.get('% Orders We Could Win'))
        won_spread = _format_percent(metrics.get('% Orders Won W/ Spread'))
        return f"Savings: {savings} | Deal Size: {deal_size} | Total Spread: {spread} | % We Could Win: {could_win} | % Won W/ Spread: {won_spread}"

    def _available_carriers_for_job(job_dir, mapping_config, excluded):
        raw_csv = job_dir / 'raw_invoice.csv'
        if raw_csv.exists() and mapping_config:
            try:
                raw_df = pd.read_csv(raw_csv)
                available = available_merchant_carriers(raw_df, mapping_config)
                if available:
                    return [c for c in available if c not in excluded]
            except Exception:
                pass
        return [c for c in MERCHANT_CARRIERS if c not in excluded]

    def _deal_inputs_from_mapping(mapping_config):
        inputs = mapping_config.get('deal_sizing_inputs') if mapping_config else None
        return inputs if isinstance(inputs, dict) else {}

    def _build_deal_sizing_view(data):
        headers = data.get('headers') or []
        rows = data.get('rows') or []
        row_ids = data.get('row_ids') or []
        header_index = {str(header): idx for idx, header in enumerate(headers)}

        def value_for(row, *keys):
            for key in keys:
                idx = header_index.get(key)
                if idx is not None and idx < len(row):
                    return row[idx]
            return ''

        target_headers = [
            'Timestamp',
            'Merchant Name',
            'SE/AE On Deal',
            'Annual Orders',
            '% USPS',
            '% FedEx',
            '% UPS',
            '% Amazon',
            '% UniUni',
            'UPS Average Label Cost',
            'FedEx Actual Spread',
            'Amazon Actual Spread',
            'UniUni Actual Spread',
            'Carrier Spread',
            'Adjusted Annual Orders',
            'Attach Rate',
            'SaaS Fee',
            'SaaS Tier',
            'Per Label Fee',
            '% Of Orders With Fee',
            'CommentSold',
            'Ebay',
            'Ebay Live Selling',
            'Printing Today',
            'Total Deal Size'
        ]

        updated_rows = []
        for row in rows:
            timestamp = value_for(row, 'Timestamp')
            merchant_name = value_for(row, 'Merchant Name', 'Merchant')
            se_ae_on_deal = value_for(row, 'SE/AE On Deal')
            annual_orders = value_for(row, 'Annual Orders')
            pct_usps = value_for(row, '% USPS')
            pct_fedex = value_for(row, '% FedEx')
            pct_ups = value_for(row, '% UPS')
            pct_amazon = value_for(row, '% Amazon')
            pct_uniuni = value_for(row, '% UniUni')
            ups_avg_label = value_for(row, 'UPS Average Label Cost')
            fedex_spread = value_for(row, 'FedEx Actual Spread')
            amazon_spread = value_for(row, 'Amazon Actual Spread')
            uniuni_spread = value_for(row, 'UniUni Actual Spread')
            attach_rate = value_for(row, 'Attach Rate')
            saas_fee = value_for(row, 'SaaS Fee', 'Monthly SaaS')
            per_label_fee = value_for(row, 'Per Label Fee')
            fee_order_pct = value_for(row, '% Of Orders With Fee')
            comment_sold = value_for(row, 'CommentSold', 'CommentSold?')
            ebay = value_for(row, 'Ebay')
            live_selling = value_for(row, 'Ebay Live Selling', 'Live Selling')
            printing = value_for(row, 'Printing Today', 'Printing?')
            total_deal_size = value_for(row, 'Total Deal Size', 'Total Size')

            usps_size = _parse_number(value_for(row, 'USPS Size'))
            fedex_size = _parse_number(value_for(row, 'FedEx Size'))
            ups_size = _parse_number(value_for(row, 'UPS Size'))
            amazon_size = _parse_number(value_for(row, 'Amazon Size'))
            uniuni_size = _parse_number(value_for(row, 'UniUni Size'))
            carrier_spread = usps_size + fedex_size + ups_size + amazon_size + uniuni_size
            carrier_spread_value = carrier_spread if carrier_spread else value_for(row, 'Carrier Spread')

            annual_orders_value = _parse_number(annual_orders)
            comment_sold_flag = _parse_bool(comment_sold)
            ebay_flag = _parse_bool(ebay)
            live_flag = _parse_bool(live_selling)
            adjusted_orders = _effective_orders(annual_orders_value, comment_sold_flag, ebay_flag, live_flag)
            adjusted_orders_value = adjusted_orders if adjusted_orders else value_for(row, 'Adjusted Annual Orders')

            updated_rows.append([
                timestamp,
                merchant_name,
                se_ae_on_deal,
                annual_orders,
                pct_usps,
                pct_fedex,
                pct_ups,
                pct_amazon,
                pct_uniuni,
                ups_avg_label,
                fedex_spread,
                amazon_spread,
                uniuni_spread,
                carrier_spread_value,
                adjusted_orders_value,
                attach_rate,
                saas_fee,
                _saas_tier_name(annual_orders_value),
                per_label_fee,
                fee_order_pct,
                _format_bool(comment_sold_flag) if comment_sold != '' else '',
                _format_bool(ebay_flag) if ebay != '' else '',
                _format_bool(live_flag) if live_selling != '' else '',
                _format_bool(_parse_bool(printing)) if printing != '' else '',
                total_deal_size
            ])

        return {'headers': target_headers, 'rows': updated_rows, 'row_ids': row_ids}
    if rate_data.get('headers'):
        headers = rate_data['headers']
        header_index = {str(header): idx for idx, header in enumerate(headers)}
        job_idx = header_index.get('Job ID')
        ts_idx = header_index.get('Timestamp')
        merchant_name_idx = header_index.get('Merchant Name')
        merchant_id_idx = header_index.get('Merchant ID')
        existing_idx = header_index.get('Existing Customer')
        origin_idx = header_index.get('Origin ZIP')
        annual_idx = header_index.get('Annual Orders')
        structure_idx = header_index.get('Structure')
        mapping_idx = header_index.get('Mapping JSON')
        pricing_idx = header_index.get('Merchant Pricing JSON')
        redo_idx = header_index.get('Redo Carriers JSON')
        pct_off_idx = header_index.get('USPS Market % Off')
        dollar_off_idx = header_index.get('USPS Market $ Off')

        standard_fields = STANDARD_FIELDS['required'] + STANDARD_FIELDS['optional']
        mapping_headers = []
        for field in standard_fields:
            mapping_headers.append(f'{field} Mapped To')
            if field == 'Weight':
                mapping_headers.append('Units')
        breakdown_headers = [f'{carrier} Breakdown' for carrier in DASHBOARD_CARRIERS]

        rate_headers = [
            'Timestamp',
            'Merchant Name',
            'Existing Customer',
            'Merchant ID',
            'SE/AE On Deal',
            'Origin ZIP',
            'Annual Orders',
            'Structure',
            'Raw Invoice File',
            'Amazon Eligible',
            'UniUni Eligible'
        ] + mapping_headers + [
            'Merchant Carriers',
            'Merchant Service Levels',
            'Redo Carriers',
            'Est. Merchant Annual Savings',
            'Total Spread Available',
            '% Orders We Could Win',
            '% Orders Won W/ Spread'
        ] + breakdown_headers + [
            '% Off USPS Market Rate',
            '$ Off USPS Market Rate',
            'Est. Redo Deal Size (Popup)',
            'Carrier Spread',
            'Adjusted Annual Orders',
            'Attach Rate',
            'File Attached',
            'SaaS Fee',
            'Per Label Fee',
            '% Of Orders With Fee',
            'Annual Orders (Deal Sizing)',
            'Average Label Cost',
            'CommentSold',
            'Ebay',
            'Live Selling',
            'Printing Today',
            'Download Rate Card',
            'View File In Sharepoint'
        ]

        rate_rows = []
        for row in rate_data.get('rows', []):
            job_id = row[job_idx] if job_idx is not None and job_idx < len(row) else ''
            timestamp = row[ts_idx] if ts_idx is not None and ts_idx < len(row) else ''
            job_dir = Path(app.config['UPLOAD_FOLDER']) / str(job_id)
            mapping_config = {}
            merchant_pricing = {}
            redo_config = {}
            if job_id and job_dir.exists():
                mapping_file = job_dir / 'mapping.json'
                if mapping_file.exists():
                    try:
                        with open(mapping_file, 'r') as f:
                            mapping_config = json.load(f)
                    except Exception:
                        mapping_config = {}
                pricing_file = job_dir / 'merchant_pricing.json'
                if pricing_file.exists():
                    try:
                        with open(pricing_file, 'r') as f:
                            merchant_pricing = json.load(f)
                    except Exception:
                        merchant_pricing = {}
                redo_file = job_dir / 'redo_carriers.json'
                if redo_file.exists():
                    try:
                        with open(redo_file, 'r') as f:
                            redo_config = json.load(f)
                    except Exception:
                        redo_config = {}

            mapping_json = row[mapping_idx] if mapping_idx is not None and mapping_idx < len(row) else None
            pricing_json = row[pricing_idx] if pricing_idx is not None and pricing_idx < len(row) else None
            redo_json = row[redo_idx] if redo_idx is not None and redo_idx < len(row) else None
            if not mapping_config and mapping_json:
                try:
                    mapping_config = json.loads(mapping_json)
                except Exception:
                    mapping_config = {}
            if not merchant_pricing and pricing_json:
                try:
                    merchant_pricing = json.loads(pricing_json)
                except Exception:
                    merchant_pricing = {}
            if not redo_config and redo_json:
                try:
                    redo_config = json.loads(redo_json)
                except Exception:
                    redo_config = {}

            merchant_name = mapping_config.get('merchant_name') or (
                row[merchant_name_idx] if merchant_name_idx is not None and merchant_name_idx < len(row) else ''
            )
            merchant_id = mapping_config.get('merchant_id') or (
                row[merchant_id_idx] if merchant_id_idx is not None and merchant_id_idx < len(row) else ''
            )
            existing_customer = mapping_config.get('existing_customer')
            if existing_customer is None and existing_idx is not None and existing_idx < len(row):
                existing_customer = row[existing_idx]
            origin_zip = mapping_config.get('origin_zip') or (
                row[origin_idx] if origin_idx is not None and origin_idx < len(row) else ''
            )
            annual_orders = mapping_config.get('annual_orders') or (
                row[annual_idx] if annual_idx is not None and annual_idx < len(row) else ''
            )
            structure = mapping_config.get('structure') or (
                row[structure_idx] if structure_idx is not None and structure_idx < len(row) else ''
            )
            se_ae_on_deal = mapping_config.get('se_ae_on_deal', '')
            raw_invoice_file = mapping_config.get('raw_invoice_file', '')
            raw_invoice_cell = {
                'href': f'/download/{job_id}/raw-invoice',
                'label': raw_invoice_file or 'Download'
            } if job_id else ''

            eligibility = None
            try:
                eligibility = compute_eligibility(origin_zip, annual_orders, mapping_config=mapping_config)
            except Exception:
                eligibility = None
            amazon_eligible = mapping_config.get('amazon_eligible')
            uniuni_eligible = mapping_config.get('uniuni_eligible')
            if amazon_eligible is None and eligibility:
                amazon_eligible = eligibility.get('amazon_eligible_final')
            if uniuni_eligible is None and eligibility:
                uniuni_eligible = eligibility.get('uniuni_eligible_final')

            mapped_values = mapping_config.get('mapping', {}) if mapping_config else {}
            mapped_columns = []
            for field in standard_fields:
                mapped_columns.append(mapped_values.get(field, ''))
                if field == 'Weight':
                    mapped_columns.append(mapped_values.get('Weight Unit', ''))

            excluded_carriers = merchant_pricing.get('excluded_carriers', []) if merchant_pricing else []
            merchant_carriers = _summarize_list(_available_carriers_for_job(job_dir, mapping_config, excluded_carriers)) if job_id else ''
            merchant_services = _summarize_list(merchant_pricing.get('included_services', []) if merchant_pricing else [])
            redo_carriers = _summarize_list(redo_config.get('selected_carriers', []) if redo_config else [])

            pct_off = row[pct_off_idx] if pct_off_idx is not None and pct_off_idx < len(row) else ''
            dollar_off = row[dollar_off_idx] if dollar_off_idx is not None and dollar_off_idx < len(row) else ''

            selected_carriers = redo_config.get('selected_carriers', []) if redo_config else []
            selected_dashboard = [c for c in DASHBOARD_CARRIERS if c in selected_carriers] or list(DASHBOARD_CARRIERS)

            summary_metrics = {}
            breakdown_metrics = {}
            source_mtime = None
            selection_key = None
            if job_id and job_dir.exists():
                rate_cards = list(job_dir.glob('* - Rate Card.xlsx'))
                if rate_cards:
                    try:
                        source_mtime = int(rate_cards[0].stat().st_mtime)
                    except Exception:
                        source_mtime = 0
                    selection_key = _selection_cache_key(selected_dashboard)
                    summary_metrics = _read_summary_cache(job_dir, source_mtime, selection_key) or {}
                    breakdown_cached, _ = _read_breakdown_cache(job_dir, source_mtime)
                    if breakdown_cached is not None:
                        breakdown_metrics = {
                            entry.get('carrier'): entry.get('metrics', {})
                            for entry in breakdown_cached
                            if entry.get('carrier') in DASHBOARD_CARRIERS
                        }

            breakdown_cells = []
            for carrier in DASHBOARD_CARRIERS:
                breakdown_cells.append(_format_breakdown(breakdown_metrics.get(carrier, {})))

            deal_inputs = _deal_inputs_from_mapping(mapping_config)
            deal_annual_orders = _parse_number(deal_inputs.get('annual_orders') or 0)
            deal_avg_label_cost = _parse_number(deal_inputs.get('avg_label_cost') or 0)
            deal_per_label_fee = _parse_number(deal_inputs.get('per_label_fee') or 0)
            deal_fee_order_pct = _parse_number(deal_inputs.get('fee_order_pct') or 0)
            if deal_fee_order_pct > 1:
                deal_fee_order_pct /= 100
            deal_attach_rate = _parse_number(deal_inputs.get('attach_rate') or 0)
            deal_saas_fee = _parse_number(deal_inputs.get('saas_fee') or 0)
            deal_comment_sold = _parse_bool(deal_inputs.get('comment_sold'))
            deal_ebay = _parse_bool(deal_inputs.get('ebay'))
            deal_live_selling = _parse_bool(deal_inputs.get('live_selling'))
            deal_printing = _parse_bool(deal_inputs.get('printing'))
            deal_attach_upload = deal_inputs.get('attach_upload_name')

            adjusted_orders = _effective_orders(deal_annual_orders, deal_comment_sold, deal_ebay, deal_live_selling)
            effective_orders = adjusted_orders * (deal_attach_rate / 100) if deal_attach_rate else adjusted_orders
            carrier_spread_total = 0.0
            if job_id and job_dir.exists() and selected_dashboard and source_mtime is not None and selection_key:
                try:
                    details = _read_carrier_details_cache(job_dir, source_mtime, selection_key) or {}
                    if details:
                        carrier_spread_total = _deal_spread_from_details(list(details.values()), effective_orders)
                except Exception:
                    carrier_spread_total = 0.0
            label_fee_total = deal_per_label_fee * deal_fee_order_pct * deal_annual_orders
            popup_deal_size = carrier_spread_total + deal_saas_fee + label_fee_total

            file_attached = 'Yes' if (deal_attach_rate > 85 and deal_attach_upload) else ('No' if deal_attach_rate > 85 else '')

            download_cell = {
                'href': f'/download/{job_id}/rate-card',
                'label': 'Download'
            } if job_id else ''
            sharepoint_cell = {
                'href': 'https://redotechinc.sharepoint.com/sites/Redo/Shared%20Documents/Forms/AllItems.aspx?id=%2Fsites%2FRedo%2FShared%20Documents%2FAll%20Rate%20Cards%2FUpdating&viewid=90e9a3fe%2Df7a4%2D42c6%2Da353%2D200c53d2e4f3&view=7',
                'label': 'View'
            }

            rate_rows.append([
                timestamp,
                merchant_name,
                _format_bool(_parse_bool(existing_customer)),
                merchant_id,
                se_ae_on_deal,
                origin_zip,
                annual_orders,
                structure,
                raw_invoice_cell,
                _format_bool(_parse_bool(amazon_eligible)) if amazon_eligible is not None else '',
                _format_bool(_parse_bool(uniuni_eligible)) if uniuni_eligible is not None else '',
                *mapped_columns,
                merchant_carriers,
                merchant_services,
                redo_carriers,
                _format_currency(summary_metrics.get('Est. Merchant Annual Savings')),
                _format_currency(summary_metrics.get('Spread Available')),
                _format_percent(summary_metrics.get('% Orders We Could Win')),
                _format_percent(summary_metrics.get('% Orders Won W/ Spread')),
                *breakdown_cells,
                pct_off,
                dollar_off,
                _format_currency(popup_deal_size),
                _format_currency(carrier_spread_total),
                _format_number(adjusted_orders),
                _format_percent(deal_attach_rate),
                file_attached,
                _format_currency(deal_saas_fee),
                _format_currency(deal_per_label_fee),
                _format_percent(deal_fee_order_pct),
                _format_number(deal_annual_orders),
                _format_currency(deal_avg_label_cost),
                _format_bool(deal_comment_sold),
                _format_bool(deal_ebay),
                _format_bool(deal_live_selling),
                _format_bool(deal_printing),
                download_cell,
                sharepoint_cell
            ])

        rate_data = {'headers': rate_headers, 'rows': rate_rows, 'row_ids': rate_data.get('row_ids') or []}
        rate_data['groups'] = _build_admin_groups(rate_data.get('headers') or [], 'rate')

    deal_data = _build_deal_sizing_view(deal_data)
    deal_data['groups'] = _build_admin_groups(deal_data.get('headers') or [], 'deal')
    wb.close()
    return deal_data, rate_data

@app.route('/admin')
def admin_page():
    deal_data, rate_data = _build_admin_view_data()
    return render_template('admin.html', deal_data=deal_data, rate_data=rate_data)

@app.route('/admin/download')
def admin_download():
    deal_data, rate_data = _build_admin_view_data()

    def _flatten_cell(cell):
        if isinstance(cell, dict):
            return cell.get('label') or cell.get('href') or ''
        return '' if cell is None else cell

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    deal_ws = wb.create_sheet('Deal sizing')
    deal_ws.append(deal_data.get('headers') or [])
    for row in deal_data.get('rows') or []:
        deal_ws.append([_flatten_cell(cell) for cell in row])

    rate_ws = wb.create_sheet('Rate card + deal sizing')
    rate_ws.append(rate_data.get('headers') or [])
    for row in rate_data.get('rows') or []:
        rate_ws.append([_flatten_cell(cell) for cell in row])

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        temp_path = tmp.name
    wb.save(temp_path)

    @after_this_request
    def _cleanup(response):
        try:
            os.remove(temp_path)
        except OSError:
            pass
        return response

    return send_file(temp_path, as_attachment=True, download_name='admin_log.xlsx')

@app.route('/api/admin/clear', methods=['POST'])
def admin_clear():
    payload = request.get_json(silent=True) or {}
    sheet_key = (payload.get('sheet') or '').strip().lower()
    sheet_name = 'Deal sizing' if sheet_key == 'deal' else 'Rate card + deal sizing'
    _ensure_admin_log()
    wb = openpyxl.load_workbook(ADMIN_LOG_PATH)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return jsonify({'error': 'Sheet not found'}), 404
    ws = wb[sheet_name]
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    wb.save(ADMIN_LOG_PATH)
    wb.close()
    return jsonify({'success': True})

@app.route('/api/admin/delete', methods=['POST'])
def admin_delete_row():
    payload = request.get_json(silent=True) or {}
    sheet_key = (payload.get('sheet') or '').strip().lower()
    row_id = payload.get('row_id')
    sheet_name = 'Deal sizing' if sheet_key == 'deal' else 'Rate card + deal sizing'
    try:
        row_id = int(row_id)
    except Exception:
        return jsonify({'error': 'Invalid row'}), 400
    if row_id < 2:
        return jsonify({'error': 'Invalid row'}), 400
    _ensure_admin_log()
    wb = openpyxl.load_workbook(ADMIN_LOG_PATH)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return jsonify({'error': 'Sheet not found'}), 404
    ws = wb[sheet_name]
    if row_id > ws.max_row:
        wb.close()
        return jsonify({'error': 'Row not found'}), 404
    ws.delete_rows(row_id, 1)
    wb.save(ADMIN_LOG_PATH)
    wb.close()
    return jsonify({'success': True})

@app.route('/api/deal-sizing-inputs/<job_id>', methods=['POST'])
def save_deal_sizing_inputs(job_id):
    payload = request.get_json(silent=True) or {}
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return jsonify({'error': 'Job not found'}), 404
    mapping_file = job_dir / 'mapping.json'
    if not mapping_file.exists():
        return jsonify({'error': 'Mapping not found'}), 404
    try:
        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)
    except Exception:
        mapping_config = {}
    mapping_config['deal_sizing_inputs'] = payload
    
    # If annual_orders is provided in the payload, update mapping and recalculate eligibility
    annual_orders = payload.get('annual_orders')
    if annual_orders is not None:
        annual_orders_value = _parse_numeric_value(annual_orders)
        if annual_orders_value is not None and annual_orders_value > 0:
            mapping_config['annual_orders'] = int(annual_orders_value)
            # Clear explicit eligibility overrides
            for key in ['amazon_eligible', 'uniuni_eligible', 'uniuni_qualified', 'uniuni']:
                if key in mapping_config:
                    del mapping_config[key]
            
            # Save updated mapping
            with open(mapping_file, 'w') as f:
                json.dump(mapping_config, f)
            
            # Recalculate eligibility
            eligibility = compute_eligibility(
                mapping_config.get('origin_zip'),
                annual_orders_value,
                mapping_config=mapping_config
            )
            app.logger.info(f"Deal sizing annual orders update - job_id={job_id}, annual_orders={annual_orders_value}, amazon_eligible={eligibility['amazon_eligible_final']}, uniuni_eligible={eligibility['uniuni_eligible_final']}")
            
            # Sync redo carriers
            redo_file = job_dir / 'redo_carriers.json'
            if redo_file.exists():
                with open(redo_file, 'r') as f:
                    redo_config = json.load(f)
                selected = redo_config.get('selected_carriers', [])
                changed = False
                if eligibility['amazon_eligible_final'] and 'Amazon' not in selected:
                    selected.append('Amazon')
                    changed = True
                if eligibility['uniuni_eligible_final'] and 'UniUni' not in selected:
                    selected.append('UniUni')
                    changed = True
                if not eligibility['amazon_eligible_final'] and 'Amazon' in selected:
                    selected = [c for c in selected if c != 'Amazon']
                    changed = True
                if not eligibility['uniuni_eligible_final'] and 'UniUni' in selected:
                    selected = [c for c in selected if c != 'UniUni']
                    changed = True
                if changed:
                    with open(redo_file, 'w') as f:
                        json.dump({'selected_carriers': selected}, f)
            
            # Sync merchant pricing
            pricing_file = job_dir / 'merchant_pricing.json'
            if pricing_file.exists():
                with open(pricing_file, 'r') as f:
                    merchant_pricing = json.load(f)
                excluded = merchant_pricing.get('excluded_carriers', [])
                changed = False
                if eligibility['amazon_eligible_final'] and 'Amazon' in excluded:
                    excluded = [c for c in excluded if c != 'Amazon']
                    changed = True
                if eligibility['uniuni_eligible_final'] and 'UniUni' in excluded:
                    excluded = [c for c in excluded if c != 'UniUni']
                    changed = True
                if not eligibility['amazon_eligible_final'] and 'Amazon' not in excluded:
                    excluded.append('Amazon')
                    changed = True
                if not eligibility['uniuni_eligible_final'] and 'UniUni' not in excluded:
                    excluded.append('UniUni')
                    changed = True
                if changed:
                    merchant_pricing['excluded_carriers'] = excluded
                    with open(pricing_file, 'w') as f:
                        json.dump(merchant_pricing, f)
            
            # Clear dashboard caches
            for cache_path in [_summary_cache_path(job_dir), _cache_path_for_job(job_dir), _carrier_details_cache_path(job_dir)]:
                if cache_path.exists():
                    cache_path.unlink()
            
            # Clear in-memory caches
            job_prefix = f"{job_dir.name}:"
            with summary_jobs_lock:
                for key in list(summary_jobs.keys()):
                    if key.startswith(job_prefix):
                        summary_jobs.pop(key, None)
            with dashboard_jobs_lock:
                for key in list(dashboard_jobs.keys()):
                    if key.startswith(job_prefix):
                        dashboard_jobs.pop(key, None)
            
            return jsonify({
                'success': True,
                'amazon_eligible': eligibility['amazon_eligible_final'],
                'uniuni_eligible': eligibility['uniuni_eligible_final']
            })
    
    with open(mapping_file, 'w') as f:
        json.dump(mapping_config, f)
    return jsonify({'success': True})

@app.route('/api/deal-sizing-standalone', methods=['POST'])
def deal_sizing_standalone():
    try:
        payload = request.get_json(silent=True) or {}
        merchant = (payload.get('merchant') or '').strip()
        annual_orders = payload.get('annual_orders')
        if not merchant:
            return jsonify({'error': 'Merchant is required'}), 400
        try:
            annual_orders_value = float(annual_orders) if annual_orders not in (None, '') else 0
        except Exception:
            annual_orders_value = 0
        if annual_orders_value <= 0:
            return jsonify({'error': 'Annual orders is required'}), 400

        _ensure_admin_log()
        wb = openpyxl.load_workbook(ADMIN_LOG_PATH)
        ws = wb['Deal sizing']
        def _num(value):
            try:
                return float(value)
            except Exception:
                return 0.0

        pct_usps = payload.get('pct_usps', '')
        pct_fedex = payload.get('pct_fedex', '')
        pct_ups = payload.get('pct_ups', '')
        pct_amazon = payload.get('pct_amazon', '')
        pct_uniuni = payload.get('pct_uniuni', '')
        ups_avg_label = payload.get('ups_avg_label_cost', '')
        fedex_spread = payload.get('fedex_actual_spread', '')
        amazon_spread = payload.get('amazon_actual_spread', '')
        uniuni_spread = payload.get('uniuni_actual_spread', '')
        attach_rate = payload.get('attach_rate', '')
        per_label_fee = payload.get('per_label_fee', '')
        fee_order_pct = payload.get('fee_order_pct', '')
        comment_sold = payload.get('commentsold', False)
        ebay = payload.get('ebay', False)
        live_selling = payload.get('live_selling', False)
        printing = payload.get('printing', False)
        saas_fee = payload.get('monthly_saas', '')
        usps_size = _num(payload.get('usps_size', 0))
        fedex_size = _num(payload.get('fedex_size', 0))
        ups_size = _num(payload.get('ups_size', 0))
        amazon_size = _num(payload.get('amazon_size', 0))
        uniuni_size = _num(payload.get('uniuni_size', 0))
        carrier_spread = usps_size + fedex_size + ups_size + amazon_size + uniuni_size
        total_size = payload.get('total_size', '')
        adjusted_orders = _effective_orders(annual_orders_value, _parse_bool(comment_sold), _parse_bool(ebay), _parse_bool(live_selling))
        saas_tier = _saas_tier_name(annual_orders_value)

        row = [
            merchant,
            payload.get('se_ae_on_deal', ''),
            annual_orders_value,
            pct_usps,
            pct_fedex,
            pct_ups,
            pct_amazon,
            pct_uniuni,
            ups_avg_label,
            fedex_spread,
            amazon_spread,
            uniuni_spread,
            carrier_spread,
            adjusted_orders,
            attach_rate,
            saas_fee,
            saas_tier,
            per_label_fee,
            fee_order_pct,
            comment_sold,
            ebay,
            live_selling,
            printing,
            total_size
        ]
        ws.append(row)
        wb.save(ADMIN_LOG_PATH)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/mapping')
def mapping_page():
    """Render mapping page"""
    job_id = request.args.get('job_id')
    if not job_id:
        return render_template('screen1.html'), 400
    
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return render_template('screen1.html'), 404
    
    # Load CSV to get columns
    raw_csv_path = job_dir / 'raw_invoice.csv'
    if not raw_csv_path.exists():
        return render_template('screen1.html'), 404
    
    structure = detect_structure(raw_csv_path)
    df = pd.read_csv(raw_csv_path, nrows=1)
    df_sample = pd.read_csv(raw_csv_path, nrows=200)
    columns = list(df.columns)
    
    # Load existing mapping if available
    mapping_file = job_dir / 'mapping.json'
    suggested_mapping = {}
    if mapping_file.exists():
        with open(mapping_file, 'r') as f:
            config = json.load(f)
            suggested_mapping = config.get('mapping', {})
    
    # Suggest mappings
    suggestions = {}
    all_fields = STANDARD_FIELDS['required'] + STANDARD_FIELDS['optional']
    required_fields = required_fields_for_structure(structure)
    optional_fields = [f for f in all_fields if f not in required_fields]
    for field in all_fields:
        suggested = suggest_mapping(columns, field)
        if suggested:
            suggestions[field] = suggested
    
    weight_unit_by_column = {}
    for col in columns:
        unit = _detect_weight_unit_from_text(col)
        if unit:
            weight_unit_by_column[col] = unit
    weight_unit_fallback = detect_weight_unit_fallback(df_sample)

    return render_template('screen2.html', 
                         job_id=job_id,
                         columns=columns,
                         suggestions=suggestions,
                         standard_fields={'required': required_fields, 'optional': optional_fields},
                         weight_unit_by_column=weight_unit_by_column,
                         weight_unit_fallback=weight_unit_fallback)

@app.route('/service-levels')
def service_levels_page():
    """Render service levels selection page"""
    job_id = request.args.get('job_id')
    if not job_id:
        return render_template('screen1.html'), 400
    
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return render_template('screen1.html'), 404
    
    mapping_file = job_dir / 'mapping.json'
    if not mapping_file.exists():
        return render_template('screen1.html'), 404

    with open(mapping_file, 'r') as f:
        mapping_config = json.load(f)

    raw_df = pd.read_csv(job_dir / 'raw_invoice.csv')
    available_services = available_merchant_services(raw_df, mapping_config)

    selected_services = []
    service_file = job_dir / 'service_levels.json'
    if service_file.exists():
        with open(service_file, 'r') as f:
            config = json.load(f)
            selected_services = config.get('selected_services', [])

    return render_template('service_levels.html',
                          job_id=job_id,
                          service_levels=available_services,
                          selected_services=selected_services)

@app.route('/merchant-pricing')
def merchant_pricing_page():
    """Render merchant pricing selection page"""
    job_id = request.args.get('job_id')
    if not job_id:
        return "Missing job_id", 400
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return "Job not found", 404
    return render_template('merchant_pricing.html', job_id=job_id)

@app.route('/redo-carriers')
def redo_carriers_page():
    """Render redo carrier selection page"""
    job_id = request.args.get('job_id')
    if not job_id:
        return render_template('screen1.html'), 400

    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return render_template('screen1.html'), 404

    return render_template('redo_carriers.html', job_id=job_id)

@app.route('/loading')
def loading_page():
    """Render loading page"""
    job_id = request.args.get('job_id') or request.args.get('token')
    if not job_id:
        return render_template('screen1.html'), 400
    
    return render_template('loading.html', token=job_id)

@app.route('/ready')
def ready_page():
    """Render rate card ready page"""
    job_id = request.args.get('job_id')
    if not job_id:
        return render_template('screen1.html'), 400
    
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return render_template('screen1.html'), 404
    
    # Load merchant name
    merchant_name = 'Merchant'
    mapping_file = job_dir / 'mapping.json'
    if mapping_file.exists():
        with open(mapping_file, 'r') as f:
            config = json.load(f)
            merchant_name = config.get('merchant_name', 'Merchant')
    
    return render_template('screen3.html', job_id=job_id, merchant_name=merchant_name)

@app.route('/dashboard')
def dashboard_page():
    """Render summary dashboard page"""
    job_id = request.args.get('job_id')
    if not job_id:
        return render_template('screen1.html'), 400
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return render_template('screen1.html'), 404
    merchant_name = 'Merchant'
    mapping_file = job_dir / 'mapping.json'
    if mapping_file.exists():
        with open(mapping_file, 'r') as f:
            config = json.load(f)
            merchant_name = config.get('merchant_name', 'Merchant')
    return render_template('dashboard.html', job_id=job_id, merchant_name=merchant_name)

@app.route('/api/upload', methods=['POST'])
def upload():
    """Handle file upload and detect structure"""
    try:
        # Accept multiple field names
        file = request.files.get('invoice') or request.files.get('invoice_file') or request.files.get('invoice_csv')
        if not file or file.filename == '':
            return jsonify({'error': 'No file uploaded'}), 400
        
        # Create job directory
        job_id = str(uuid.uuid4())
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        job_dir.mkdir(parents=True, exist_ok=True)
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        ext = Path(filename).suffix.lower()
        raw_csv_path = job_dir / 'raw_invoice.csv'
        if ext == '.xlsx':
            raw_xlsx_path = job_dir / 'raw_invoice.xlsx'
            file.save(raw_xlsx_path)
            try:
                sheets = pd.read_excel(raw_xlsx_path, sheet_name=None, dtype=str)
            except Exception as e:
                return jsonify({'error': f'Failed to read XLSX: {str(e)}'}), 400
            if not sheets:
                return jsonify({'error': 'XLSX file has no readable sheets'}), 400

            def score_sheet(df):
                columns = [str(c).strip().lower() for c in df.columns if c is not None]
                if not columns:
                    return 0
                keywords = ('service', 'carrier', 'shipping', 'order', 'zip', 'postal', 'weight', 'zone')
                keyword_hits = sum(1 for c in columns if any(k in c for k in keywords))
                non_empty_cols = sum(1 for c in columns if c)
                return keyword_hits * 10 + non_empty_cols

            df_upload = max(sheets.values(), key=score_sheet)
            df_upload.columns = [
                str(c).strip() if c is not None else ''
                for c in df_upload.columns
            ]
            def _ensure_shipping_service_column(frame):
                normalized = {
                    re.sub(r'\W+', '', str(c).strip().lower()): c for c in frame.columns
                }
                if 'shippingservice' in normalized:
                    return frame
                if frame.shape[1] > 28:
                    cols = list(frame.columns)
                    cols[28] = 'ShippingService'
                    frame.columns = cols
                return frame

            df_upload = _ensure_shipping_service_column(df_upload)
            normalized_cols = {
                re.sub(r'\W+', '', str(c).strip().lower()): c for c in df_upload.columns
            }
            if 'shippingservice' not in normalized_cols and 'shipping_service' not in normalized_cols:
                raw_sheet_name = None
                for name, sheet_df in sheets.items():
                    if sheet_df is df_upload:
                        raw_sheet_name = name
                        break
                if raw_sheet_name is None:
                    raw_sheet_name = list(sheets.keys())[0]
                df_raw = pd.read_excel(raw_xlsx_path, sheet_name=raw_sheet_name, header=None, dtype=str)
                header_row = None
                for idx in range(min(20, len(df_raw))):
                    row_values = df_raw.iloc[idx].fillna('').astype(str)
                    normalized_row = [
                        re.sub(r'\W+', '', val.strip().lower()) for val in row_values
                    ]
                    if any('shippingservice' in val or ('shipping' in val and 'service' in val) for val in normalized_row):
                        header_row = idx
                        break
                if header_row is not None:
                    header_values = df_raw.iloc[header_row].fillna('').astype(str).tolist()
                    df_upload = df_raw.iloc[header_row + 1:].copy()
                    df_upload.columns = [str(c).strip() for c in header_values]
                    df_upload = df_upload.loc[:, df_upload.columns != '']
                    df_upload = _ensure_shipping_service_column(df_upload)
            df_upload.to_csv(raw_csv_path, index=False)
        else:
            file.save(raw_csv_path)
        
        # Detect structure
        structure = detect_structure(raw_csv_path)
        
        # Read CSV to get columns and suggest merchant name
        df = pd.read_csv(raw_csv_path, nrows=5)
        columns = list(df.columns)
        
        # Merchant name is manual; do not auto-suggest.
        merchant_name_suggestion = None
        
        # Record upload phase timestamp with started_at as baseline
        progress_file = job_dir / 'progress.json'
        now = datetime.now(timezone.utc).isoformat()
        with open(progress_file, 'w') as f:
            json.dump({
                'started_at': now,
                'phase_timestamps': {
                    'upload': now
                }
            }, f)
        
        return jsonify({
            'job_id': job_id,
            'detected_structure': structure,
            'columns': columns,
            'merchant_name_suggestion': merchant_name_suggestion or '',
            'requires_origin_zip': structure == 'zip'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mapping', methods=['POST'])
def mapping():
    """Save mapping configuration"""
    try:
        data = request.json
        job_id = data.get('job_id')
        merchant_name = data.get('merchant_name', '')
        merchant_id = data.get('merchant_id', '')
        existing_customer = data.get('existing_customer', False)
        se_ae_on_deal = data.get('se_ae_on_deal', '')
        origin_zip = data.get('origin_zip', '')
        annual_orders = data.get('annual_orders', '')
        raw_invoice_file = data.get('raw_invoice_file', '')
        amazon_eligible = data.get('amazon_eligible')
        uniuni_eligible = data.get('uniuni_eligible')
        mapping_config = data.get('mapping', {})
        
        if not job_id:
            return jsonify({'error': 'job_id required'}), 400
        
        if existing_customer and not merchant_id:
            return jsonify({'error': 'Merchant ID required for existing customers'}), 400
        
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404

        weight_column = mapping_config.get('Weight')
        weight_unit = (mapping_config.get('Weight Unit') or '').lower()
        if weight_column and not weight_unit:
            return jsonify({'error': 'Weight unit required when Weight is mapped'}), 400
        if weight_unit and weight_unit not in {'oz', 'lb', 'kg'}:
            return jsonify({'error': 'Invalid weight unit'}), 400
        detected_weight_unit = weight_unit
        
        # Validate required mappings
        structure = data.get('structure', 'zone')
        required_fields = required_fields_for_structure(structure)
        missing = [f for f in required_fields if f not in mapping_config or not mapping_config[f]]
        
        if missing:
            return jsonify({
                'error': 'Missing required field mappings',
                'missing': missing
            }), 400
        
        # Normalize and save CSV
        raw_csv_path = job_dir / 'raw_invoice.csv'
        df = pd.read_csv(raw_csv_path)
        for col in df.select_dtypes(include=['category']).columns:
            df[col] = df[col].astype(object)
        
        # Find zone column name for later use
        zone_col_name = None
        structure = data.get('structure', 'zone')
        if structure == 'zone':
            for col in df.columns:
                if 'zone' in col.lower():
                    zone_col_name = col
                    break
        
        # Save mapping config
        config = {
            'merchant_name': merchant_name,
            'merchant_id': merchant_id,
            'existing_customer': existing_customer,
            'se_ae_on_deal': se_ae_on_deal,
            'origin_zip': origin_zip,
            'annual_orders': annual_orders,
            'raw_invoice_file': raw_invoice_file,
            'amazon_eligible': amazon_eligible,
            'uniuni_eligible': uniuni_eligible,
            'flow_type': data.get('flow_type', 'rate_card_plus_deal_sizing'),
            'mapping': mapping_config,
            'structure': structure,
            'zone_column': zone_col_name
        }
        
        # Save config
        with open(job_dir / 'mapping.json', 'w') as f:
            json.dump(config, f)
        
        # Record mapping phase timestamp
        write_progress(job_dir, 'mapping', True)
        
        # Apply mapping
        normalized_data = {}
        for std_field, invoice_col in mapping_config.items():
            if invoice_col and invoice_col in df.columns:
                normalized_data[std_field] = df[invoice_col]
        
        # Zone will be included if mapped by user
        # If zone-based and zone column exists but wasn't mapped, we'll handle it in generation
        
        # Create normalized DataFrame
        normalized_df = pd.DataFrame(normalized_data)
        for col in normalized_df.select_dtypes(include=['category']).columns:
            normalized_df[col] = normalized_df[col].astype(object)
        if 'Weight (oz)' in normalized_df.columns and 'Weight' not in normalized_df.columns:
            normalized_df['Weight'] = normalized_df['Weight (oz)']
            if not detected_weight_unit:
                detected_weight_unit = 'oz'

        # Derived/computed fields
        country_series = None
        for col in df.columns:
            if 'country' in str(col).lower():
                country_series = df[col]
                break

        if country_series is not None:
            country_str = country_series.fillna("").astype(str).str.strip()
            country_upper = country_str.str.upper()
            is_two_letter = country_upper.str.len().eq(2) & country_upper.str.isalpha()

            mapped_codes = country_upper.map(COUNTRY_NAME_TO_CODE).fillna("")
            normalized_df['TWO_LETTER_COUNTRY_CODE'] = np.where(
                is_two_letter, country_upper, mapped_codes
            )
            normalized_df['FULL_COUNTRY_NAME'] = np.where(
                is_two_letter,
                country_upper.map(CODE_TO_COUNTRY_NAME).fillna(""),
                country_str
            )
        else:
            normalized_df['TWO_LETTER_COUNTRY_CODE'] = ""
            normalized_df['FULL_COUNTRY_NAME'] = ""

        two_letter = normalized_df['TWO_LETTER_COUNTRY_CODE'].fillna("").astype(str)
        full_name = normalized_df['FULL_COUNTRY_NAME'].fillna("").astype(str)
        mapped_from_name = full_name.str.upper().map(COUNTRY_NAME_TO_CODE).fillna("")
        zip_series = normalized_df['Zip'] if 'Zip' in normalized_df.columns else pd.Series([""] * len(normalized_df))
        zip_match = zip_series.fillna("").astype(str).str.extract(r'(\d{5})', expand=False)
        has_zip = zip_match.notna()

        calculated_code = two_letter.mask(two_letter.eq(""), mapped_from_name)
        calculated_code = calculated_code.mask(calculated_code.eq(""), np.where(has_zip, "US", ""))
        normalized_df['CALCULATED_TWO_LETTER_COUNTRY_CODE'] = calculated_code

        if structure == 'zip' and 'Zip' in normalized_df.columns:
            origin_zip3 = _zip3_from_zip(origin_zip)
            zone_map = _fetch_usps_zone_chart(origin_zip3)
            if zone_map:
                dest_zip3 = normalized_df['Zip'].apply(_zip3_from_zip)
                zone_values = dest_zip3.map(zone_map)
                normalized_df['Zone'] = pd.to_numeric(zone_values, errors='coerce')
            else:
                return jsonify({
                    'error': (
                        'Unable to fetch USPS zone chart for origin ZIP. '
                        'Check network access or prefill the cache at '
                        f'{USPS_ZONE_CACHE_PATH}.'
                    )
                }), 500

        shipping_service_series = (
            normalized_df['Shipping Service']
            if 'Shipping Service' in normalized_df.columns
            else pd.Series([""] * len(normalized_df))
        )
        cleaned_service = shipping_service_series.fillna("").astype(str)
        cleaned_service = cleaned_service.str.replace('Â', '', regex=False).str.replace('®', '', regex=False)
        cleaned_service = cleaned_service.str.split(r'\s*[-–—]\s*', n=1, expand=True)[0]
        cleaned_service = cleaned_service.str.replace(r'[^\w\s]', ' ', regex=True)
        cleaned_service = cleaned_service.str.replace(r'\s+', ' ', regex=True).str.upper().str.strip()
        normalized_df['CLEANED_SHIPPING_SERVICE'] = cleaned_service

        priority = pd.Series([""] * len(normalized_df))
        non_empty = cleaned_service.ne("")
        priority = priority.mask(non_empty & cleaned_service.str.contains('GROUND', regex=False), 'GROUND')
        air_mask = non_empty & cleaned_service.str.contains(r'2ND DAY|2 DAY|2DAY', regex=True)
        priority = priority.mask(priority.eq("") & air_mask, 'AIR')
        exp_mask = non_empty & cleaned_service.str.contains('EXPEDITED', regex=False)
        priority = priority.mask(priority.eq("") & exp_mask, 'EXPEDITED')
        priority = priority.mask((priority.eq("")) & non_empty, 'OTHER')
        normalized_df['SHIPPING_PRIORITY'] = priority

        weight_series = None
        if 'Weight' in normalized_df.columns:
            weight_series = pd.to_numeric(normalized_df['Weight'], errors='coerce')
        else:
            weight_series = pd.Series([np.nan] * len(normalized_df))

        if detected_weight_unit == 'oz':
            normalized_df['WEIGHT_IN_OZ'] = weight_series.round(4)
            normalized_df['WEIGHT_IN_LBS'] = (weight_series / 16).round(4)
        elif detected_weight_unit == 'lb':
            normalized_df['WEIGHT_IN_LBS'] = weight_series.round(4)
            normalized_df['WEIGHT_IN_OZ'] = pd.Series([np.nan] * len(normalized_df))
        elif detected_weight_unit == 'kg':
            normalized_df['WEIGHT_IN_LBS'] = (weight_series * 2.2046226218).round(4)
            normalized_df['WEIGHT_IN_OZ'] = (weight_series * 35.27396195).round(4)
        else:
            normalized_df['WEIGHT_IN_LBS'] = pd.Series([np.nan] * len(normalized_df))
            normalized_df['WEIGHT_IN_OZ'] = pd.Series([np.nan] * len(normalized_df))

        def _numeric_series(series_name):
            if series_name in normalized_df.columns:
                return pd.to_numeric(normalized_df[series_name], errors='coerce')
            return pd.Series([None] * len(normalized_df))

        length = _numeric_series('Package Length')
        width = _numeric_series('Package Width')
        height = _numeric_series('Package Height')
        volume = length * width * height
        normalized_df['PACKAGE_DIMENSION_VOLUME'] = volume

        size_bins = [-float('inf'), SMALL_MAX_VOLUME, MEDIUM_MAX_VOLUME, float('inf')]
        size_labels = ['SMALL', 'MEDIUM', 'LARGE']
        size_class = pd.Series(
            pd.cut(volume, bins=size_bins, labels=size_labels, right=False),
            index=volume.index
        ).astype(object)
        normalized_df['PACKAGE_SIZE_STATUS'] = size_class.where(volume.notna(), "")

        weight_lbs = normalized_df['WEIGHT_IN_LBS']
        weight_bins = [-float('inf'), 1, 5, 10, float('inf')]
        weight_labels = ['<1', '1-5', '5-10', '10+']
        weight_class = pd.Series(
            pd.cut(weight_lbs, bins=weight_bins, labels=weight_labels, right=False),
            index=weight_lbs.index
        ).astype(object)
        normalized_df['WEIGHT_CLASSIFICATION'] = weight_class.where(weight_lbs.notna(), "")

        origin_zip_value = extract_origin_zip(origin_zip)
        if structure == 'zip':
            normalized_df['ORIGIN_ZIP_CODE'] = [origin_zip_value] * len(normalized_df)
        else:
            normalized_df['ORIGIN_ZIP_CODE'] = ["" for _ in range(len(normalized_df))]
        
        # Save normalized CSV
        normalized_csv_path = job_dir / 'normalized.csv'
        normalized_df.to_csv(normalized_csv_path, index=False)
        
        return jsonify({'success': True})
    except Exception as e:
        app.logger.exception('Error saving mapping config')
        return jsonify({'error': str(e)}), 500

@app.route('/api/service-levels', methods=['POST'])
def service_levels():
    """Save selected service levels"""
    try:
        data = request.json
        job_id = data.get('job_id')
        selected_services = data.get('selected_services', [])
        
        if not job_id:
            return jsonify({'error': 'job_id required'}), 400
        
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404
        
        # Save service levels
        with open(job_dir / 'service_levels.json', 'w') as f:
            json.dump({'selected_services': selected_services}, f)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/service-levels/<job_id>', methods=['GET'])
def get_service_levels(job_id):
    """Return available merchant service levels for a job"""
    try:
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404

        mapping_file = job_dir / 'mapping.json'
        if not mapping_file.exists():
            return jsonify({'error': 'Mapping not found'}), 404

        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)

        raw_df = pd.read_csv(job_dir / 'raw_invoice.csv')
        available_services = available_merchant_services(raw_df, mapping_config)

        return jsonify({'available_services': available_services})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/amazon-eligibility', methods=['POST'])
def amazon_eligibility():
    """Return Amazon eligibility based on origin ZIP."""
    try:
        data = request.json or {}
        origin_zip = data.get('origin_zip', '')
        annual_orders = data.get('annual_orders')
        eligibility = compute_eligibility(origin_zip, annual_orders)
        payload = {
            'zip_eligible': eligibility['zip_eligible_amazon'],
            'amazon_volume_avg': eligibility['amazon_volume_avg'],
            'amazon_volume_eligible': eligibility['amazon_volume_eligible'],
            'uniuni_volume_avg': eligibility['uniuni_volume_avg'],
            'uniuni_volume_eligible': eligibility['uniuni_volume_eligible'],
            'amazon_eligible_final': eligibility['amazon_eligible_final'],
            'uniuni_eligible_final': eligibility['uniuni_eligible_final'],
            'eligible': eligibility['amazon_eligible_final']
        }
        return jsonify(payload)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/uniuni-eligibility', methods=['POST'])
def uniuni_eligibility():
    """Return UniUni eligibility based on origin ZIP."""
    try:
        data = request.json or {}
        origin_zip = data.get('origin_zip', '')
        annual_orders = data.get('annual_orders')
        eligibility = compute_eligibility(origin_zip, annual_orders)
        payload = {
            'zip_eligible': eligibility['zip_eligible_uniuni'],
            'amazon_volume_avg': eligibility['amazon_volume_avg'],
            'amazon_volume_eligible': eligibility['amazon_volume_eligible'],
            'uniuni_volume_avg': eligibility['uniuni_volume_avg'],
            'uniuni_volume_eligible': eligibility['uniuni_volume_eligible'],
            'amazon_eligible_final': eligibility['amazon_eligible_final'],
            'uniuni_eligible_final': eligibility['uniuni_eligible_final'],
            'eligible': eligibility['uniuni_eligible_final']
        }
        return jsonify(payload)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/merchant-pricing/<job_id>', methods=['GET', 'POST'])
def merchant_pricing(job_id):
    """Get or save merchant pricing selections."""
    try:
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404

        pricing_file = job_dir / 'merchant_pricing.json'

        if request.method == 'POST':
            data = request.json or {}
            excluded = data.get('excluded_carriers', [])
            included = data.get('included_services', [])
            if not isinstance(excluded, list) or not isinstance(included, list):
                return jsonify({'error': 'Invalid payload'}), 400
            payload = {
                'excluded_carriers': excluded,
                'included_services': included
            }
            with open(pricing_file, 'w') as f:
                json.dump(payload, f)
            return jsonify({'success': True})

        saved = {'excluded_carriers': [], 'included_services': []}
        has_saved = False
        if pricing_file.exists():
            with open(pricing_file, 'r') as f:
                saved = json.load(f)
            has_saved = True

        mapping_file = job_dir / 'mapping.json'
        if not mapping_file.exists():
            return jsonify({'error': 'Mapping not found'}), 404

        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)

        raw_df = pd.read_csv(job_dir / 'raw_invoice.csv')
        available_services = available_merchant_services(raw_df, mapping_config)
        available_carriers = available_merchant_carriers(raw_df, mapping_config)

        # Sync eligibility with merchant pricing
        eligibility = compute_eligibility(
            mapping_config.get('origin_zip'),
            mapping_config.get('annual_orders'),
            mapping_config=mapping_config
        )
        
        excluded = saved.get('excluded_carriers', [])
        
        # Auto-remove from excluded if eligible, auto-add if not
        if eligibility['amazon_eligible_final']:
            if 'Amazon' in excluded:
                excluded = [c for c in excluded if c != 'Amazon']
        else:
            if 'Amazon' not in excluded:
                excluded.append('Amazon')
                
        if eligibility['uniuni_eligible_final']:
            if 'UniUni' in excluded:
                excluded = [c for c in excluded if c != 'UniUni']
        else:
            if 'UniUni' not in excluded:
                excluded.append('UniUni')
        
        # Save back if changed
        if excluded != saved.get('excluded_carriers', []):
            saved['excluded_carriers'] = excluded
            with open(pricing_file, 'w') as f:
                json.dump(saved, f)

        included_services = saved.get('included_services', [])
        if not has_saved and not included_services:
            included_services = default_included_services(available_services)
        else:
            available_by_norm = {normalize_service_name(s): s for s in available_services}
            saved_norm = {normalize_service_name(s) for s in (included_services or [])}
            default_norm = {
                normalize_service_name(s)
                for s in default_included_services(available_services)
            }
            merged = []
            for norm in saved_norm:
                if norm in available_by_norm:
                    merged.append(available_by_norm[norm])
            added = False
            for service in available_services:
                norm = normalize_service_name(service)
                if norm not in saved_norm and norm in default_norm:
                    merged.append(service)
                    added = True
            if merged:
                included_services = merged
            if added and pricing_file.exists():
                payload = {
                    'excluded_carriers': excluded,
                    'included_services': included_services
                }
                with open(pricing_file, 'w') as f:
                    json.dump(payload, f)

        eligibility = compute_eligibility(
            mapping_config.get('origin_zip'),
            mapping_config.get('annual_orders'),
            mapping_config=mapping_config
        )
        if not eligibility['amazon_eligible_final'] and 'Amazon' not in excluded:
            excluded.append('Amazon')
        if not eligibility['uniuni_eligible_final'] and 'UniUni' not in excluded:
            excluded.append('UniUni')

        if available_carriers:
            excluded = [c for c in excluded if c in available_carriers]

        return jsonify({
            'carriers': available_carriers,
            'service_levels': available_services,
            'excluded_carriers': excluded,
            'included_services': included_services
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/redo-carriers/<job_id>', methods=['GET', 'POST'])
def redo_carriers(job_id):
    """Get or save redo carrier selections"""
    try:
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404

        mapping_file = job_dir / 'mapping.json'
        if not mapping_file.exists():
            return jsonify({'error': 'Mapping not found'}), 404

        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)

        eligibility = compute_eligibility(
            mapping_config.get('origin_zip'),
            mapping_config.get('annual_orders'),
            mapping_config=mapping_config
        )
        app.logger.info(f"Redo carriers eligibility check - annual_orders={mapping_config.get('annual_orders')}, amazon_eligible={eligibility['amazon_eligible_final']}, uniuni_eligible={eligibility['uniuni_eligible_final']}, amazon_volume_avg={eligibility.get('amazon_volume_avg')}")

        if request.method == 'POST':
            data = request.json or {}
            selected = data.get('selected_carriers', [])
            selected = [c for c in selected if c in REDO_CARRIERS]
            if not eligibility['amazon_eligible_final']:
                selected = [c for c in selected if c != 'Amazon']
            if not eligibility['uniuni_eligible_final']:
                selected = [c for c in selected if c != 'UniUni']

            with open(job_dir / 'redo_carriers.json', 'w') as f:
                json.dump({'selected_carriers': selected}, f)
            return jsonify({'success': True})

        raw_df = pd.read_csv(job_dir / 'raw_invoice.csv')
        detected_redo = detect_redo_carriers(raw_df, mapping_config)
        available = list(REDO_FORCED_ON)
        for carrier in detected_redo:
            if carrier not in available:
                available.append(carrier)
        for fallback in ('FedEx',):
            if fallback not in available:
                available.append(fallback)
        
        # Ensure UniUni and Amazon are always available in the selection list if eligible
        if eligibility['amazon_eligible_final'] and 'Amazon' not in available:
            available.append('Amazon')
        if eligibility['uniuni_eligible_final'] and 'UniUni' not in available:
            available.append('UniUni')

        selected = list(REDO_FORCED_ON)
        if eligibility['amazon_eligible_final'] and 'Amazon' not in selected:
            selected.append('Amazon')
        if eligibility['uniuni_eligible_final'] and 'UniUni' not in selected:
            selected.append('UniUni')

        # Sync redo_carriers.json file if it exists to ensure persistence
        redo_file = job_dir / 'redo_carriers.json'
        if redo_file.exists():
            try:
                with open(redo_file, 'r') as f:
                    saved_redo = json.load(f)
                rs = saved_redo.get('selected_carriers', [])
                changed = False
                if eligibility['amazon_eligible_final'] and 'Amazon' not in rs:
                    rs.append('Amazon')
                    changed = True
                if eligibility['uniuni_eligible_final'] and 'UniUni' not in rs:
                    rs.append('UniUni')
                    changed = True
                if changed:
                    with open(redo_file, 'w') as f:
                        json.dump({'selected_carriers': rs}, f)
                # Use saved selections (without DHL)
                selected = [c for c in rs if c in REDO_CARRIERS]
                for forced in REDO_FORCED_ON:
                    if forced not in selected:
                        selected.append(forced)
            except:
                pass

        # Filter out any invalid carriers from selected
        selected = [c for c in selected if c in REDO_CARRIERS]
        app.logger.info(f"Redo carriers response - available={available}, selected={selected}")

        return jsonify({
            'detected_carriers': available,
            'selected_carriers': selected,
            'default_selected': list(REDO_FORCED_ON)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate', methods=['POST'])
def generate():
    """Start rate card generation - generates Excel synchronously for accurate calculations"""
    try:
        data = request.json
        job_id = data.get('job_id')
        
        if not job_id:
            return jsonify({'error': 'job_id required'}), 400
        
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404
        
        # Load configs
        with open(job_dir / 'mapping.json', 'r') as f:
            mapping_config = json.load(f)
        
        merchant_pricing = {'excluded_carriers': [], 'included_services': []}
        pricing_file = job_dir / 'merchant_pricing.json'
        if pricing_file.exists():
            with open(pricing_file, 'r') as f:
                merchant_pricing = json.load(f)
        
        # Initialize progress
        progress_file = job_dir / 'progress.json'
        normalized_csv = job_dir / 'normalized.csv'
        # Base estimate is ~50s based on actual measurements:
        # - Template parsing ~23s
        # - Data processing ~3s
        # - Workbook write ~24s
        estimated_seconds = 50
        if normalized_csv.exists():
            try:
                with open(normalized_csv, newline='') as f:
                    rows = sum(1 for _ in f) - 1
                rows = max(rows, 0)
                # Larger files take slightly longer: base 50s + 0.01s per row over 1000
                if rows > 1000:
                    estimated_seconds = min(90, 50 + int((rows - 1000) * 0.01))
            except Exception:
                pass

        # Merge with existing progress to preserve upload/mapping phase timestamps
        existing_progress = {}
        if progress_file.exists():
            try:
                with open(progress_file, 'r') as f:
                    existing_progress = json.load(f)
            except Exception:
                pass
        
        # Update with generation start info, preserving phase_timestamps
        existing_progress['eta_seconds'] = estimated_seconds
        if 'started_at' not in existing_progress:
            existing_progress['started_at'] = datetime.now(timezone.utc).isoformat()
        if 'phase_timestamps' not in existing_progress:
            existing_progress['phase_timestamps'] = {}
        existing_progress['phase_timestamps']['generation_start'] = datetime.now(timezone.utc).isoformat()
        
        with open(progress_file, 'w') as f:
            json.dump(existing_progress, f)
        
        # Generate Excel and dashboard metrics in background thread
        def run_generation():
            try:
                generate_rate_card(job_dir, mapping_config, merchant_pricing)
            except Exception as e:
                write_error(job_dir, f'Generation failed: {str(e)}')

        if app.config.get('TESTING'):
            run_generation()
            return jsonify({'success': True, 'status': 'completed'})

        thread = threading.Thread(target=run_generation, daemon=True)
        thread.start()
        
        return jsonify({'success': True, 'status': 'started'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def write_progress(job_dir, step, value=True):
    """Write progress update with timestamp for phase tracking"""
    progress_file = job_dir / 'progress.json'
    progress = {}
    if progress_file.exists():
        with open(progress_file, 'r') as f:
            progress = json.load(f)
    progress[step] = value
    if 'phase_timestamps' not in progress:
        progress['phase_timestamps'] = {}
    progress['phase_timestamps'][step] = datetime.now(timezone.utc).isoformat()
    with open(progress_file, 'w') as f:
        json.dump(progress, f)

def _load_progress_stats():
    """Load historical phase timings to estimate ETA."""
    stats = {'phases': {}}
    try:
        with _PROGRESS_STATS_LOCK:
            if _PROGRESS_STATS_FILE.exists():
                with open(_PROGRESS_STATS_FILE, 'r') as f:
                    stats = json.load(f)
    except Exception:
        stats = {'phases': {}}
    if 'phases' not in stats:
        stats['phases'] = {}
    return stats

def _write_progress_stats(stats):
    """Persist updated timing stats."""
    stats.setdefault('phases', {})
    stats['last_updated'] = datetime.now(timezone.utc).isoformat()
    with _PROGRESS_STATS_LOCK:
        with open(_PROGRESS_STATS_FILE, 'w') as f:
            json.dump(stats, f)

def _compute_phase_durations(timestamps):
    """Compute durations between generator phases."""
    durations = {}
    base_ts = timestamps.get('generation_start')
    if not base_ts and timestamps:
        base_ts = min(timestamps.values())
    if not base_ts:
        return durations
    try:
        prev_time = datetime.fromisoformat(base_ts)
    except Exception:
        return durations

    for phase in PROGRESS_PHASE_SEQUENCE:
        phase_ts = timestamps.get(phase)
        if not phase_ts:
            break
        try:
            current_time = datetime.fromisoformat(phase_ts)
        except Exception:
            break
        durations[phase] = max(0.0, (current_time - prev_time).total_seconds())
        prev_time = current_time
    return durations

def _record_progress_stats(job_dir):
    """Capture finished phase durations for future ETA estimates."""
    progress_file = job_dir / 'progress.json'
    if not progress_file.exists():
        return
    try:
        with open(progress_file, 'r') as f:
            progress = json.load(f)
    except Exception:
        return
    timestamps = progress.get('phase_timestamps', {})
    durations = _compute_phase_durations(timestamps)
    if not durations:
        return
    stats = _load_progress_stats()
    phases_data = stats.setdefault('phases', {})
    for phase, duration in durations.items():
        entry = phases_data.setdefault(phase, {'count': 0, 'total_seconds': 0.0})
        entry['count'] += 1
        entry['total_seconds'] += duration
    _write_progress_stats(stats)

def _phase_estimates_from_stats(stats):
    """Translate stored stats into per-phase averages."""
    estimates = {}
    phases_data = stats.get('phases', {})
    for phase in PROGRESS_PHASE_SEQUENCE:
        entry = phases_data.get(phase, {})
        count = entry.get('count', 0)
        total = entry.get('total_seconds', 0.0)
        if count > 0:
            estimates[phase] = total / count
        else:
            estimates[phase] = DEFAULT_PHASE_ESTIMATES.get(phase, 5.0)
    return estimates

def _estimate_eta_from_stats(progress):
    """Estimate remaining seconds using historical phase averages."""
    stats = _load_progress_stats()
    estimates = _phase_estimates_from_stats(stats)
    total_estimate = sum(estimates.values())
    if total_estimate <= 0:
        return None
    started_at = progress.get('started_at')
    if not started_at:
        return round(total_estimate)
    try:
        started_dt = datetime.fromisoformat(started_at)
    except Exception:
        return round(total_estimate)
    elapsed = (datetime.now(timezone.utc) - started_dt).total_seconds()
    remaining = max(0.0, total_estimate - elapsed)
    return int(round(remaining))

def _estimate_eta_from_progress(progress):
    """Fallback ETA based on earlier row-based heuristic."""
    eta_total = progress.get('eta_seconds')
    if eta_total is None:
        return None
    started_at = progress.get('started_at')
    if started_at:
        try:
            started_dt = datetime.fromisoformat(started_at)
            elapsed = (datetime.now(timezone.utc) - started_dt).total_seconds()
            return max(0, int(round(eta_total - elapsed)))
        except Exception:
            pass
    return int(round(eta_total))

def _calculate_eta(progress):
    eta = _estimate_eta_from_stats(progress)
    if eta is not None:
        return eta, 'history'
    fallback = _estimate_eta_from_progress(progress)
    if fallback is not None:
        return fallback, 'row_estimate'
    return None, None

def write_error(job_dir, message):
    """Write error to progress file."""
    progress_file = job_dir / 'progress.json'
    progress = {}
    if progress_file.exists():
        with open(progress_file, 'r') as f:
            progress = json.load(f)
    progress['error'] = message
    with open(progress_file, 'w') as f:
        json.dump(progress, f)

def generate_rate_card(job_dir, mapping_config, merchant_pricing):
    """Generate the rate card Excel file"""
    import logging
    gen_start = time.time()
    
    # Save output path first
    merchant_name = mapping_config.get('merchant_name', 'Merchant')
    output_filename = f"{merchant_name} - Rate Card.xlsx"
    output_path = job_dir / output_filename
    
    # Step 1: Normalize/qualify data (keep 'normalize' for compatibility)
    write_progress(job_dir, 'normalize', True)
    write_progress(job_dir, 'qualification', True)
    
    # Load template from cache (fast BytesIO copy instead of 25s file read)
    load_start = time.time()
    template_buffer = _get_cached_template()
    wb = openpyxl.load_workbook(template_buffer, keep_vba=False, data_only=False)
    logging.info(f"Template load time: {time.time() - load_start:.1f}s")
    if 'Raw Data' not in wb.sheetnames:
        raise ValueError("Template must contain 'Raw Data' sheet")
    ws = wb['Raw Data']
    
    headers = [cell.value for cell in ws[1]]
    header_to_col = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        header_str = str(header).strip()
        if header_str:
            header_to_col[header_str] = idx + 1
    
    redo_config = {}
    redo_file = job_dir / 'redo_carriers.json'
    if redo_file.exists():
        with open(redo_file, 'r') as f:
            redo_config = json.load(f)

    # Load normalized data
    normalized_df = pd.read_csv(job_dir / 'normalized.csv')

    if not normalized_df.empty:
        debug_cols = [
            'WEIGHT_IN_LBS',
            'TWO_LETTER_COUNTRY_CODE',
            'CALCULATED_TWO_LETTER_COUNTRY_CODE',
            'FULL_COUNTRY_NAME',
            'CLEANED_SHIPPING_SERVICE',
            'SHIPPING_PRIORITY',
            'PACKAGE_DIMENSION_VOLUME',
            'PACKAGE_SIZE_STATUS',
            'WEIGHT_CLASSIFICATION',
            'ORIGIN_ZIP_CODE'
        ]
        sample = {}
        for col in debug_cols:
            if col in normalized_df.columns:
                sample[col] = normalized_df.iloc[0].get(col)
        if sample:
            app.logger.info("Computed columns sample: %s", sample)
    
    # Step 2: Write into template
    write_progress(job_dir, 'write_template', True)
    
    start_row = 2
    origin_zip_value = mapping_config.get('origin_zip')
    
    # Pre-calculate zone fallback values if needed
    zone_values_fallback = None
    if 'Zone' in normalized_df.columns:
        zone_values_fallback = normalized_df['Zone'].tolist()
    
    # Map standard fields to Excel columns
    field_to_excel = {
        'Order Number': 'ORDER_NUMBER',
        'Order Date': 'DATE',
        'Zip': 'DESTINATION_ZIP_CODE',
        'Weight': 'WEIGHT_IN_OZ',
        'WEIGHT_IN_LBS': 'WEIGHT_IN_LBS',
        'Shipping Carrier': 'SHIPPING_CARRIER',
        'CLEANED_SHIPPING_SERVICE': 'CLEANED_SHIPPING_SERVICE',
        'Package Height': 'PACKAGE_HEIGHT',
        'Package Width': 'PACKAGE_WIDTH',
        'Package Length': 'PACKAGE_LENGTH',
        'PACKAGE_DIMENSION_VOLUME': 'PACKAGE_DIMENSION_VOLUME',
        'ORIGIN_ZIP_CODE': 'ORIGIN_ZIP_CODE',
        'Shipping Service': 'SHIPPING_SERVICE',
        'Label Cost': 'LABEL_COST',
        'Zone': 'ZONE'
    }

    numeric_cols = {
        'WEIGHT_IN_OZ', 'WEIGHT_IN_LBS', 'PACKAGE_HEIGHT', 'PACKAGE_WIDTH',
        'PACKAGE_LENGTH', 'PACKAGE_DIMENSION_VOLUME', 'LABEL_COST'
    }

    # Get merchant pricing selections
    excluded_carriers = merchant_pricing.get('excluded_carriers', [])
    included_services = merchant_pricing.get('included_services', [])
    normalized_selected = {normalize_service_name(s) for s in included_services}
    normalized_excluded = {normalize_merchant_carrier(c) for c in excluded_carriers}

    # Speed optimization: Extract data to lists first
    column_data = {
        field: normalized_df[field].tolist()
        for field in field_to_excel.keys()
        if field in normalized_df.columns
    }
    
    # Pre-calculate flags
    service_series = normalized_df.get('Shipping Service')
    if service_series is None:
        service_series = pd.Series([""] * len(normalized_df))
    carrier_series = normalized_df.get('Shipping Carrier')
    if carrier_series is None:
        carrier_series = pd.Series([""] * len(normalized_df))
    service_norm = service_series.fillna("").astype(str).apply(normalize_service_name)
    carrier_norm = carrier_series.fillna("").astype(str).apply(normalize_merchant_carrier)
    carrier_allowed = ~carrier_norm.isin(normalized_excluded)
    qualified_flags = (service_norm.isin(normalized_selected) & carrier_allowed).tolist()

    write_cols = set()
    write_fields = []
    for std_field, excel_col in field_to_excel.items():
        col_idx = header_to_col.get(excel_col)
        if col_idx:
            write_cols.add(col_idx)
            if std_field in column_data:
                write_fields.append((std_field, excel_col, col_idx))

    # Identify formula columns (AI-AN, 1-indexed)
    formula_cols = set(range(35, 41))

    # Optimize: Pre-calculate lookups
    write_fields_prepared = []
    for std_field, excel_col, col_idx in write_fields:
        if col_idx not in formula_cols:
            write_fields_prepared.append((std_field, excel_col, col_idx))

    # Single pass for writing data
    total_rows = len(normalized_df)
    
    # Pre-calculate formula values (they are constant formulas)
    formula_values = {}
    for col_idx in formula_cols:
        source_cell = ws.cell(2, col_idx)
        if source_cell.value and str(source_cell.value).startswith('='):
            formula_values[col_idx] = source_cell.value

    # Identify other columns
    m_id = mapping_config.get('merchant_id')
    m_col = header_to_col.get('MERCHANT_ID')
    q_col = header_to_col.get('QUALIFIED')
    z_col = header_to_col.get('ZONE')

    for idx in range(total_rows):
        excel_row = start_row + idx
        
        # Write mapped fields
        for std_field, excel_col, col_idx in write_fields_prepared:
            value = column_data[std_field][idx]
            
            if pd.isna(value):
                value = None
            elif std_field == 'ORIGIN_ZIP_CODE' and value is None:
                value = origin_zip_value
            elif std_field == 'Order Date' and excel_col == 'DATE':
                try:
                    if isinstance(value, str):
                        value = pd.to_datetime(value)
                    if hasattr(value, 'to_pydatetime'):
                        value = value.to_pydatetime()
                except:
                    pass
            elif std_field == 'Zip' and excel_col == 'DESTINATION_ZIP_CODE':
                zip_str = str(value).strip()
                zip_match = re.search(r'\d{5}', zip_str)
                value = int(zip_match.group()) if zip_match else None
            elif excel_col in numeric_cols:
                try:
                    value = float(value) if value else None
                except:
                    value = None
            
            if value is not None:
                ws.cell(row=excel_row, column=col_idx, value=value)
        
        # Write fallback ZONE
        if zone_values_fallback is not None and z_col and z_col not in formula_cols:
            zone_val = zone_values_fallback[idx] if idx < len(zone_values_fallback) else None
            if zone_val is not None and not pd.isna(zone_val):
                try:
                    ws.cell(row=excel_row, column=z_col, value=int(float(zone_val)))
                except:
                    pass
        
        # Write MERCHANT_ID
        if m_id and m_col and m_col not in formula_cols:
            ws.cell(row=excel_row, column=m_col, value=m_id)
        
        # Write QUALIFIED
        if q_col and q_col not in formula_cols:
            is_qualified = qualified_flags[idx] if idx < len(qualified_flags) else False
            ws.cell(row=excel_row, column=q_col, value=is_qualified)

        # Write Formulas directly in the loop to avoid second pass
        for col_idx, formula in formula_values.items():
            ws.cell(row=excel_row, column=col_idx, value=formula)

    # Update Pricing & Summary redo carrier selections
    if 'Pricing & Summary' in wb.sheetnames:
        selected_redo = redo_config.get('selected_carriers', [])
        selected_services = merchant_pricing.get('included_services', [])
        excluded_carriers = merchant_pricing.get('excluded_carriers', [])
        summary_ws = wb['Pricing & Summary']
        # Populate Annual Orders to avoid blank Deal Info calculations.
        annual_orders_value = mapping_config.get('annual_orders')
        try:
            annual_orders_value = int(float(annual_orders_value)) if annual_orders_value else None
        except Exception:
            annual_orders_value = None
        summary_ws['C9'] = annual_orders_value if annual_orders_value is not None else 13968
        if origin_zip_value is not None:
            summary_ws['C29'] = origin_zip_value
        pct_off, dollar_off = _usps_market_discount_values(mapping_config)
        summary_ws['C19'] = pct_off
        summary_ws['C20'] = dollar_off
        update_pricing_summary_redo_carriers(summary_ws, selected_redo)
        update_pricing_summary_merchant_carriers(summary_ws, excluded_carriers)
        update_pricing_summary_merchant_service_levels(
            summary_ws, selected_services, normalized_df
        )

    # Step 3: Save and finalize
    write_progress(job_dir, 'saving', True)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"
    if hasattr(wb, "_calcChain"):
        wb._calcChain = None
    
    # Save the workbook atomically
    temp_file = None
    save_start = time.time()
    try:
        with tempfile.NamedTemporaryFile(delete=False, dir=job_dir, suffix='.xlsx') as tmp:
            temp_file = Path(tmp.name)
        wb.save(temp_file)
        logging.info(f"Workbook save time: {time.time() - save_start:.1f}s")
        if not zipfile.is_zipfile(temp_file):
            raise Exception("Generated file is not a valid XLSX archive")
        os.replace(temp_file, output_path)
    except Exception as e:
        if temp_file and temp_file.exists():
            try:
                temp_file.unlink()
            except Exception:
                pass
        raise Exception(f"Failed to save Excel file: {str(e)}")
    finally:
        wb.close()
    
    # Pre-compute dashboard metrics and save to JSON cache (fast, no Excel re-load needed)
    try:
        _precompute_dashboard_metrics(job_dir, mapping_config, redo_config)
    except Exception as e:
        logging.warning(f"Dashboard metrics pre-computation failed: {e}")
    
    try:
        log_admin_entry(job_dir.name, mapping_config, merchant_pricing, redo_config)
    except Exception:
        pass

    # Mark Excel generation complete and record timings
    write_progress(job_dir, 'excel_complete', True)
    _record_progress_stats(job_dir)
    
    logging.info(f"Total generation time: {time.time() - gen_start:.1f}s")
    return output_path

@app.route('/api/status/<job_id>')
def status(job_id=None):
    """Get job status and file links"""
    if not job_id:
        return jsonify({'error': 'Job ID required'}), 400
    
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return jsonify({'error': 'Job not found'}), 404
    
    # Check progress file
    progress_file = job_dir / 'progress.json'
    progress = {}
    if progress_file.exists():
        with open(progress_file, 'r') as f:
            progress = json.load(f)
    
    # Surface generation error if present
    if 'error' in progress:
        return jsonify({
            'ready': False,
            'error': progress['error'],
            'progress': progress
        })

    # Check if generation is complete
    rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
    is_complete = len(rate_card_files) > 0
    
    if is_complete:
        # Load merchant name for redirect
        merchant_name = 'Merchant'
        mapping_file = job_dir / 'mapping.json'
        if mapping_file.exists():
            with open(mapping_file, 'r') as f:
                config = json.load(f)
                merchant_name = config.get('merchant_name', 'Merchant')
        return jsonify({
            'ready': True,
            'redirect_url': f'/dashboard?job_id={job_id}',
            'progress': progress
        })
    
    elapsed_seconds = None
    if progress.get('started_at'):
        try:
            started_at = datetime.fromisoformat(progress['started_at'])
            elapsed_seconds = (datetime.now(timezone.utc) - started_at).total_seconds()
        except Exception:
            elapsed_seconds = None
    eta_remaining, eta_source = _calculate_eta(progress)

    phase_timestamps = progress.get('phase_timestamps', {})
    phase_durations = {}
    if phase_timestamps:
        try:
            sorted_phases = sorted(phase_timestamps.items(), key=lambda x: x[1])
            # Use first phase as baseline if started_at is missing
            baseline = progress.get('started_at')
            if not baseline and sorted_phases:
                baseline = sorted_phases[0][1]
            if baseline:
                prev_time = datetime.fromisoformat(baseline)
                for phase, ts in sorted_phases:
                    phase_time = datetime.fromisoformat(ts)
                    duration = (phase_time - prev_time).total_seconds()
                    phase_durations[phase] = round(max(0, duration), 2)
                    prev_time = phase_time
        except Exception:
            pass

    response = {
        'ready': False,
        'progress': progress,
        'eta_seconds_remaining': eta_remaining,
        'elapsed_seconds': round(elapsed_seconds, 2) if elapsed_seconds is not None else None,
        'phase_timestamps': phase_timestamps,
        'phase_durations': phase_durations
    }
    if eta_source:
        response['eta_source'] = eta_source
    return jsonify(response)

@app.route('/api/dashboard/<job_id>', methods=['GET', 'POST'])
def dashboard_data(job_id):
    """Return dashboard metrics - fast JSON read only, no Excel loading."""
    try:
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404
        
        mapping_file = job_dir / 'mapping.json'
        mapping_config = {}
        if mapping_file.exists():
            with open(mapping_file, 'r') as f:
                mapping_config = json.load(f)
        
        redo_file = job_dir / 'redo_carriers.json'
        redo_config = {}
        if redo_file.exists():
            with open(redo_file, 'r') as f:
                redo_config = json.load(f)
        
        annual_orders_missing = _annual_orders_missing(mapping_config)
        pct_off, dollar_off = _usps_market_discount_values(mapping_config)
        
        # Require rate card Excel file for accurate calculations
        rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
        if not rate_card_files:
            return jsonify({'error': 'Rate card not found'}), 404
        
        current_hash = _compute_full_cache_hash(job_dir, mapping_config, redo_config)
        
        refresh = request.args.get('refresh') == '1'
        if refresh:
            for cache_file in [_summary_cache_path(job_dir), _cache_path_for_job(job_dir), _carrier_details_cache_path(job_dir)]:
                if cache_file.exists():
                    cache_file.unlink()
        
        cache = _read_dashboard_cache(job_dir)
        
        eligibility = None
        if mapping_config:
            eligibility = compute_eligibility(
                mapping_config.get('origin_zip'),
                mapping_config.get('annual_orders'),
                mapping_config=mapping_config
            )

        available_carriers = list(DASHBOARD_CARRIERS)
        if eligibility and not eligibility['amazon_eligible_final']:
            available_carriers = [c for c in available_carriers if c != 'Amazon']
        if eligibility and not eligibility['uniuni_eligible_final']:
            available_carriers = [c for c in available_carriers if c != 'UniUni']

        redo_selected = redo_config.get('selected_carriers', [])
        if not redo_selected:
            redo_selected = list(REDO_FORCED_ON)
            if eligibility and eligibility['amazon_eligible_final']:
                redo_selected.append('Amazon')
            if eligibility and eligibility['uniuni_eligible_final']:
                redo_selected.append('UniUni')

        selected_set = _dashboard_selected_from_redo(redo_selected)
        selected_dashboard = [c for c in available_carriers if c in selected_set]
        
        if request.method == 'POST':
            data = request.json or {}
            incoming = data.get('selected_carriers', [])
            if isinstance(incoming, list) and incoming:
                selected_dashboard = [c for c in available_carriers if c in incoming]

        show_usps_market_discount = bool(
            (eligibility and (eligibility['amazon_eligible_final'] or eligibility['uniuni_eligible_final']))
            or ('Amazon' in redo_selected or 'UniUni' in redo_selected)
        )
        carrier_percentages = _carrier_distribution(job_dir, mapping_config, available_carriers)
        
        # Only check rate card file hash (first part) for cache validity
        cached_hash = cache.get('source_hash', '')
        cached_file_hash = cached_hash.split(':')[0] if cached_hash else ''
        current_file_hash = current_hash.split(':')[0] if current_hash else ''
        cache_valid = cache.get('ready') and cached_file_hash == current_file_hash
        
        if not cache_valid:
            _precompute_dashboard_metrics(job_dir, mapping_config, redo_config)
            cache = _read_dashboard_cache(job_dir)
        
        carrier_metrics = cache.get('breakdown', {})
        summary_cache = cache.get('summary', {})
        
        include_per_carrier = request.args.get('per_carrier') == '1'
        per_carrier = []
        if include_per_carrier:
            for carrier in selected_dashboard or available_carriers:
                metrics = carrier_metrics.get(carrier, {})
                if annual_orders_missing:
                    metrics = {k: v for k, v in metrics.items() if k not in ('Est. Merchant Annual Savings', 'Est. Redo Deal Size', 'Spread Available')}
                per_carrier.append({'carrier': carrier, 'metrics': metrics})
            
            return jsonify({
                'selected_carriers': selected_dashboard,
                'available_carriers': available_carriers,
                'per_carrier': per_carrier,
                'pending': False,
                'summary_pending': False,
                'per_carrier_count': len(per_carrier),
                'annual_orders_missing': annual_orders_missing,
                'show_usps_market_discount': show_usps_market_discount,
                'usps_market_pct_off': pct_off,
                'usps_market_dollar_off': dollar_off,
                'carrier_percentages': carrier_percentages,
                'per_carrier_total': len(selected_dashboard or available_carriers)
            })
        
        selection_key = _selection_cache_key(selected_dashboard)
        overall_metrics = summary_cache.get(selection_key)
        
        if not overall_metrics:
            overall_metrics = _calculate_metrics_fast(job_dir, selected_dashboard, mapping_config)
            if overall_metrics:
                summary_cache[selection_key] = overall_metrics
                _write_dashboard_cache(job_dir, carrier_metrics, summary_cache, current_hash)
        
        if annual_orders_missing:
            overall_metrics = {k: v for k, v in (overall_metrics or {}).items() 
                            if k not in ('Est. Merchant Annual Savings', 'Est. Redo Deal Size', 'Spread Available')}

        return jsonify({
            'selected_carriers': selected_dashboard,
            'available_carriers': available_carriers,
            'overall': overall_metrics,
            'per_carrier': [],
            'pending': False,
            'summary_pending': False,
            'annual_orders_missing': annual_orders_missing,
            'show_usps_market_discount': show_usps_market_discount,
            'usps_market_pct_off': pct_off,
            'usps_market_dollar_off': dollar_off,
            'carrier_percentages': carrier_percentages
        })
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/<job_id>/rate-card')
def download_rate_card(job_id):
    """Download generated rate card"""
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    if not job_dir.exists():
        return jsonify({'error': 'Job not found'}), 404
    
    rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
    if not rate_card_files:
        return jsonify({'error': 'Rate card not found'}), 404
    
    return send_file(rate_card_files[0], as_attachment=True)

@app.route('/api/annual-orders/<job_id>', methods=['POST'])
def update_annual_orders(job_id):
    """Update annual orders and clear dashboard cache (fast, no Excel regeneration)."""
    try:
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404
        mapping_file = job_dir / 'mapping.json'
        if not mapping_file.exists():
            return jsonify({'error': 'Mapping not found'}), 404
        data = request.json or {}
        annual_orders = data.get('annual_orders')
        annual_orders_value = _parse_numeric_value(annual_orders)
        if annual_orders_value is None:
            return jsonify({'error': 'Invalid annual orders'}), 400
        annual_orders_value = int(annual_orders_value)
        if annual_orders_value <= 0:
            return jsonify({'error': 'Annual orders must be greater than 0'}), 400

        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)
        mapping_config['annual_orders'] = annual_orders_value
        # Clear explicit eligibility overrides so volume-based calculation takes effect
        if 'amazon_eligible' in mapping_config:
            del mapping_config['amazon_eligible']
        if 'uniuni_eligible' in mapping_config:
            del mapping_config['uniuni_eligible']
        if 'uniuni_qualified' in mapping_config:
            del mapping_config['uniuni_qualified']
        if 'uniuni' in mapping_config:
            del mapping_config['uniuni']
        with open(mapping_file, 'w') as f:
            json.dump(mapping_config, f)

        # Compute new eligibility based on updated annual orders (no overrides now)
        eligibility = compute_eligibility(
            mapping_config.get('origin_zip'),
            annual_orders_value,
            mapping_config=mapping_config
        )
        app.logger.info(f"Annual orders update - job_id={job_id}, annual_orders={annual_orders_value}, origin_zip={mapping_config.get('origin_zip')}")
        app.logger.info(f"Eligibility result - amazon_volume_avg={eligibility['amazon_volume_avg']:.2f}, amazon_eligible={eligibility['amazon_eligible_final']}, uniuni_volume_avg={eligibility['uniuni_volume_avg']:.2f}, uniuni_eligible={eligibility['uniuni_eligible_final']}")

        # Sync redo carriers based on new eligibility
        redo_file = job_dir / 'redo_carriers.json'
        if redo_file.exists():
            with open(redo_file, 'r') as f:
                redo_config = json.load(f)
            selected = redo_config.get('selected_carriers', [])
            changed = False
            if eligibility['amazon_eligible_final'] and 'Amazon' not in selected:
                selected.append('Amazon')
                changed = True
            if eligibility['uniuni_eligible_final'] and 'UniUni' not in selected:
                selected.append('UniUni')
                changed = True
            if not eligibility['amazon_eligible_final'] and 'Amazon' in selected:
                selected = [c for c in selected if c != 'Amazon']
                changed = True
            if not eligibility['uniuni_eligible_final'] and 'UniUni' in selected:
                selected = [c for c in selected if c != 'UniUni']
                changed = True
            if changed:
                with open(redo_file, 'w') as f:
                    json.dump({'selected_carriers': selected}, f)

        # Sync merchant pricing excluded carriers
        pricing_file = job_dir / 'merchant_pricing.json'
        if pricing_file.exists():
            with open(pricing_file, 'r') as f:
                merchant_pricing = json.load(f)
            excluded = merchant_pricing.get('excluded_carriers', [])
            changed = False
            if eligibility['amazon_eligible_final'] and 'Amazon' in excluded:
                excluded = [c for c in excluded if c != 'Amazon']
                changed = True
            if eligibility['uniuni_eligible_final'] and 'UniUni' in excluded:
                excluded = [c for c in excluded if c != 'UniUni']
                changed = True
            if not eligibility['amazon_eligible_final'] and 'Amazon' not in excluded:
                excluded.append('Amazon')
                changed = True
            if not eligibility['uniuni_eligible_final'] and 'UniUni' not in excluded:
                excluded.append('UniUni')
                changed = True
            if changed:
                merchant_pricing['excluded_carriers'] = excluded
                with open(pricing_file, 'w') as f:
                    json.dump(merchant_pricing, f)

        # Clear dashboard caches so they recalculate with new annual orders
        summary_cache = _summary_cache_path(job_dir)
        breakdown_cache = _cache_path_for_job(job_dir)
        carrier_details_cache = _carrier_details_cache_path(job_dir)
        if summary_cache.exists():
            summary_cache.unlink()
        if breakdown_cache.exists():
            breakdown_cache.unlink()
        if carrier_details_cache.exists():
            carrier_details_cache.unlink()
        
        # Clear in-memory job caches
        job_prefix = f"{job_dir.name}:"
        with summary_jobs_lock:
            for key in list(summary_jobs.keys()):
                if key.startswith(job_prefix):
                    summary_jobs.pop(key, None)
        with dashboard_jobs_lock:
            for key in list(dashboard_jobs.keys()):
                if key.startswith(job_prefix):
                    dashboard_jobs.pop(key, None)

        # Return success with updated eligibility so frontend can refresh UI
        app.logger.info(f"Annual orders response - amazon_eligible={eligibility['amazon_eligible_final']}, uniuni_eligible={eligibility['uniuni_eligible_final']}")
        return jsonify({
            'success': True,
            'eligibility': eligibility,
            'amazon_eligible': eligibility['amazon_eligible_final'],
            'uniuni_eligible': eligibility['uniuni_eligible_final'],
            'annual_orders': annual_orders_value,
            'amazon_volume_avg': eligibility['amazon_volume_avg'],
            'uniuni_volume_avg': eligibility['uniuni_volume_avg']
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/usps-market-discount/<job_id>', methods=['POST'])
def update_usps_market_discount(job_id):
    """Update USPS Market discount settings and clear cache (fast, no Excel regeneration)."""
    try:
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404
        mapping_file = job_dir / 'mapping.json'
        if not mapping_file.exists():
            return jsonify({'error': 'Mapping not found'}), 404
        data = request.json or {}
        pct_off = data.get('pct_off')
        dollar_off = data.get('dollar_off')
        try:
            pct_off = float(pct_off)
            dollar_off = float(dollar_off)
        except Exception:
            return jsonify({'error': 'Invalid discount values'}), 400
        if pct_off > 1:
            pct_off = pct_off / 100
        if pct_off < 0 or dollar_off < 0:
            return jsonify({'error': 'Discount values must be non-negative'}), 400

        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)
        mapping_config['usps_market_pct_off'] = pct_off
        mapping_config['usps_market_dollar_off'] = dollar_off
        with open(mapping_file, 'w') as f:
            json.dump(mapping_config, f)

        # Clear all dashboard and carrier details caches so they recalculate with new discounts
        summary_cache = _summary_cache_path(job_dir)
        breakdown_cache = _cache_path_for_job(job_dir)
        carrier_details_cache = _carrier_details_cache_path(job_dir)
        if summary_cache.exists():
            summary_cache.unlink()
        if breakdown_cache.exists():
            breakdown_cache.unlink()
        if carrier_details_cache.exists():
            carrier_details_cache.unlink()
        job_prefix = f"{job_dir.name}:"
        with summary_jobs_lock:
            for key in list(summary_jobs.keys()):
                if key.startswith(job_prefix):
                    summary_jobs.pop(key, None)
        with dashboard_jobs_lock:
            for key in list(dashboard_jobs.keys()):
                if key.startswith(job_prefix):
                    dashboard_jobs.pop(key, None)
        with carrier_details_jobs_lock:
            for key in list(carrier_details_jobs.keys()):
                if key.startswith(job_prefix):
                    carrier_details_jobs.pop(key, None)

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/deal-sizing/<job_id>', methods=['GET', 'POST'])
def deal_sizing_data(job_id):
    """Return inputs needed for deal sizing."""
    try:
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        if not job_dir.exists():
            return jsonify({'error': 'Job not found'}), 404
        mapping_file = job_dir / 'mapping.json'
        if not mapping_file.exists():
            return jsonify({'error': 'Mapping not found'}), 404
        with open(mapping_file, 'r') as f:
            mapping_config = json.load(f)
        selected_carriers = []
        if request.method == 'POST':
            payload = request.get_json(silent=True) or {}
            selected_carriers = payload.get('selected_carriers', [])
        else:
            raw_selected = request.args.get('selected_carriers')
            if raw_selected:
                selected_carriers = [c.strip() for c in raw_selected.split(',') if c.strip()]
        if not isinstance(selected_carriers, list):
            selected_carriers = []
        selected_carriers = [c for c in selected_carriers if c in DASHBOARD_CARRIERS]
        annual_orders = mapping_config.get('annual_orders')
        avg_label_cost = _avg_label_cost_from_job(job_dir)
        carrier_details = []
        carrier_details_pending = False
        rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
        if rate_card_files:
            try:
                source_mtime = int(rate_card_files[0].stat().st_mtime)
            except Exception:
                source_mtime = 0
            selection_key = _selection_cache_key(selected_carriers)
            detail_map, carrier_details_pending = _start_carrier_details_cache(
                job_dir,
                source_mtime,
                selection_key,
                selected_carriers,
                mapping_config
            )
            if detail_map is None:
                detail_map = _load_carrier_details(rate_card_files[0], source_mtime) or {}
            if not isinstance(detail_map, dict):
                detail_map = {}
            carrier_details = list(detail_map.values())
        return jsonify({
            'annual_orders': annual_orders,
            'avg_label_cost': avg_label_cost,
            'carrier_details': carrier_details,
            'carrier_details_pending': carrier_details_pending
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/<job_id>/raw-invoice')
def download_raw_invoice(job_id):
    """Download original raw invoice"""
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    raw_csv = job_dir / 'raw_invoice.csv'
    
    if not raw_csv.exists():
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(raw_csv, as_attachment=True, download_name='raw_invoice.csv')

@app.route('/download/<job_id>/normalized')
def download_normalized(job_id):
    """Download normalized CSV"""
    job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
    normalized_csv = job_dir / 'normalized.csv'
    
    if not normalized_csv.exists():
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(normalized_csv, as_attachment=True, download_name='normalized.csv')

def _preload_resources():
    """Preload template data and rate tables when the module is imported."""
    import time as _time
    preload_start = _time.time()
    try:
        template_path = Path('#New Template - Rate Card.xlsx')
        if not template_path.exists():
            template_path = Path('Rate Card Template.xlsx')

        logging.info(f"[PRELOAD] Starting template preload at {datetime.now(timezone.utc).isoformat()}")
        
        t0 = _time.time()
        _get_cached_template()
        logging.info(f"[PRELOAD] Template buffer loaded in {_time.time() - t0:.2f}s")
        
        t0 = _time.time()
        _load_rate_tables(str(template_path))
        logging.info(f"[PRELOAD] Rate tables parsed in {_time.time() - t0:.2f}s")
        
        t0 = _time.time()
        _get_pricing_controls(str(template_path))
        logging.info(f"[PRELOAD] Pricing controls loaded in {_time.time() - t0:.2f}s")
        
        logging.info(f"[PRELOAD] All resources preloaded in {_time.time() - preload_start:.2f}s - workers are warm")
    except Exception as exc:
        logging.warning(f"[PRELOAD] Could not preload template: {exc}")

_preload_resources()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True, use_reloader=False)
