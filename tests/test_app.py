import pytest
import os
import csv
import json
import tempfile
import shutil
import time
from pathlib import Path
import pandas as pd
from app import (
    app,
    normalize_service_name,
    detect_structure,
    suggest_mapping,
    normalize_redo_label,
    update_pricing_summary_redo_carriers,
    update_pricing_summary_merchant_service_levels,
    update_pricing_summary_merchant_carriers
)

@pytest.fixture
def client():
    app.config['TESTING'] = True
    app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()
    with app.test_client() as client:
        yield client
    shutil.rmtree(app.config['UPLOAD_FOLDER'])

@pytest.fixture
def zone_based_csv():
    """Create a zone-based CSV fixture"""
    csv_content = """Order - Number,Date - Shipped Date,Ship To - Postal Code,Ship To - Country,Shipment - Weight (Oz),Carrier - Service Selected,Shipment - Zone,Length,Width,Height,Shipping Rate
HT001,2025-01-01,12345,United States,10,UPS Ground,5,12,10,8,5.50
HT002,2025-01-02,23456,CA,15,DHL SM Parcel Expedited,7,14,12,10,7.25"""
    return csv_content

@pytest.fixture
def zip_based_csv():
    """Create a zip-based CSV fixture (no zone column)"""
    csv_content = """Order - Number,Date - Shipped Date,Ship To - Postal Code,Ship To - Country,Shipment - Weight (Oz),Carrier - Service Selected,Length,Width,Height,Shipping Rate
HT001,2025-01-01,12345,US,10,UPS Ground,12,10,8,5.50
HT002,2025-01-02,23456,Mexico,15,DHL SM Parcel Expedited,14,12,10,7.25"""
    return csv_content

@pytest.fixture
def redo_carrier_csv():
    """CSV fixture with multiple redo carrier signals"""
    csv_content = """Order - Number,Date - Shipped Date,Ship To - Postal Code,Ship To - Country,Shipment - Weight (Oz),Carrier - Name,Carrier - Service Selected,Shipment - Zone,Length,Width,Height,Shipping Rate
HT001,2025-01-01,12345,United States,10,First Mile,First Mile 1-3 Days,5,12,10,8,5.50
HT002,2025-01-02,23456,CA,15,DHL,DHL Express,7,14,12,10,7.25
HT003,2025-01-03,34567,US,12,Amazon,Amazon Logistics,6,10,8,6,6.00
HT004,2025-01-04,45678,US,9,UPS,UPS Ground,4,11,9,7,4.75"""
    return csv_content

@pytest.fixture
def merchant_service_csv():
    """CSV fixture for merchant service level intersection tests"""
    csv_content = """Order - Number,Date - Shipped Date,Ship To - Postal Code,Ship To - Country,Shipment - Weight (Oz),Carrier - Service Selected,Shipment - Zone,Length,Width,Height,Shipping Rate
HT001,2025-01-01,12345,US,10,UPSÂ® Ground,3,12,10,8,5.50
HT002,2025-01-02,23456,US,15,USPS Ground Advantage,4,14,12,10,7.25
HT003,2025-01-03,34567,US,8,Some Unknown Service,5,10,8,6,4.25"""
    return csv_content

@pytest.fixture
def carrier_service_csv():
    """CSV fixture with UPS and FedEx services for qualification tests"""
    csv_content = """Order - Number,Date - Shipped Date,Ship To - Postal Code,Ship To - Country,Shipment - Weight (Oz),Carrier - Name,Carrier - Service Selected,Shipment - Zone,Length,Width,Height,Shipping Rate
HT001,2025-01-01,12345,US,10,UPS,UPS Ground,5,12,10,8,5.50
HT002,2025-01-02,23456,US,15,FedEx,FedEx Ground,6,14,12,10,7.25"""
    return csv_content

def test_structure_detection_zone_vs_zip(client, zone_based_csv, zip_based_csv):
    """Test that structure detection correctly identifies zone vs zip-based invoices"""
    # Test zone-based
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        zone_path = f.name
    
    try:
        structure = detect_structure(zone_path)
        assert structure == 'zone'
    finally:
        os.unlink(zone_path)
    
    # Test zip-based
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zip_based_csv)
        zip_path = f.name
    
    try:
        structure = detect_structure(zip_path)
        assert structure == 'zip'
    finally:
        os.unlink(zip_path)

def test_get_uploaded_file_accepts_all_field_names(client, zone_based_csv):
    """Test that upload endpoint accepts invoice, invoice_file, and invoice_csv field names"""
    # Test with 'invoice'
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name
    
    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            assert response.status_code == 200
            data1 = json.loads(response.data)
            assert 'job_id' in data1
            assert data1['detected_structure'] == 'zone'
    finally:
        os.unlink(csv_path)
    
    # Test with 'invoice_file'
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name
    
    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice_file': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            assert response.status_code == 200
            data2 = json.loads(response.data)
            assert 'job_id' in data2
    finally:
        os.unlink(csv_path)
    
    # Test with 'invoice_csv'
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name
    
    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice_csv': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            assert response.status_code == 200
            data3 = json.loads(response.data)
            assert 'job_id' in data3
    finally:
        os.unlink(csv_path)

def test_service_normalization_handles_special_chars():
    """Test that service normalization handles ® and Â characters"""
    # Test various formats
    test_cases = [
        ('UPS® Ground', 'UPS GROUND'),
        ('UPSÂ® Ground', 'UPS GROUND'),
        ('UPS Ground', 'UPS GROUND'),
        ('DHL SM Parcel Expedited', 'DHL SM PARCEL EXPEDITED'),
        ('USPS Ground Advantage', 'USPS GROUND ADVANTAGE'),
    ]
    
    for input_service, expected in test_cases:
        normalized = normalize_service_name(input_service)
        assert normalized == expected, f"Failed for {input_service}: got {normalized}, expected {expected}"

def test_normalized_computed_fields_present(client, zone_based_csv):
    """Test that normalization fills computed fields when source data exists"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name

    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']

        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Service Selected',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }

        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '94105',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200

        normalized_csv = Path(app.config['UPLOAD_FOLDER']) / job_id / 'normalized.csv'
        with open(normalized_csv, newline='') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert rows, "Normalized CSV should have at least one row"
        row = rows[0]

        required_cols = [
            'WEIGHT_IN_LBS',
            'TWO_LETTER_COUNTRY_CODE',
            'CALCULATED_TWO_LETTER_COUNTRY_CODE',
            'FULL_COUNTRY_NAME',
            'CLEANED_SHIPPING_SERVICE',
            'SHIPPING_PRIORITY',
            'PACKAGE_DIMENSION_VOLUME',
            'PACKAGE_SIZE_STATUS',
            'WEIGHT_CLASSIFICATION'
        ]
        for col in required_cols:
            assert row.get(col) not in (None, ''), f"{col} should be populated"
        assert row.get('ORIGIN_ZIP_CODE') in (None, ''), "ORIGIN_ZIP_CODE should be blank for zone runs"
    finally:
        os.unlink(csv_path)

def test_excel_generation_populates_fill_columns(client, zone_based_csv):
    """Test that fill columns are written into the Excel output"""
    import openpyxl

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name

    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']

        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Service Selected',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }

        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '94105',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200

        response = client.post(f'/api/merchant-pricing/{job_id}', json={
            'excluded_carriers': [],
            'included_services': ['UPS Ground']
        })
        assert response.status_code == 200

        response = client.post('/api/generate', json={'job_id': job_id})
        assert response.status_code == 200

        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
        assert rate_card_files, "Expected a generated rate card file"

        wb = openpyxl.load_workbook(rate_card_files[0])
        ws = wb['Raw Data']
        headers = [cell.value for cell in ws[1]]
        header_index = {h: i + 1 for i, h in enumerate(headers) if h}

        fill_headers = [
            'ORDER_NUMBER',
            'DATE',
            'DESTINATION_ZIP_CODE',
            'WEIGHT_IN_OZ',
            'WEIGHT_IN_LBS',
            'SHIPPING_CARRIER',
            'CLEANED_SHIPPING_SERVICE',
            'SHIPPING_SERVICE',
            'LABEL_COST',
            'ZONE',
            'QUALIFIED'
        ]
        for header in fill_headers:
            col_idx = header_index[header]
            value = ws.cell(2, col_idx).value
            assert value not in (None, ''), f"{header} should be populated in Excel output"

        blank_header = 'TWO_LETTER_COUNTRY_CODE'
        col_idx = header_index[blank_header]
        assert ws.cell(2, col_idx).value in (None, ''), f"{blank_header} should remain blank"

        wb.close()
    finally:
        os.unlink(csv_path)

def _find_redo_pricing_section(ws):
    title_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == 'Redo Carriers':
                title_cell = cell
                break
        if title_cell:
            break

    assert title_cell is not None, "Redo Carriers section not found"

    header_row_idx = title_cell.row
    use_col = None
    label_col = None
    for cell in ws[header_row_idx]:
        if cell.value and str(cell.value).strip() == 'Use in Pricing':
            use_col = cell.column
        if cell.value and str(cell.value).strip() == 'Redo Carriers':
            label_col = cell.column

    if use_col is None:
        for cell in ws[header_row_idx + 1]:
            if cell.value and str(cell.value).strip() == 'Use in Pricing':
                use_col = cell.column
        header_row_idx = header_row_idx + 1
        if label_col is None:
            for cell in ws[header_row_idx]:
                if cell.value and 'CARRIER' in str(cell.value).upper():
                    label_col = cell.column
                    break

    if label_col is None:
        label_col = title_cell.column

    assert use_col is not None, "Use in Pricing column not found"
    return header_row_idx + 1, label_col, use_col

def _find_pricing_section_values(ws, section_title, stop_titles):
    header_row_idx = None
    label_col = None
    use_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == section_title:
                header_row_idx = cell.row
                label_col = cell.column
                break
        if header_row_idx:
            break

    assert header_row_idx is not None, f"{section_title} section not found"

    for cell in ws[header_row_idx]:
        if cell.value and str(cell.value).strip() == 'Use in Pricing':
            use_col = cell.column
            break
    if use_col is None:
        for cell in ws[header_row_idx + 1]:
            if cell.value and str(cell.value).strip() == 'Use in Pricing':
                use_col = cell.column
                header_row_idx = header_row_idx + 1
                break
    assert use_col is not None, "Use in Pricing column not found"

    values = {}
    row_idx = header_row_idx + 1
    while True:
        label_val = ws.cell(row_idx, label_col).value
        if not label_val:
            break
        normalized = normalize_redo_label(label_val)
        if normalized in stop_titles:
            break
        values[str(label_val).strip()] = ws.cell(row_idx, use_col).value
        row_idx += 1
    return values

def test_redo_carrier_use_in_pricing_written(client, zone_based_csv):
    """Test redo carrier selections written to Pricing & Summary."""
    import openpyxl

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name

    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']

        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Service Selected',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }

        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '94105',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200

        response = client.post(f'/api/merchant-pricing/{job_id}', json={
            'excluded_carriers': [],
            'included_services': ['UPS Ground']
        })
        assert response.status_code == 200

        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        with open(job_dir / 'redo_carriers.json', 'w') as f:
            json.dump({'selected_carriers': ['DHL', 'Amazon']}, f)

        response = client.post('/api/generate', json={'job_id': job_id})
        assert response.status_code == 200

        rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
        assert rate_card_files, "Expected a generated rate card file"

        wb = openpyxl.load_workbook(rate_card_files[0])
        ws = wb['Pricing & Summary']
        start_row, label_col, use_col = _find_redo_pricing_section(ws)

        values = {}
        row_idx = start_row
        while True:
            label_val = ws.cell(row_idx, label_col).value
            if not label_val:
                break
            normalized = normalize_redo_label(label_val)
            if normalized in {'MERCHANT CARRIERS', 'MERCHANT CARRIER', 'MERCHANT SERVICE LEVELS'}:
                break
            label = str(label_val).strip()
            values[label] = ws.cell(row_idx, use_col).value
            row_idx += 1

        def find_value(prefix):
            for key, val in values.items():
                if key.startswith(prefix):
                    return val
            return None

        assert find_value('USPS Market') == 'No'
        assert find_value('UPS Ground Saver') == 'No'
        assert find_value('UPS Ground') == 'No'
        assert find_value('DHL') == 'Yes'
        assert find_value('Amazon') == 'Yes'
        assert find_value('FedEx') == 'No'

        wb.close()
    finally:
        os.unlink(csv_path)

def test_pricing_summary_overwrite_use_in_pricing(client, zone_based_csv):
    """Test deterministic Yes/No overwrite for redo and merchant sections."""
    import openpyxl

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name

    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']

        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Service Selected',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }

        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '94105',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200

        response = client.post(f'/api/merchant-pricing/{job_id}', json={
            'excluded_carriers': ['FedEx', 'Amazon'],
            'included_services': ['UPSÂ® Ground']
        })
        assert response.status_code == 200

        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        with open(job_dir / 'redo_carriers.json', 'w') as f:
            json.dump({'selected_carriers': ['DHL']}, f)

        response = client.post('/api/generate', json={'job_id': job_id})
        assert response.status_code == 200

        rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
        assert rate_card_files, "Expected a generated rate card file"

        wb = openpyxl.load_workbook(rate_card_files[0])
        ws = wb['Pricing & Summary']

        start_row, label_col, use_col = _find_redo_pricing_section(ws)
        redo_values = {}
        row_idx = start_row
        while True:
            label_val = ws.cell(row_idx, label_col).value
            if not label_val:
                break
            normalized = normalize_redo_label(label_val)
            if normalized in {'MERCHANT CARRIERS', 'MERCHANT CARRIER', 'MERCHANT SERVICE LEVELS'}:
                break
            redo_values[str(label_val).strip()] = ws.cell(row_idx, use_col).value
            row_idx += 1

        def find_redo(prefix):
            for key, val in redo_values.items():
                if key.startswith(prefix):
                    return val
            return None

        assert find_redo('USPS Market') == 'No'
        assert find_redo('UPS Ground Saver') == 'No'
        assert find_redo('UPS Ground') == 'No'
        assert find_redo('DHL') == 'Yes'
        assert find_redo('FedEx') == 'No'
        assert find_redo('Amazon') == 'No'
        assert find_redo('First Mile') == 'No'

        header_row_idx, label_col, use_col = None, None, None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip() == 'Merchant Service Levels':
                    header_row_idx = cell.row
                    label_col = cell.column
                    break
            if header_row_idx:
                break
        assert header_row_idx is not None
        for cell in ws[header_row_idx]:
            if cell.value and str(cell.value).strip() == 'Use in Pricing':
                use_col = cell.column
                break
        if use_col is None:
            for cell in ws[header_row_idx + 1]:
                if cell.value and str(cell.value).strip() == 'Use in Pricing':
                    use_col = cell.column
            header_row_idx = header_row_idx + 1
        assert use_col is not None

        first_label = ws.cell(header_row_idx + 1, label_col).value
        if isinstance(first_label, str) and first_label.startswith('='):
            for idx in range(3):
                row_idx = header_row_idx + 1 + idx
                val = ws.cell(row_idx, use_col).value
                assert val in ('Yes', 'No', None)
        else:
            service_values = {}
            row_idx = header_row_idx + 1
            while True:
                label_val = ws.cell(row_idx, label_col).value
                if not label_val:
                    break
                normalized = normalize_redo_label(label_val)
                if normalized in {'REDO CARRIERS'}:
                    break
                service_values[str(label_val).strip()] = ws.cell(row_idx, use_col).value
                row_idx += 1

            def find_service(prefix):
                for key, val in service_values.items():
                    if key.startswith(prefix):
                        return val
                return None

            normalized_labels = {normalize_service_name(label) for label in service_values}
            expected_labels = {normalize_service_name(s) for s in [
                'UPSÂ® Ground',
                'DHL Parcel International Direct - DDU',
                'DHL SM Parcel Expedited',
                'USPS Ground Advantage',
                'UPS 2nd Day AirÂ®',
                'DHL SM Parcel Expedited Max'
            ]}
            assert expected_labels.issubset(normalized_labels)

            assert find_service('UPS') == 'Yes'
            assert find_service('DHL Parcel International Direct') == 'No'
            assert find_service('DHL SM Parcel Expedited') == 'No'
            assert find_service('USPS Ground Advantage') == 'No'
            assert find_service('UPS 2nd Day Air') == 'No'
            assert find_service('DHL SM Parcel Expedited Max') == 'No'

        header_row_idx, label_col, use_col = None, None, None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip() == 'Merchant Carriers':
                    header_row_idx = cell.row
                    label_col = cell.column
                    break
            if header_row_idx:
                break
        assert header_row_idx is not None
        for cell in ws[header_row_idx]:
            if cell.value and str(cell.value).strip() == 'Use in Pricing':
                use_col = cell.column
                break
        if use_col is None:
            for cell in ws[header_row_idx + 1]:
                if cell.value and str(cell.value).strip() == 'Use in Pricing':
                    use_col = cell.column
            header_row_idx = header_row_idx + 1
        assert use_col is not None

        carrier_values = {}
        row_idx = header_row_idx + 1
        while True:
            label_val = ws.cell(row_idx, label_col).value
            if not label_val:
                break
            normalized = normalize_redo_label(label_val)
            if normalized in {'MERCHANT SERVICE LEVELS', 'REDO CARRIERS'}:
                break
            carrier_values[str(label_val).strip()] = ws.cell(row_idx, use_col).value
            row_idx += 1

        def find_carrier(prefix):
            for key, val in carrier_values.items():
                if key.startswith(prefix):
                    return val
            return None

        assert find_carrier('USPS') == 'Yes'
        assert find_carrier('UPS') == 'Yes'
        assert find_carrier('DHL') == 'Yes'
        assert find_carrier('FedEx') == 'No'
        assert find_carrier('Amazon') == 'No'
        assert find_carrier('First Mile') == 'No'
        wb.close()
    finally:
        os.unlink(csv_path)

def test_redo_carrier_detection_defaults(client, redo_carrier_csv):
    """Test redo carrier defaults"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(redo_carrier_csv)
        csv_path = f.name

    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']

        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Name',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }

        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '94105',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200

        response = client.get(f'/api/redo-carriers/{job_id}')
        assert response.status_code == 200
        data = json.loads(response.data)

        detected = set(data['detected_carriers'])
        expected = {
            "USPS Market",
            "UPS Ground",
            "UPS Ground Saver",
            "FedEx",
            "Amazon",
            "DHL"
        }
        assert detected == expected

        default_selected = set(data['default_selected'])
        assert default_selected == {
            "USPS Market",
            "UPS Ground",
            "UPS Ground Saver"
        }
    finally:
        os.unlink(csv_path)

def test_merchant_service_levels_are_filtered_to_invoice_intersection(client, merchant_service_csv):
    """Test merchant service levels filtered to invoice intersection"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(merchant_service_csv)
        csv_path = f.name

    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']

        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Service Selected',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }

        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '94105',
            'mapping': mapping,
            'structure': 'zip'
        })
        assert response.status_code == 200

        response = client.get(f'/api/service-levels/{job_id}')
        assert response.status_code == 200
        data = json.loads(response.data)

        available = set(data['available_services'])
        assert 'UPSÂ® Ground' in available
        assert 'USPS Ground Advantage' in available
        assert 'DHL SM Parcel Expedited' not in available
    finally:
        os.unlink(csv_path)

def test_excel_generation_preserves_formulas(client, zone_based_csv):
    """Test that Excel generation preserves formula columns"""
    import openpyxl
    
    # Upload file
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name
    
    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']
        
        # Create mapping
        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Service Selected',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }
        
        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200
        
        response = client.post(f'/api/merchant-pricing/{job_id}', json={
            'excluded_carriers': [],
            'included_services': ['UPS Ground', 'DHL SM Parcel Expedited']
        })
        assert response.status_code == 200
        
        # Generate
        response = client.post('/api/generate', json={'job_id': job_id})
        assert response.status_code == 200
        
        # Check generated file
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
        assert len(rate_card_files) > 0
        
        # Load and check formulas
        wb = openpyxl.load_workbook(rate_card_files[0])
        ws = wb['Raw Data']
        
        # Check that formula columns (AI-AN) still have formulas
        # Note: formulas might be in row 2 (template) or copied to data rows
        formula_cols = [35, 36, 37, 38, 39, 40]
        for col_idx in formula_cols:
            # Check row 2 (template row)
            cell = ws.cell(2, col_idx)
            if cell.value:
                assert str(cell.value).startswith('='), f"Column {col_idx} should contain a formula"
    
    finally:
        os.unlink(csv_path)

def test_qualified_written_true_false_based_on_selected_services(client, zone_based_csv):
    """Test that QUALIFIED column is written correctly based on selected services"""
    import openpyxl
    
    # Upload file
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(zone_based_csv)
        csv_path = f.name
    
    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']
        
        # Create mapping
        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Service Selected',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }
        
        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200
        
        response = client.post(f'/api/merchant-pricing/{job_id}', json={
            'excluded_carriers': [],
            'included_services': ['UPS Ground']
        })
        assert response.status_code == 200
        
        # Generate
        response = client.post('/api/generate', json={'job_id': job_id})
        assert response.status_code == 200
        
        # Check generated file
        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id
        rate_card_files = list(job_dir.glob('* - Rate Card.xlsx'))
        assert len(rate_card_files) > 0
        
        # Load and check QUALIFIED column
        wb = openpyxl.load_workbook(rate_card_files[0])
        ws = wb['Raw Data']
        
        # Find QUALIFIED column
        headers = [cell.value for cell in ws[1]]
        qualified_col = headers.index('QUALIFIED') + 1
        
        # First row has "UPS Ground" - should be True
        row2_qualified = ws.cell(2, qualified_col).value
        assert row2_qualified == True, "First row should be qualified (UPS Ground selected)"
        
        # Second row has "DHL SM Parcel Expedited" - should be False
        row3_qualified = ws.cell(3, qualified_col).value
        assert row3_qualified == False, "Second row should not be qualified (DHL not selected)"
    
    finally:
        os.unlink(csv_path)

def test_merchant_pricing_controls_qualified_and_carriers(client, carrier_service_csv):
    """Test merchant pricing selections affect qualification and carrier pricing flags."""
    import openpyxl

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(carrier_service_csv)
        csv_path = f.name

    try:
        with open(csv_path, 'rb') as f:
            data = {'invoice': (f, 'test.csv')}
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')
            upload_data = json.loads(response.data)
            job_id = upload_data['job_id']

        mapping = {
            'Order Number': 'Order - Number',
            'Order Date': 'Date - Shipped Date',
            'Zip': 'Ship To - Postal Code',
            'Weight (oz)': 'Shipment - Weight (Oz)',
            'Shipping Carrier': 'Carrier - Name',
            'Shipping Service': 'Carrier - Service Selected',
            'Package Height': 'Height',
            'Package Width': 'Width',
            'Package Length': 'Length',
            'Zone': 'Shipment - Zone',
            'Label Cost': 'Shipping Rate'
        }

        response = client.post('/api/mapping', json={
            'job_id': job_id,
            'merchant_name': 'Test Merchant',
            'merchant_id': '',
            'existing_customer': False,
            'origin_zip': '',
            'mapping': mapping,
            'structure': 'zone'
        })
        assert response.status_code == 200

        job_dir = Path(app.config['UPLOAD_FOLDER']) / job_id

        def generate_with_pricing(included_services, excluded_carriers):
            response = client.post(f'/api/merchant-pricing/{job_id}', json={
                'excluded_carriers': excluded_carriers,
                'included_services': included_services
            })
            assert response.status_code == 200
            response = client.post('/api/generate', json={'job_id': job_id})
            assert response.status_code == 200

            deadline = time.time() + 5
            rate_card = None
            while time.time() < deadline:
                files = list(job_dir.glob('* - Rate Card.xlsx'))
                if files:
                    rate_card = files[0]
                    break
                time.sleep(0.1)
            assert rate_card is not None, "Expected a generated rate card file"
            return openpyxl.load_workbook(rate_card)

        wb = generate_with_pricing(['UPS Ground'], [])
        ws = wb['Raw Data']
        headers = [cell.value for cell in ws[1]]
        qualified_col = headers.index('QUALIFIED') + 1
        assert ws.cell(2, qualified_col).value is True
        assert ws.cell(3, qualified_col).value is False
        wb.close()

        wb = generate_with_pricing(['FedEx Ground'], [])
        ws = wb['Raw Data']
        headers = [cell.value for cell in ws[1]]
        qualified_col = headers.index('QUALIFIED') + 1
        assert ws.cell(2, qualified_col).value is False
        assert ws.cell(3, qualified_col).value is True
        wb.close()

        wb = generate_with_pricing(['FedEx Ground'], ['FedEx'])
        summary_ws = wb['Pricing & Summary']
        values = _find_pricing_section_values(
            summary_ws,
            'Merchant Carriers',
            {'MERCHANT SERVICE LEVELS', 'REDO CARRIERS'}
        )
        fedex_value = None
        for key, val in values.items():
            if key.startswith('FedEx'):
                fedex_value = val
                break
        assert fedex_value == 'No'
        wb.close()
    finally:
        os.unlink(csv_path)

if __name__ == '__main__':
    pytest.main([__file__, '-v'])
